from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden, HttpResponseRedirect, HttpResponse, Http404
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q, Avg
from django.db.models.functions import TruncMonth
from django.db import models
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.mail import send_mail, get_connection
from django.conf import settings as django_settings
from decimal import Decimal
from datetime import timedelta, date
from io import BytesIO
import re
import json
import csv

from xhtml2pdf import pisa

from .models import (
    UserModel, Categoria, Equipo, Jugador,
    Entrenamiento, Torneo, Partido, Finanza,
    Evaluacion, Pago, ConceptoPago, Notificacion, Asistencia,
    PartidoCalendario, Convocatoria, EstadisticaPartido, ObservacionDeportista,
    RespuestaConvocatoria, PagoEntrenador
)


def _notificar_partido(partido, es_nuevo, emisor):
    """
    Crea notificacion para deportistas del equipo cuando se crea/modifica un partido.
    """
    accion  = 'programado' if es_nuevo else 'modificado'
    asunto  = f'Partido {accion}: {partido.equipo_propio.nombre} vs {partido.equipo_rival}'
    mensaje = (
        f'Se ha {accion} un partido.\n\n'
        f'Equipo:  {partido.equipo_propio.nombre}\n'
        f'Rival:   {partido.equipo_rival}\n'
        f'Fecha:   {partido.fecha.strftime("%d/%m/%Y")}\n'
        f'Hora:    {partido.hora.strftime("%H:%M")}\n'
        f'Lugar:   {partido.cancha or "Por confirmar"}'
    )
    jugadores = Jugador.objects.filter(
        equipo=partido.equipo_propio, activo=True
    ).select_related('usuario')
    notifs = [
        Notificacion(usuario=j.usuario, asunto=asunto, mensaje=mensaje, emisor=emisor)
        for j in jugadores
    ]
    Notificacion.objects.bulk_create(notifs)
    for j in jugadores:
        if j.usuario.email and '@' in j.usuario.email:
            try:
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[j.usuario.email],
                    fail_silently=True,
                )
            except Exception:
                pass


def _notificar_torneo(torneo, es_nuevo, emisor):
    """
    Notifica a todos los deportistas activos cuando se crea/modifica un torneo.
    """
    accion  = 'creado' if es_nuevo else 'modificado'
    asunto  = f'Torneo {accion}: {torneo.nombre}'
    mensaje = (
        f'Se ha {accion} un torneo.\n\n'
        f'Nombre:     {torneo.nombre}\n'
        f'Categoría:  {torneo.categoria.nombre}\n'
        f'Inicio:     {torneo.fecha_inicio.strftime("%d/%m/%Y")}\n'
        f'Fin:        {torneo.fecha_fin.strftime("%d/%m/%Y")}\n'
        f'Estado:     {torneo.get_estado_display()}\n'
        f'Lugar:      {torneo.lugar or "Por confirmar"}'
    )
    jugadores = Jugador.objects.filter(
        equipo__categoria=torneo.categoria, activo=True
    ).select_related('usuario')
    notifs = [
        Notificacion(usuario=j.usuario, asunto=asunto, mensaje=mensaje, emisor=emisor)
        for j in jugadores
    ]
    Notificacion.objects.bulk_create(notifs)
    for j in jugadores:
        if j.usuario.email and '@' in j.usuario.email:
            try:
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[j.usuario.email],
                    fail_silently=True,
                )
            except Exception:
                pass


def _notificar_usuario_cambio(usuario, es_nuevo, emisor):
    """
    Notifica al usuario cuando su cuenta es creada o editada.
    """
    accion  = 'creada' if es_nuevo else 'actualizada'
    asunto  = f'Tu cuenta ha sido {accion}'
    mensaje = (
        f'Hola {usuario.get_full_name() or usuario.username},\n\n'
        f'Tu cuenta ha sido {accion} en el sistema.\n\n'
        f'Usuario:   {usuario.username}\n'
        f'Rol:       {usuario.get_role_display()}\n'
        f'Estado:    {"Activo" if usuario.is_active else "Inactivo"}'
    )
    Notificacion.objects.create(usuario=usuario, asunto=asunto, mensaje=mensaje, emisor=emisor)
    if usuario.email and '@' in usuario.email:
        try:
            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=django_settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                fail_silently=True,
            )
        except Exception:
            pass


def _notificar_entrenamiento(entrenamiento, es_nuevo, emisor):
    """
    Crea una Notificacion para cada deportista activo del equipo
    e intenta enviar correo cuando se crea o modifica un entrenamiento.
    """
    accion  = 'programado' if es_nuevo else 'modificado'
    asunto  = f'Entrenamiento {accion}: {entrenamiento.titulo}'
    mensaje = (
        f'Se ha {accion} un entrenamiento.\n\n'
        f'Título:     {entrenamiento.titulo}\n'
        f'Fecha:      {entrenamiento.fecha.strftime("%d/%m/%Y")}\n'
        f'Horario:    {entrenamiento.hora_inicio.strftime("%H:%M")} – {entrenamiento.hora_fin.strftime("%H:%M")}\n'
        f'Lugar:      {entrenamiento.lugar or "Por confirmar"}\n'
        f'Equipo:     {entrenamiento.equipo.nombre}'
    )
    jugadores = Jugador.objects.filter(
        equipo=entrenamiento.equipo, activo=True
    ).select_related('usuario')
    notifs = [
        Notificacion(usuario=j.usuario, asunto=asunto, mensaje=mensaje, emisor=emisor)
        for j in jugadores
    ]
    Notificacion.objects.bulk_create(notifs)
    for j in jugadores:
        if j.usuario.email and '@' in j.usuario.email:
            try:
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[j.usuario.email],
                    fail_silently=True,
                )
            except Exception:
                pass


def _notificar_pago(pago, accion, emisor):
    """
    Envia notificacion al deportista sobre un pago.
    accion: 'registrado', 'actualizado', 'pagado', 'pendiente'
    """
    etiquetas = {
        'registrado': 'Nuevo pago registrado',
        'actualizado': 'Pago actualizado',
        'pagado':      'Pago confirmado',
        'pendiente':   'Pago revertido a pendiente',
    }
    asunto  = f'{etiquetas.get(accion, "Notificacion de pago")}: {pago.descripcion}'
    estado_display = dict(pago.ESTADO_CHOICES).get(pago.estado, pago.estado)
    fv = pago.fecha_vencimiento
    if isinstance(fv, str):
        fv = date.fromisoformat(fv)
    fp = pago.fecha_pago
    if isinstance(fp, str):
        fp = date.fromisoformat(fp)
    mensaje = (
        f'Informacion sobre tu pago:\n\n'
        f'Concepto:     {pago.get_concepto_display()}\n'
        f'Descripcion:  {pago.descripcion}\n'
        f'Monto:        {pago.monto}\n'
        f'Vencimiento:  {fv.strftime("%d/%m/%Y")}\n'
        f'Estado:       {estado_display}'
        + (f'\nFecha pago:   {fp.strftime("%d/%m/%Y")}' if fp else '')
    )
    usuario = pago.jugador.usuario
    Notificacion.objects.create(usuario=usuario, asunto=asunto, mensaje=mensaje, emisor=emisor)
    if usuario.email and '@' in usuario.email:
        try:
            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=django_settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                fail_silently=True,
            )
        except Exception:
            pass


def solo_admin(view_func):
    """Decorador que verifica que el usuario sea administrador"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, 'auth/login.html')
        if request.user.role != 'administrador':
            return HttpResponseForbidden("Sin permisos de administrador.")
        return view_func(request, *args, **kwargs)
    return wrapper

def solo_entrenador(view_func):
    """Decorador que verifica que el usuario sea entrenador"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, 'auth/login.html')
        if request.user.role != 'entrenador':
            return HttpResponseForbidden("Sin permisos de entrenador.")
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────
# DASHBOARD PRINCIPAL
# ─────────────────────────────────────────────

@login_required
def admin_dashboard(request):
    """Dashboard principal con estadísticas generales"""
    if request.user.role != 'administrador':
        return HttpResponseForbidden("Sin permisos.")

    # Estadísticas de usuarios
    total_users = UserModel.objects.count()
    total_admins = UserModel.objects.filter(role='administrador').count()
    total_entrenadores = UserModel.objects.filter(role='entrenador').count()
    total_deportistas = UserModel.objects.filter(role='deportista').count()

    # Estadísticas de entidades
    total_categorias = Categoria.objects.filter(activo=True).count()
    total_equipos = Equipo.objects.filter(activo=True).count()
    total_entrenamientos = Entrenamiento.objects.count()
    total_torneos = Torneo.objects.count()
    torneos_activos = Torneo.objects.filter(estado='en_curso').count()

    # Estadísticas financieras
    total_ingresos = Finanza.objects.filter(tipo='ingreso').aggregate(
        total=Sum('monto'))['total'] or Decimal('0')
    total_egresos = Finanza.objects.filter(tipo='egreso').aggregate(
        total=Sum('monto'))['total'] or Decimal('0')
    balance = total_ingresos - total_egresos

    # Últimos registros
    ultimos_usuarios = UserModel.objects.order_by('-date_joined')[:5]
    ultimos_entrenamientos = Entrenamiento.objects.order_by('-fecha')[:5]
    ultimas_finanzas = Finanza.objects.order_by('-fecha')[:5]

    context = {
        'total_users': total_users,
        'total_admins': total_admins,
        'total_entrenadores': total_entrenadores,
        'total_deportistas': total_deportistas,
        'total_categorias': total_categorias,
        'total_equipos': total_equipos,
        'total_entrenamientos': total_entrenamientos,
        'total_torneos': total_torneos,
        'torneos_activos': torneos_activos,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'balance': balance,
        'ultimos_usuarios': ultimos_usuarios,
        'ultimos_entrenamientos': ultimos_entrenamientos,
        'ultimas_finanzas': ultimas_finanzas,
    }
    return render(request, 'panel/dashboard.html', context)


# ─────────────────────────────────────────────
# GESTIÓN DE USUARIOS
# ─────────────────────────────────────────────

@solo_admin
def usuarios_lista(request):
    """Lista de usuarios con filtros multi criterio"""
    qs = UserModel.objects.all()

    buscar  = request.GET.get('buscar', '').strip()
    rol     = request.GET.get('rol', '')
    estado  = request.GET.get('estado', '')

    if buscar:
        qs = qs.filter(
            models.Q(username__icontains=buscar) |
            models.Q(first_name__icontains=buscar) |
            models.Q(last_name__icontains=buscar) |
            models.Q(email__icontains=buscar)     |
            models.Q(documento__icontains=buscar)
        )
    if rol:
        qs = qs.filter(role=rol)
    if estado == 'activo':
        qs = qs.filter(is_active=True)
    elif estado == 'inactivo':
        qs = qs.filter(is_active=False)

    qs = qs.order_by('-date_joined')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/usuarios/lista.html', {
        'usuarios': page,
        'buscar': buscar,
        'rol': rol,
        'estado': estado,
        'roles': UserModel.ROLE_CHOICES,
        'total_filtrados': qs.count(),
    })


def _validar_usuario_form(request, usuario_id=None):
    """
    Valida y extrae los campos del formulario de usuario.
    Retorna (errores: list, datos: dict).
    Si errores está vacío, los datos son válidos.
    """
    birth_date_raw = request.POST.get('birth_date', '').strip()
    birth_date = None
    if birth_date_raw:
        try:
            from datetime import date as _date
            birth_date = _date.fromisoformat(birth_date_raw)
        except ValueError:
            birth_date = None

    datos = {
        'email':      request.POST.get('email', '').strip(),
        'first_name': request.POST.get('first_name', '').strip(),
        'last_name':  request.POST.get('last_name', '').strip(),
        'role':       request.POST.get('role', 'deportista').strip(),
        'phone':      request.POST.get('phone', '').strip(),
        'documento':  request.POST.get('documento', '').strip(),
        'password':   request.POST.get('password', '').strip(),
        'is_active':  request.POST.get('is_active') == 'on',
        'equipo_id':  request.POST.get('equipo') or None,
        'posicion':   request.POST.get('posicion', '').strip(),
        'jugador_id': request.POST.get('jugador_id') or None,
        'parentesco': request.POST.get('parentesco', '').strip(),
        'birth_date': birth_date,
    }
    errores = []

    # ── Email obligatorio y formato válido ───────────────────────────────────
    if not datos['email']:
        errores.append('El email es obligatorio.')
    elif not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', datos['email']):
        errores.append('El formato del email no es válido.')
    else:
        qs_email = UserModel.objects.filter(email__iexact=datos['email'])
        if usuario_id:
            qs_email = qs_email.exclude(id=usuario_id)
        if qs_email.exists():
            errores.append(f'El email "{datos["email"]}" ya está registrado.')

    # ── Documento: solo números, obligatorio y único ─────────────────────────
    if not datos['documento']:
        errores.append('El documento es obligatorio.')
    elif not re.match(r'^\d+$', datos['documento']):
        errores.append('El documento debe contener solo números.')
    else:
        qs_doc = UserModel.objects.filter(documento=datos['documento'])
        if usuario_id:
            qs_doc = qs_doc.exclude(id=usuario_id)
        if qs_doc.exists():
            errores.append(f'El documento "{datos["documento"]}" ya está registrado.')

    # ── Teléfono: solo números si se ingresa ─────────────────────────────────
    if datos['phone']:
        if not re.match(r'^\d{7,15}$', datos['phone']):
            errores.append('El teléfono debe contener solo números (7-15 dígitos).')

    # ── Contraseña obligatoria al crear ──────────────────────────────────────
    if not usuario_id and not datos['password']:
        errores.append('La contraseña es obligatoria al crear un usuario.')

    # ── Reglas por rol ───────────────────────────────────────────────────────
    if datos['role'] == 'deportista':
        # Deportista requiere equipo y posición
        if not datos['equipo_id']:
            errores.append('El equipo es obligatorio para deportistas.')
        if not datos['posicion']:
            errores.append('La posición es obligatoria para deportistas.')

    return errores, datos


def _context_form_usuario(usuario=None, post_data=None):
    """
    Construye el contexto base para el formulario de usuario.
    Incluye equipos, categorías y datos del jugador si aplica.
    """
    jugador = None
    if usuario:
        try:
            jugador = usuario.jugador
        except Exception:
            pass

    acudiente = None
    if usuario:
        try:
            acudiente = usuario.acudiente
        except Exception:
            pass

    ctx = {
        'roles':      UserModel.ROLE_CHOICES,
        'equipos':    Equipo.objects.filter(activo=True).select_related('categoria'),
        'categorias': Categoria.objects.filter(activo=True),
        'jugador':    jugador,
        'acudiente':  acudiente,
        'jugadores_disponibles': Jugador.objects.filter(activo=True).select_related('usuario', 'equipo'),
        'es_mayor':   _deportista_es_mayor_de_edad(usuario) if usuario else None,
    }
    if usuario:
        ctx['usuario'] = usuario
    if post_data:
        ctx['post_data'] = post_data
    return ctx


@solo_admin
def usuario_crear(request):
    """
    Crear nuevo usuario con validaciones completas.
    - Username generado automáticamente según rol.
    - Si el rol es deportista, crea o actualiza el perfil Jugador.
    - Valida email único, documento único, teléfono y campos por rol.
    """
    if request.method == 'POST':
        errores, datos = _validar_usuario_form(request)

        if errores:
            for e in errores:
                messages.error(request, e)
            ctx = _context_form_usuario(post_data=request.POST)
            ctx['titulo'] = 'Crear Usuario'
            return render(request, 'panel/usuarios/form.html', ctx)

        # Generar username único según rol
        prefijos = {'administrador': 'admin', 'entrenador': 'entrenador', 'deportista': 'deportista'}
        prefijo  = prefijos.get(datos['role'], 'usuario')
        count    = UserModel.objects.filter(role=datos['role']).count() + 1
        username = f"{prefijo}{count:03d}"
        while UserModel.objects.filter(username=username).exists():
            count += 1
            username = f"{prefijo}{count:03d}"

        # Crear usuario
        usuario = UserModel.objects.create_user(
            username=username,
            email=datos['email'],
            password=datos['password'],
            first_name=datos['first_name'],
            last_name=datos['last_name'],
            role=datos['role'],
            phone=datos['phone'],
            documento=datos['documento'],
            is_active=datos['is_active'],
            birth_date=datos['birth_date'],
        )

        # Si es deportista, crear/actualizar perfil Jugador
        if datos['role'] == 'deportista' and datos['equipo_id']:
            Jugador.objects.update_or_create(
                usuario=usuario,
                defaults={
                    'equipo_id': datos['equipo_id'],
                    'posicion':  datos['posicion'],
                    'activo':    datos['is_active'],
                }
            )

        _notificar_usuario_cambio(usuario, es_nuevo=True, emisor=request.user)
        messages.success(request, f'Usuario creado correctamente. Username asignado: {username}')
        return usuarios_lista(request)

    ctx = _context_form_usuario()
    ctx['titulo'] = 'Crear Usuario'
    return render(request, 'panel/usuarios/form.html', ctx)


@solo_admin
def usuario_editar(request, user_id):
    """
    Editar usuario existente con validaciones completas.
    - Contraseña opcional: si está vacía no se actualiza.
    - Si el rol es deportista, sincroniza el perfil Jugador.
    - Si cambia de rol desde deportista, desvincula el perfil Jugador.
    """
    usuario = get_object_or_404(UserModel, id=user_id)

    if request.method == 'POST':
        errores, datos = _validar_usuario_form(request, usuario_id=user_id)

        if errores:
            for e in errores:
                messages.error(request, e)
            ctx = _context_form_usuario(usuario=usuario, post_data=request.POST)
            ctx['titulo'] = 'Editar Usuario'
            return render(request, 'panel/usuarios/form.html', ctx)

        # Actualizar campos del usuario
        usuario.first_name = datos['first_name']
        usuario.last_name  = datos['last_name']
        usuario.email      = datos['email']
        usuario.role       = datos['role']
        usuario.phone      = datos['phone']
        usuario.documento  = datos['documento']
        usuario.is_active  = datos['is_active']
        if datos['birth_date'] is not None:
            usuario.birth_date = datos['birth_date']

        # Contraseña: solo actualizar si se ingresó una nueva
        if datos['password']:
            usuario.set_password(datos['password'])

        usuario.save()

        # Sincronizar perfil Jugador según el nuevo rol
        if datos['role'] == 'deportista' and datos['equipo_id']:
            # Crear o actualizar el perfil Jugador
            Jugador.objects.update_or_create(
                usuario=usuario,
                defaults={
                    'equipo_id': datos['equipo_id'],
                    'posicion':  datos['posicion'],
                    'activo':    datos['is_active'],
                }
            )
        elif datos['role'] != 'deportista':
            # Si ya no es deportista, desvincular el perfil (sin eliminar)
            Jugador.objects.filter(usuario=usuario).update(equipo=None, activo=False)

        # Sincronizar perfil Acudiente
        if datos['role'] == 'acudiente' and datos.get('jugador_id'):
            from .models import Acudiente
            jugador_obj = Jugador.objects.filter(id=datos['jugador_id']).first()
            if jugador_obj:
                Acudiente.objects.update_or_create(
                    usuario=usuario,
                    defaults={'jugador': jugador_obj, 'parentesco': datos.get('parentesco', '')},
                )
        elif datos['role'] != 'acudiente':
            from .models import Acudiente
            Acudiente.objects.filter(usuario=usuario).delete()

        messages.success(request, f'Usuario "{usuario.username}" actualizado correctamente.')
        _notificar_usuario_cambio(usuario, es_nuevo=False, emisor=request.user)
        return usuarios_lista(request)

    ctx = _context_form_usuario(usuario=usuario)
    ctx['titulo'] = 'Editar Usuario'
    return render(request, 'panel/usuarios/form.html', ctx)


@solo_admin
def usuario_detalle(request, user_id):
    """Detalle de usuario con su perfil de jugador si aplica"""
    usuario = get_object_or_404(UserModel, id=user_id)
    jugador = None
    try:
        jugador = usuario.jugador
    except Exception:
        pass
    return render(request, 'panel/usuarios/form.html', {
        'titulo': 'Detalle de Usuario',
        'usuario': usuario,
        'jugador': jugador,
        'roles': UserModel.ROLE_CHOICES,
        'solo_lectura': True,
    })


@solo_admin
def usuario_eliminar(request, user_id):
    """Eliminar usuario"""
    usuario = get_object_or_404(UserModel, id=user_id)
    if usuario.id == request.user.id:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
    else:
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{nombre}" eliminado correctamente.')
    usuarios = Paginator(UserModel.objects.all().order_by('-date_joined'), 10).get_page(1)
    return render(request, 'panel/usuarios/lista.html', {'usuarios': usuarios})


# ─────────────────────────────────────────────
# GESTIÓN DE CATEGORÍAS Y EQUIPOS
# ─────────────────────────────────────────────

@solo_admin
def categorias_lista(request):
    """Lista de categorías con filtros por nombre y rango de edad"""
    qs = Categoria.objects.annotate(num_equipos=Count('equipos'))

    buscar       = request.GET.get('buscar', '').strip()
    subcategoria = request.GET.get('subcategoria', '').strip()

    if buscar:
        qs = qs.filter(nombre__icontains=buscar)
    if subcategoria:
        qs = qs.filter(subcategoria=subcategoria)

    return render(request, 'panel/categorias/lista.html', {
        'categorias':    qs.order_by('nombre'),
        'buscar':        buscar,
        'subcategoria':  subcategoria,
        'subcategorias': Categoria.SUBCATEGORIA_CHOICES,
    })


@solo_admin
def categoria_crear(request):
    """Crear categoría"""
    if request.method == 'POST':
        nombre       = request.POST.get('nombre', '').strip()
        descripcion  = request.POST.get('descripcion', '').strip()
        subcategoria = request.POST.get('subcategoria', '').strip()

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        elif Categoria.objects.filter(nombre=nombre).exists():
            messages.error(request, f'La categoría "{nombre}" ya existe.')
        else:
            edad_minima, edad_maxima = Categoria.EDAD_RANGOS.get(subcategoria, (None, None))
            Categoria.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                subcategoria=subcategoria,
                edad_minima=edad_minima,
                edad_maxima=edad_maxima,
            )
            messages.success(request, f'Categoría "{nombre}" creada correctamente.')
    return render(request, 'panel/categorias/form.html', {
        'titulo': 'Crear Categoría',
        'subcategorias': Categoria.SUBCATEGORIA_CHOICES,
    })


@solo_admin
def categoria_editar(request, cat_id):
    """Editar categoría"""
    categoria = get_object_or_404(Categoria, id=cat_id)
    if request.method == 'POST':
        subcategoria = request.POST.get('subcategoria', '').strip()
        categoria.nombre       = request.POST.get('nombre', '').strip()
        categoria.descripcion  = request.POST.get('descripcion', '').strip()
        categoria.subcategoria = subcategoria
        categoria.activo       = request.POST.get('activo') == 'on'
        categoria.edad_minima, categoria.edad_maxima = Categoria.EDAD_RANGOS.get(subcategoria, (None, None))
        categoria.save()
        messages.success(request, f'Categoría "{categoria.nombre}" actualizada.')
    return render(request, 'panel/categorias/form.html', {
        'titulo': 'Editar Categoría',
        'categoria': categoria,
        'subcategorias': Categoria.SUBCATEGORIA_CHOICES,
    })


@solo_admin
def categoria_eliminar(request, cat_id):
    """Eliminar categoría solo si no tiene equipos asociados."""
    categoria = get_object_or_404(Categoria, id=cat_id)
    if categoria.equipos.exists():
        messages.error(
            request,
            f'No se puede eliminar "{categoria.nombre}" porque tiene '
            f'{categoria.equipos.count()} equipo(s) asociado(s). '
            f'Eliminá o reasigná los equipos primero.'
        )
    else:
        nombre = categoria.nombre
        categoria.delete()
        messages.success(request, f'Categoría "{nombre}" eliminada.')
    return categorias_lista(request)


@solo_admin
def equipos_lista(request):
    """Lista de equipos con filtros por nombre, categoría y estado"""
    qs = Equipo.objects.select_related('categoria', 'entrenador').annotate(
        num_jugadores=Count('jugadores')).order_by('nombre')

    buscar       = request.GET.get('buscar', '').strip()
    categoria_id = request.GET.get('categoria', '')
    estado       = request.GET.get('estado', '')

    if buscar:
        qs = qs.filter(nombre__icontains=buscar)
    if categoria_id:
        qs = qs.filter(categoria_id=categoria_id)
    if estado == 'activo':
        qs = qs.filter(activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(activo=False)

    return render(request, 'panel/equipos/lista.html', {
        'equipos':      qs,
        'buscar':       buscar,
        'categoria_id': categoria_id,
        'estado':       estado,
        'categorias':   Categoria.objects.filter(activo=True).order_by('nombre'),
        'total_filtrados': qs.count(),
    })


@solo_admin
def admin_equipo_jugadores(request, equipo_id):
    """Lista de jugadores de un equipo con nombre, apellido, documento, posición y estado."""
    equipo    = get_object_or_404(Equipo, id=equipo_id)
    jugadores = Jugador.objects.filter(
        equipo=equipo
    ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        jugadores = jugadores.filter(
            Q(usuario__first_name__icontains=buscar) |
            Q(usuario__last_name__icontains=buscar)  |
            Q(usuario__documento__icontains=buscar)  |
            Q(posicion__icontains=buscar)
        )

    return render(request, 'panel/equipos/jugadores.html', {
        'equipo':    equipo,
        'jugadores': jugadores,
        'buscar':    buscar,
        'total':     jugadores.count(),
    })


@solo_admin
def equipo_crear(request):
    """Crear equipo - al seleccionar categoria filtra entrenadores disponibles"""
    if request.method == 'POST':
        nombre        = request.POST.get('nombre', '').strip()
        categoria_id  = request.POST.get('categoria')
        entrenador_id = request.POST.get('entrenador') or None
        activo        = request.POST.get('activo') == 'on'

        if not nombre or not categoria_id:
            messages.error(request, 'Nombre y categoría son obligatorios.')
        elif Equipo.objects.filter(nombre=nombre, categoria_id=categoria_id).exists():
            messages.error(request, f'Ya existe un equipo "{nombre}" en esa categoría.')
        else:
            Equipo.objects.create(
                nombre=nombre,
                categoria_id=categoria_id,
                entrenador_id=entrenador_id,
                activo=activo,
            )
            messages.success(request, f'Equipo "{nombre}" creado correctamente.')
            return equipos_lista(request)

    return render(request, 'panel/equipos/form.html', {
        'titulo':      'Crear Equipo',
        'categorias':  Categoria.objects.filter(activo=True),
        'entrenadores': UserModel.objects.filter(role='entrenador', is_active=True),
    })


@solo_admin
def equipo_editar(request, equipo_id):
    """Editar equipo - valida que el entrenador no tenga conflicto de categoria"""
    equipo = get_object_or_404(Equipo, id=equipo_id)
    if request.method == 'POST':
        equipo.nombre        = request.POST.get('nombre', '').strip()
        equipo.categoria_id  = request.POST.get('categoria')
        equipo.entrenador_id = request.POST.get('entrenador') or None
        equipo.activo        = request.POST.get('activo') == 'on'
        equipo.save()
        messages.success(request, f'Equipo "{equipo.nombre}" actualizado.')
        return equipos_lista(request)

    return render(request, 'panel/equipos/form.html', {
        'titulo':      'Editar Equipo',
        'equipo':      equipo,
        'categorias':  Categoria.objects.filter(activo=True),
        'entrenadores': UserModel.objects.filter(role='entrenador', is_active=True),
    })


@solo_admin
def equipo_eliminar(request, equipo_id):
    """Eliminar equipo solo si no tiene jugadores asociados."""
    equipo = get_object_or_404(Equipo, id=equipo_id)
    num_jugadores = equipo.jugadores.count()
    if num_jugadores > 0:
        messages.error(
            request,
            f'No se puede eliminar "{equipo.nombre}" porque tiene '
            f'{num_jugadores} jugador(es) asociado(s). '
            f'Desvinculá los jugadores primero.'
        )
    else:
        nombre = equipo.nombre
        equipo.delete()
        messages.success(request, f'Equipo "{nombre}" eliminado.')
    return equipos_lista(request)


# ─────────────────────────────────────────────
# GESTIÓN DE ENTRENAMIENTOS
# ─────────────────────────────────────────────

@solo_admin
def entrenamientos_lista(request):
    """Lista de entrenamientos con filtros multi criterio"""
    qs = Entrenamiento.objects.select_related('equipo', 'entrenador')

    buscar     = request.GET.get('buscar', '').strip()
    equipo_id  = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(titulo__icontains=buscar) |
            Q(lugar__icontains=buscar) |
            Q(descripcion__icontains=buscar)
        )
    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by('-fecha')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/entrenamientos/lista.html', {
        'entrenamientos': page,
        'buscar': buscar,
        'equipo_id': equipo_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'equipos': Equipo.objects.filter(activo=True),
        'total_filtrados': qs.count(),
    })


@solo_admin
def entrenamiento_crear(request):
    """Crear entrenamiento"""
    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        equipo_id = request.POST.get('equipo')
        entrenador_id = request.POST.get('entrenador') or None
        fecha = request.POST.get('fecha')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin = request.POST.get('hora_fin')
        lugar = request.POST.get('lugar', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if not all([titulo, equipo_id, fecha, hora_inicio, hora_fin]):
            messages.error(request, 'Completa todos los campos obligatorios.')
        else:
            ent = Entrenamiento.objects.create(
                titulo=titulo, equipo_id=equipo_id, entrenador_id=entrenador_id,
                fecha=fecha, hora_inicio=hora_inicio, hora_fin=hora_fin,
                lugar=lugar, descripcion=descripcion
            )
            _notificar_entrenamiento(ent, es_nuevo=True, emisor=request.user)
            messages.success(request, f'Entrenamiento "{titulo}" creado correctamente.')

    return render(request, 'panel/entrenamientos/form.html', {
        'titulo': 'Crear Entrenamiento',
        'equipos': Equipo.objects.filter(activo=True),
        'entrenadores': UserModel.objects.filter(role='entrenador', is_active=True),
    })


@solo_admin
def entrenamiento_editar(request, ent_id):
    """Editar entrenamiento"""
    entrenamiento = get_object_or_404(Entrenamiento, id=ent_id)
    if request.method == 'POST':
        entrenamiento.titulo = request.POST.get('titulo', '').strip()
        entrenamiento.equipo_id = request.POST.get('equipo')
        entrenamiento.entrenador_id = request.POST.get('entrenador') or None
        entrenamiento.fecha = request.POST.get('fecha')
        entrenamiento.hora_inicio = request.POST.get('hora_inicio')
        entrenamiento.hora_fin = request.POST.get('hora_fin')
        entrenamiento.lugar = request.POST.get('lugar', '').strip()
        entrenamiento.descripcion = request.POST.get('descripcion', '').strip()
        entrenamiento.save()
        _notificar_entrenamiento(entrenamiento, es_nuevo=False, emisor=request.user)
        messages.success(request, f'Entrenamiento "{entrenamiento.titulo}" actualizado.')
    return render(request, 'panel/entrenamientos/form.html', {
        'titulo': 'Editar Entrenamiento',
        'entrenamiento': entrenamiento,
        'equipos': Equipo.objects.filter(activo=True),
        'entrenadores': UserModel.objects.filter(role='entrenador', is_active=True),
    })


@solo_admin
def entrenamiento_eliminar(request, ent_id):
    """Eliminar entrenamiento"""
    entrenamiento = get_object_or_404(Entrenamiento, id=ent_id)
    nombre = entrenamiento.titulo
    entrenamiento.delete()
    messages.success(request, f'Entrenamiento "{nombre}" eliminado.')
    return entrenamientos_lista(request)


# ─────────────────────────────────────────────
# GESTIÓN DE TORNEOS
# ─────────────────────────────────────────────

@solo_admin
def torneos_lista(request):
    """Lista de torneos con filtros multi criterio"""
    qs = Torneo.objects.select_related('categoria').annotate(num_partidos=Count('partidos'))

    buscar      = request.GET.get('buscar', '').strip()
    categoria_id = request.GET.get('categoria', '')
    estado      = request.GET.get('estado', '')

    if buscar:
        qs = qs.filter(
            Q(nombre__icontains=buscar) |
            Q(lugar__icontains=buscar)
        )
    if categoria_id:
        qs = qs.filter(categoria_id=categoria_id)
    if estado:
        qs = qs.filter(estado=estado)

    qs = qs.order_by('-fecha_inicio')

    return render(request, 'panel/torneos/lista.html', {
        'torneos': qs,
        'buscar': buscar,
        'categoria_id': categoria_id,
        'estado': estado,
        'categorias': Categoria.objects.filter(activo=True),
        'estados': Torneo.ESTADO_CHOICES,
        'total_filtrados': qs.count(),
    })


@solo_admin
def torneo_crear(request):
    """Crear torneo"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        categoria_id = request.POST.get('categoria')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        lugar = request.POST.get('lugar', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        estado = request.POST.get('estado', 'planificado')

        if not all([nombre, categoria_id, fecha_inicio, fecha_fin]):
            messages.error(request, 'Completa todos los campos obligatorios.')
        else:
            torneo_obj = Torneo.objects.create(
                nombre=nombre, categoria_id=categoria_id,
                fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                lugar=lugar, descripcion=descripcion, estado=estado
            )
            _notificar_torneo(torneo_obj, es_nuevo=True, emisor=request.user)
            messages.success(request, f'Torneo "{nombre}" creado correctamente.')

    return render(request, 'panel/torneos/form.html', {
        'titulo': 'Crear Torneo',
        'categorias': Categoria.objects.filter(activo=True),
        'estados': Torneo.ESTADO_CHOICES,
    })


@solo_admin
def torneo_editar(request, torneo_id):
    """Editar torneo"""
    torneo = get_object_or_404(Torneo, id=torneo_id)
    if request.method == 'POST':
        torneo.nombre = request.POST.get('nombre', '').strip()
        torneo.categoria_id = request.POST.get('categoria')
        torneo.fecha_inicio = request.POST.get('fecha_inicio')
        torneo.fecha_fin = request.POST.get('fecha_fin')
        torneo.lugar = request.POST.get('lugar', '').strip()
        torneo.descripcion = request.POST.get('descripcion', '').strip()
        torneo.estado = request.POST.get('estado', torneo.estado)
        torneo.save()
        _notificar_torneo(torneo, es_nuevo=False, emisor=request.user)
        messages.success(request, f'Torneo "{torneo.nombre}" actualizado.')
    return render(request, 'panel/torneos/form.html', {
        'titulo': 'Editar Torneo',
        'torneo': torneo,
        'categorias': Categoria.objects.filter(activo=True),
        'estados': Torneo.ESTADO_CHOICES,
    })


@solo_admin
def torneo_eliminar(request, torneo_id):
    """Eliminar torneo"""
    torneo = get_object_or_404(Torneo, id=torneo_id)
    nombre = torneo.nombre
    torneo.delete()
    messages.success(request, f'Torneo "{nombre}" eliminado.')
    return torneos_lista(request)


# ─────────────────────────────────────────────
# GESTIÓN FINANCIERA
# ─────────────────────────────────────────────

@solo_admin
def finanzas_lista(request):
    """Lista de movimientos financieros con filtros multi criterio"""
    qs = Finanza.objects.select_related('registrado_por')

    buscar      = request.GET.get('buscar', '').strip()
    tipo        = request.GET.get('tipo', '')
    categoria   = request.GET.get('categoria', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(descripcion__icontains=buscar)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if categoria:
        qs = qs.filter(categoria=categoria)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by('-fecha')

    # Totales sobre el queryset filtrado
    total_ingresos = qs.filter(tipo='ingreso').aggregate(total=Sum('monto'))['total'] or Decimal('0')
    total_egresos  = qs.filter(tipo='egreso').aggregate(total=Sum('monto'))['total'] or Decimal('0')

    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/finanzas/lista.html', {
        'finanzas': page,
        'buscar': buscar,
        'tipo': tipo,
        'categoria': categoria,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos': Finanza.TIPO_CHOICES,
        'categorias': Finanza.CATEGORIA_CHOICES,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'balance': total_ingresos - total_egresos,
        'total_filtrados': qs.count(),
    })


@solo_admin
def finanza_crear(request):
    """Crear movimiento financiero"""
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        categoria = request.POST.get('categoria', 'otro')
        descripcion = request.POST.get('descripcion', '').strip()
        monto = request.POST.get('monto')
        fecha = request.POST.get('fecha')

        if not all([tipo, descripcion, monto, fecha]):
            messages.error(request, 'Completa todos los campos obligatorios.')
        else:
            Finanza.objects.create(
                tipo=tipo, categoria=categoria, descripcion=descripcion,
                monto=monto, fecha=fecha, registrado_por=request.user
            )
            messages.success(request, f'{tipo.capitalize()} registrado correctamente.')

    return render(request, 'panel/finanzas/form.html', {
        'titulo': 'Registrar Movimiento',
        'tipos': Finanza.TIPO_CHOICES,
        'categorias': Finanza.CATEGORIA_CHOICES,
    })


@solo_admin
def finanza_eliminar(request, fin_id):
    """Eliminar movimiento financiero"""
    finanza = get_object_or_404(Finanza, id=fin_id)
    finanza.delete()
    messages.success(request, 'Movimiento eliminado correctamente.')
    return finanzas_lista(request)


# ─────────────────────────────────────────────
# GESTIÓN DE PAGOS (ADMIN)
# ─────────────────────────────────────────────

def _marcar_vencidos():
    """Ejecuta el marcado automático de pagos vencidos antes de listar."""
    Pago.marcar_vencidos()


@solo_admin
def admin_pagos_lista(request):
    """
    Lista todos los pagos del sistema con filtros.
    Marca automáticamente como VENCIDO los pagos pendientes con fecha pasada.
    """
    _marcar_vencidos()

    qs = Pago.objects.select_related(
        'jugador__usuario', 'jugador__equipo__categoria', 'registrado_por'
    )

    # Filtros
    buscar    = request.GET.get('buscar', '').strip()
    estado    = request.GET.get('estado', '')
    concepto  = request.GET.get('concepto', '')
    equipo_id = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(descripcion__icontains=buscar) |
            Q(jugador__usuario__first_name__icontains=buscar) |
            Q(jugador__usuario__last_name__icontains=buscar) |
            Q(jugador__usuario__documento__icontains=buscar)
        )
    if estado:
        qs = qs.filter(estado=estado)
    if concepto:
        qs = qs.filter(concepto=concepto)
    if equipo_id:
        qs = qs.filter(jugador__equipo_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)

    # Totales sobre el queryset filtrado
    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    alertas_vencidos  = qs.filter(estado='vencido').count()
    alertas_revision   = Pago.objects.filter(estado='en_revision').count()

    paginator = Paginator(qs.order_by('-fecha_vencimiento'), 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/pagos/lista.html', {
        'pagos':             page,
        'buscar':            buscar,
        'estado':            estado,
        'concepto':          concepto,
        'equipo_id':         equipo_id,
        'fecha_desde':       fecha_desde,
        'fecha_hasta':       fecha_hasta,
        'estados':           Pago.ESTADO_CHOICES,
        'conceptos':         Pago.CONCEPTO_CHOICES,
        'equipos':           Equipo.objects.filter(activo=True),
        'total_pagado':      total_pagado,
        'total_pendiente':   total_pendiente,
        'total_vencido':     total_vencido,
        'alertas_vencidos':  alertas_vencidos,
        'alertas_revision':  alertas_revision,
        'total_filtrados':   qs.count(),
    })


@solo_admin
def admin_pago_crear(request):
    """
    Crear pago para un deportista.
    Validaciones: jugador válido, solo rol deportista, monto > 0, estado controlado.
    """
    jugadores = Jugador.objects.filter(activo=True).select_related('usuario', 'equipo')

    if request.method == 'POST':
        jugador_id        = request.POST.get('jugador')
        concepto          = request.POST.get('concepto', 'cuota_mensual')
        descripcion       = request.POST.get('descripcion', '').strip()
        monto             = request.POST.get('monto', '').strip()
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        fecha_pago        = request.POST.get('fecha_pago') or None
        estado            = request.POST.get('estado', 'pendiente')
        metodo_pago       = request.POST.get('metodo_pago', '')

        errores = []
        if not jugador_id:
            errores.append('Debes seleccionar un deportista.')
        else:
            jugador_obj = Jugador.objects.filter(id=jugador_id, activo=True).first()
            if not jugador_obj:
                errores.append('El deportista seleccionado no es válido.')
            elif jugador_obj.usuario.role != 'deportista':
                errores.append('Solo se pueden registrar pagos a usuarios con rol Deportista.')
        try:
            monto_dec = Decimal(monto)
            if monto_dec <= 0:
                errores.append('El monto debe ser mayor a 0.')
        except Exception:
            errores.append('El monto ingresado no es válido.')
        if not descripcion:
            errores.append('La descripción es obligatoria.')
        if not fecha_vencimiento:
            errores.append('La fecha de vencimiento es obligatoria.')
        if estado not in [v for v, _ in Pago.ESTADO_CHOICES]:
            errores.append('El estado seleccionado no es válido.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            pago_obj = Pago.objects.create(
                jugador_id=jugador_id,
                concepto=concepto,
                descripcion=descripcion,
                monto=monto_dec,
                fecha_vencimiento=fecha_vencimiento,
                fecha_pago=fecha_pago,
                estado=estado,
                metodo_pago=metodo_pago,
                registrado_por=request.user,
            )
            pago_obj.refresh_from_db()
            _notificar_pago(pago_obj, 'registrado', request.user)
            messages.success(request, 'Pago registrado correctamente.')
            return admin_pagos_lista(request)

    return render(request, 'panel/pagos/form.html', {
        'titulo':    'Registrar Pago',
        'jugadores': jugadores,
        'estados':   Pago.ESTADO_CHOICES,
        'conceptos': Pago.CONCEPTO_CHOICES,
        'metodos':   Pago.METODO_CHOICES,
    })


@solo_admin
def admin_pago_editar(request, pago_id):
    """
    Editar pago existente.
    Permite actualizar estado (ej: pendiente → pagado) y método de pago.
    """
    pago      = get_object_or_404(Pago, id=pago_id)
    jugadores = Jugador.objects.filter(activo=True).select_related('usuario', 'equipo')

    if request.method == 'POST':
        jugador_id        = request.POST.get('jugador')
        concepto          = request.POST.get('concepto', pago.concepto)
        descripcion       = request.POST.get('descripcion', '').strip()
        monto             = request.POST.get('monto', '').strip()
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        fecha_pago        = request.POST.get('fecha_pago') or None
        estado            = request.POST.get('estado', pago.estado)
        metodo_pago       = request.POST.get('metodo_pago', '')

        errores = []
        try:
            monto_dec = Decimal(monto)
            if monto_dec <= 0:
                errores.append('El monto debe ser mayor a 0.')
        except Exception:
            errores.append('El monto ingresado no es válido.')

        if not descripcion:
            errores.append('La descripción es obligatoria.')
        if estado not in [v for v, _ in Pago.ESTADO_CHOICES]:
            errores.append('El estado seleccionado no es válido.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            pago.jugador_id        = jugador_id
            pago.concepto          = concepto
            pago.descripcion       = descripcion
            pago.monto             = monto_dec
            pago.fecha_vencimiento = fecha_vencimiento
            pago.fecha_pago        = fecha_pago
            pago.estado            = estado
            pago.metodo_pago       = metodo_pago
            pago.save()
            # Sincronizar con finanzas
            if estado == 'pagado':
                _crear_finanza_por_pago(pago, request.user)
            else:
                Finanza.objects.filter(pago=pago).delete()
            _notificar_pago(pago, 'actualizado', request.user)
            messages.success(request, 'Pago actualizado correctamente.')
            return admin_pagos_lista(request)

    return render(request, 'panel/pagos/form.html', {
        'titulo':    'Editar Pago',
        'pago':      pago,
        'jugadores': jugadores,
        'estados':   Pago.ESTADO_CHOICES,
        'conceptos': Pago.CONCEPTO_CHOICES,
        'metodos':   Pago.METODO_CHOICES,
    })


@solo_admin
def admin_pago_eliminar(request, pago_id):
    """Eliminar pago y su movimiento financiero asociado si existe."""
    pago = get_object_or_404(Pago, id=pago_id)
    Finanza.objects.filter(pago=pago).delete()
    pago.delete()
    messages.success(request, 'Pago eliminado correctamente.')
    return admin_pagos_lista(request)


def _crear_finanza_por_pago(pago, registrado_por):
    """Crea un ingreso en Finanza vinculado al pago si no existe ya."""
    if not Finanza.objects.filter(pago=pago).exists():
        concepto_categoria = {
            'cuota_mensual': 'cuota',
            'inscripcion':   'otro',
            'equipamiento':  'equipamiento',
            'torneo':        'otro',
            'otro':          'otro',
        }
        Finanza.objects.create(
            tipo='ingreso',
            categoria=concepto_categoria.get(pago.concepto, 'otro'),
            descripcion=f'Pago: {pago.descripcion} — {pago.jugador.usuario.get_full_name() or pago.jugador.usuario.username}',
            monto=pago.monto,
            fecha=pago.fecha_pago or timezone.now().date(),
            registrado_por=registrado_por,
            pago=pago,
        )


@solo_admin
def admin_cuenta_deportista(request, jug_id):
    """
    Historial financiero completo de un deportista: pagos, facturas y saldos.
    """
    _marcar_vencidos()
    jugador = get_object_or_404(
        Jugador.objects.select_related('usuario', 'equipo__categoria'),
        id=jug_id
    )
    qs = Pago.objects.filter(jugador=jugador).order_by('-fecha_vencimiento')

    estado      = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if estado:
        qs = qs.filter(estado=estado)
    if fecha_desde:
        qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)

    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado__in=['pendiente', 'vencido', 'en_revision']).aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')

    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/pagos/cuenta_deportista.html', {
        'jugador':         jugador,
        'pagos':           page,
        'estado':          estado,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'estados':         Pago.ESTADO_CHOICES,
        'total_pagado':    total_pagado,
        'total_pendiente': total_pendiente,
        'total_vencido':   total_vencido,
        'total_filtrados': qs.count(),
    })


@solo_admin
def admin_facturas_lista(request):
    """
    Lista de facturas (pagos) filtrables por deportista, fecha y estado.
    Permite descargar la factura PDF individual de cada pago.
    """
    _marcar_vencidos()

    qs = Pago.objects.select_related(
        'jugador__usuario', 'jugador__equipo__categoria', 'registrado_por'
    )

    jugador_id  = request.GET.get('jugador', '')
    estado      = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if jugador_id:
        qs = qs.filter(jugador_id=jugador_id)
    if estado:
        qs = qs.filter(estado=estado)
    if fecha_desde:
        qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)

    qs = qs.order_by('-fecha_vencimiento')
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))

    jugadores = Jugador.objects.filter(
        activo=True
    ).select_related('usuario').order_by('usuario__last_name', 'usuario__first_name')

    return render(request, 'panel/pagos/facturas.html', {
        'pagos':        page,
        'jugador_id':   jugador_id,
        'estado':       estado,
        'fecha_desde':  fecha_desde,
        'fecha_hasta':  fecha_hasta,
        'jugadores':    jugadores,
        'estados':      Pago.ESTADO_CHOICES,
        'total_filtrados': qs.count(),
    })


@solo_admin
def admin_conceptos_lista(request):
    """Lista todos los conceptos de pago registrados."""
    conceptos = ConceptoPago.objects.all().order_by('nombre')
    return render(request, 'panel/pagos/conceptos_lista.html', {'conceptos': conceptos})


@solo_admin
def admin_concepto_crear(request):
    """Crear nuevo concepto de pago."""
    if request.method == 'POST':
        nombre      = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        monto_base  = request.POST.get('monto_base', '0').strip() or '0'

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        elif ConceptoPago.objects.filter(nombre__iexact=nombre).exists():
            messages.error(request, f'Ya existe un concepto con el nombre "{nombre}".')
        else:
            ConceptoPago.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                monto_base=monto_base,
            )
            messages.success(request, f'Concepto "{nombre}" creado correctamente.')
            return admin_conceptos_lista(request)

    return render(request, 'panel/pagos/concepto_form.html', {
        'titulo': 'Nuevo Concepto de Pago',
    })


@solo_admin
def admin_concepto_editar(request, concepto_id):
    """Modificar nombre, descripción y monto base de un concepto de pago."""
    concepto = get_object_or_404(ConceptoPago, id=concepto_id)
    if request.method == 'POST':
        nombre      = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        monto_base  = request.POST.get('monto_base', '0').strip() or '0'
        activo      = request.POST.get('activo') == 'on'

        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        elif ConceptoPago.objects.filter(nombre__iexact=nombre).exclude(id=concepto_id).exists():
            messages.error(request, f'Ya existe otro concepto con el nombre "{nombre}".')
        else:
            concepto.nombre      = nombre
            concepto.descripcion = descripcion
            concepto.monto_base  = monto_base
            concepto.activo      = activo
            concepto.save()
            messages.success(request, f'Concepto "{nombre}" actualizado correctamente.')
            return admin_conceptos_lista(request)

    return render(request, 'panel/pagos/concepto_form.html', {
        'titulo':   'Editar Concepto de Pago',
        'concepto': concepto,
    })


@solo_admin
def admin_concepto_eliminar(request, concepto_id):
    """Eliminar concepto de pago solo si no tiene pagos asociados."""
    concepto = get_object_or_404(ConceptoPago, id=concepto_id)
    if Pago.objects.filter(descripcion__icontains=concepto.nombre).exists():
        messages.error(
            request,
            f'No se puede eliminar "{concepto.nombre}" porque tiene pagos asociados.'
        )
    else:
        nombre = concepto.nombre
        concepto.delete()
        messages.success(request, f'Concepto "{nombre}" eliminado correctamente.')
    return admin_conceptos_lista(request)


@solo_admin
def admin_factura_pago(request, pago_id):
    """Genera PDF de factura individual para un pago."""
    pago = get_object_or_404(
        Pago.objects.select_related(
            'jugador__usuario', 'jugador__equipo__categoria', 'registrado_por'
        ),
        id=pago_id
    )
    hoy = timezone.now()
    html = render_to_string('panel/pagos/factura.html', {
        'pago':           pago,
        'fecha_emision':  hoy,
        'emitido_por':    request.user.get_full_name() or request.user.username,
    })
    buffer = BytesIO()
    pisa.CreatePDF(html.encode('utf-8'), dest=buffer)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factura_pago_{pago_id}.pdf"'
    response['Content-Length'] = len(pdf)
    return response


@solo_admin
def admin_pago_marcar_pendiente(request, pago_id):
    """
    Acción rápida: revertir un pago PAGADO a PENDIENTE.
    Elimina el movimiento financiero asociado si existe.
    """
    pago = get_object_or_404(Pago, id=pago_id)
    pago.estado     = 'pendiente'
    pago.fecha_pago = None
    pago.save()
    Finanza.objects.filter(pago=pago).delete()
    messages.success(
        request,
        f'Pago de {pago.jugador.usuario.get_full_name() or pago.jugador.usuario.username} revertido a Pendiente.'
    )
    return admin_pagos_lista(request)


@solo_admin
def admin_pago_marcar_pagado(request, pago_id):
    """
    Acción rápida: marcar un pago como PAGADO desde la lista.
    Registra la fecha de pago como hoy si no tiene una y crea el ingreso en finanzas.
    """
    pago = get_object_or_404(Pago, id=pago_id)
    pago.estado = 'pagado'
    if not pago.fecha_pago:
        pago.fecha_pago = timezone.now().date()
    pago.save()
    _crear_finanza_por_pago(pago, request.user)
    _notificar_pago(pago, 'pagado', request.user)
    messages.success(request, f'Pago de {pago.jugador.usuario.get_full_name() or pago.jugador.usuario.username} marcado como pagado.')
    return admin_pagos_lista(request)


# ─────────────────────────────────────────────
# REPORTES FINANCIEROS (ADMIN)
# ─────────────────────────────────────────────

@solo_admin
def admin_reporte_financiero(request):
    """
    Reporte financiero unificado: pagos filtrados por periodo,
    ingresos confirmados, facturas y resumen por estado.
    """
    _marcar_vencidos()
    hoy = timezone.now().date()

    fecha_desde  = request.GET.get('fecha_desde', '')
    fecha_hasta  = request.GET.get('fecha_hasta', '')
    estado       = request.GET.get('estado', '')
    equipo_id    = request.GET.get('equipo', '')
    concepto     = request.GET.get('concepto', '')

    qs = Pago.objects.select_related(
        'jugador__usuario', 'jugador__equipo__categoria', 'registrado_por'
    )
    if fecha_desde:
        qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)
    if estado:
        qs = qs.filter(estado=estado)
    if equipo_id:
        qs = qs.filter(jugador__equipo_id=equipo_id)
    if concepto:
        qs = qs.filter(concepto=concepto)

    qs = qs.order_by('-fecha_vencimiento')

    # Totales sobre el queryset filtrado
    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_revision  = qs.filter(estado='en_revision').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_general   = total_pagado + total_pendiente + total_vencido + total_revision

    count_pagado    = qs.filter(estado='pagado').count()
    count_pendiente = qs.filter(estado='pendiente').count()
    count_vencido   = qs.filter(estado='vencido').count()
    count_revision  = qs.filter(estado='en_revision').count()

    # Ingresos confirmados por mes sobre el periodo filtrado
    qs_pagados = qs.filter(estado='pagado', fecha_pago__isnull=False)
    por_mes = (
        qs_pagados
        .annotate(mes=TruncMonth('fecha_pago'))
        .values('mes')
        .annotate(total=Sum('monto'), cantidad=Count('id'))
        .order_by('mes')
    )

    # Proximos a vencer en los proximos 7 dias (sobre el qs filtrado)
    proximos_vencer = qs.filter(
        estado='pendiente',
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timedelta(days=7)
    ).order_by('fecha_vencimiento')

    params = request.GET.urlencode()
    return render(request, 'panel/reportes/financiero.html', {
        'pagos':            qs,
        'fecha_desde':      fecha_desde,
        'fecha_hasta':      fecha_hasta,
        'estado':           estado,
        'equipo_id':        equipo_id,
        'concepto':         concepto,
        'equipos':          Equipo.objects.filter(activo=True).order_by('nombre'),
        'estados':          Pago.ESTADO_CHOICES,
        'conceptos':        Pago.CONCEPTO_CHOICES,
        'total_pagado':     total_pagado,
        'total_pendiente':  total_pendiente,
        'total_vencido':    total_vencido,
        'total_revision':   total_revision,
        'total_general':    total_general,
        'count_pagado':     count_pagado,
        'count_pendiente':  count_pendiente,
        'count_vencido':    count_vencido,
        'count_revision':   count_revision,
        'por_mes':          por_mes,
        'proximos_vencer':  proximos_vencer,
        'total_filtrados':  qs.count(),
        'query_params':     params,
        'hoy':              hoy,
    })


@solo_admin
def admin_reporte_pagos(request):
    """
    Reporte financiero completo:
    - Totales globales por estado
    - Ingresos por mes (últimos 12 meses)
    - Pagos por equipo
    - Pagos por concepto
    - Pagos por categoría
    """
    _marcar_vencidos()
    hoy = timezone.now().date()

    # Totales globales
    total_pagado    = Pago.objects.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = Pago.objects.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_vencido   = Pago.objects.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_general   = total_pagado + total_pendiente + total_vencido

    # Conteos por estado
    count_pagado    = Pago.objects.filter(estado='pagado').count()
    count_pendiente = Pago.objects.filter(estado='pendiente').count()
    count_vencido   = Pago.objects.filter(estado='vencido').count()

    # Ingresos confirmados por mes (últimos 12 meses)
    por_mes = (
        Pago.objects
        .filter(estado='pagado', fecha_pago__isnull=False)
        .annotate(mes=TruncMonth('fecha_pago'))
        .values('mes')
        .annotate(total=Sum('monto'), cantidad=Count('id'))
        .order_by('-mes')[:12]
    )

    # Pagos por equipo
    por_equipo = (
        Pago.objects
        .filter(jugador__equipo__isnull=False)
        .values('jugador__equipo__nombre')
        .annotate(
            total=Sum('monto'),
            pagados=Sum('monto', filter=Q(estado='pagado')),
            pendientes=Count('id', filter=Q(estado='pendiente')),
            vencidos=Count('id', filter=Q(estado='vencido')),
        )
        .order_by('-total')
    )

    # Pagos por concepto
    por_concepto = (
        Pago.objects
        .values('concepto')
        .annotate(total=Sum('monto'), cantidad=Count('id'))
        .order_by('-total')
    )

    # Pagos por categoría
    por_categoria = (
        Pago.objects
        .filter(jugador__equipo__categoria__isnull=False)
        .values('jugador__equipo__categoria__nombre')
        .annotate(total=Sum('monto'), cantidad=Count('id'))
        .order_by('-total')
    )

    # Próximos a vencer (pendientes en los próximos 7 días)
    proximos_vencer = Pago.objects.filter(
        estado='pendiente',
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=hoy + timedelta(days=7)
    ).select_related('jugador__usuario', 'jugador__equipo').order_by('fecha_vencimiento')

    return render(request, 'panel/pagos/reporte.html', {
        'total_pagado':     total_pagado,
        'total_pendiente':  total_pendiente,
        'total_vencido':    total_vencido,
        'total_general':    total_general,
        'count_pagado':     count_pagado,
        'count_pendiente':  count_pendiente,
        'count_vencido':    count_vencido,
        'por_mes':          por_mes,
        'por_equipo':       por_equipo,
        'por_concepto':     por_concepto,
        'por_categoria':    por_categoria,
        'proximos_vencer':  proximos_vencer,
        'hoy':              hoy,
    })


# ─────────────────────────────────────────────
# ESTADISTICAS DE ASISTENCIA Y PARTIDOS
# ─────────────────────────────────────────────

@solo_admin
def admin_estadisticas_asistencia(request):
    """Estadisticas de asistencia: porcentaje por deportista y por equipo (admin)."""
    equipos      = Equipo.objects.filter(activo=True).order_by('nombre')
    categorias   = Categoria.objects.filter(activo=True).order_by('nombre')
    equipo_id    = request.GET.get('equipo', '')
    categoria_id = request.GET.get('categoria', '')

    jugadores_qs = Jugador.objects.filter(activo=True).select_related('usuario', 'equipo__categoria')
    if equipo_id:    jugadores_qs = jugadores_qs.filter(equipo_id=equipo_id)
    if categoria_id: jugadores_qs = jugadores_qs.filter(equipo__categoria_id=categoria_id)

    # Porcentaje por deportista
    stats_deportistas = []
    for jug in jugadores_qs.order_by('usuario__last_name'):
        total  = Asistencia.objects.filter(jugador=jug).count()
        asistio = Asistencia.objects.filter(jugador=jug, estado='asistio').count()
        no_asistio = Asistencia.objects.filter(jugador=jug, estado='no_asistio').count()
        justificado = Asistencia.objects.filter(jugador=jug, estado='justificado').count()
        pct = round((asistio / total) * 100, 1) if total > 0 else 0
        stats_deportistas.append({
            'jugador': jug,
            'total': total,
            'asistio': asistio,
            'no_asistio': no_asistio,
            'justificado': justificado,
            'pct': pct,
        })

    # Porcentaje por equipo
    stats_equipos = []
    equipos_filtro = equipos if not equipo_id else equipos.filter(id=equipo_id)
    if categoria_id: equipos_filtro = equipos_filtro.filter(categoria_id=categoria_id)
    for eq in equipos_filtro:
        total   = Asistencia.objects.filter(entrenamiento__equipo=eq).count()
        asistio = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='asistio').count()
        no_asistio = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='no_asistio').count()
        justificado = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='justificado').count()
        pct = round((asistio / total) * 100, 1) if total > 0 else 0
        stats_equipos.append({
            'equipo': eq,
            'total': total,
            'asistio': asistio,
            'no_asistio': no_asistio,
            'justificado': justificado,
            'pct': pct,
        })

    dep_nombres = json.dumps([s['jugador'].usuario.get_full_name() or s['jugador'].usuario.username for s in stats_deportistas[:15]])
    dep_pct     = json.dumps([s['pct'] for s in stats_deportistas[:15]])
    eq_nombres  = json.dumps([s['equipo'].nombre for s in stats_equipos])
    eq_pct      = json.dumps([s['pct'] for s in stats_equipos])

    return render(request, 'panel/reportes/estadisticas_asistencia.html', {
        'equipos':           equipos,
        'categorias':        categorias,
        'equipo_id':         equipo_id,
        'categoria_id':      categoria_id,
        'stats_deportistas': stats_deportistas,
        'stats_equipos':     stats_equipos,
        'dep_nombres':       dep_nombres,
        'dep_pct':           dep_pct,
        'eq_nombres':        eq_nombres,
        'eq_pct':            eq_pct,
    })


@solo_admin
def admin_estadisticas_partidos(request):
    """Estadisticas de partidos: ganados, perdidos, empatados y goles (admin)."""
    equipos   = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id = request.GET.get('equipo', '')

    qs = PartidoCalendario.objects.filter(goles_favor__isnull=False)
    if equipo_id: qs = qs.filter(equipo_propio_id=equipo_id)

    total     = qs.count()
    ganados   = qs.filter(goles_favor__gt=Count('id', filter=Q(goles_favor__gt=0))).count()
    # Calculo directo con Python para evitar anotaciones complejas
    ganados   = sum(1 for p in qs if p.goles_favor > p.goles_contra)
    perdidos  = sum(1 for p in qs if p.goles_favor < p.goles_contra)
    empatados = sum(1 for p in qs if p.goles_favor == p.goles_contra)
    goles_favor  = qs.aggregate(t=Sum('goles_favor'))['t'] or 0
    goles_contra = qs.aggregate(t=Sum('goles_contra'))['t'] or 0

    # Estadisticas por equipo
    stats_equipos = []
    equipos_filtro = equipos if not equipo_id else equipos.filter(id=equipo_id)
    for eq in equipos_filtro:
        pqs = PartidoCalendario.objects.filter(equipo_propio=eq, goles_favor__isnull=False)
        if not pqs.exists(): continue
        g = sum(1 for p in pqs if p.goles_favor > p.goles_contra)
        pe = sum(1 for p in pqs if p.goles_favor < p.goles_contra)
        e = sum(1 for p in pqs if p.goles_favor == p.goles_contra)
        gf = pqs.aggregate(t=Sum('goles_favor'))['t'] or 0
        gc = pqs.aggregate(t=Sum('goles_contra'))['t'] or 0
        stats_equipos.append({
            'equipo': eq, 'total': pqs.count(),
            'ganados': g, 'perdidos': pe, 'empatados': e,
            'goles_favor': gf, 'goles_contra': gc,
            'diferencia': gf - gc,
        })

    eq_nombres  = json.dumps([s['equipo'].nombre for s in stats_equipos])
    eq_ganados  = json.dumps([s['ganados']  for s in stats_equipos])
    eq_perdidos = json.dumps([s['perdidos'] for s in stats_equipos])
    eq_empatados= json.dumps([s['empatados']for s in stats_equipos])

    return render(request, 'panel/reportes/estadisticas_partidos.html', {
        'equipos':      equipos,
        'equipo_id':    equipo_id,
        'total':        total,
        'ganados':      ganados,
        'perdidos':     perdidos,
        'empatados':    empatados,
        'goles_favor':  goles_favor,
        'goles_contra': goles_contra,
        'stats_equipos': stats_equipos,
        'eq_nombres':   eq_nombres,
        'eq_ganados':   eq_ganados,
        'eq_perdidos':  eq_perdidos,
        'eq_empatados': eq_empatados,
    })


@solo_entrenador
def ent_estadisticas_asistencia(request):
    """Estadisticas de asistencia: porcentaje por deportista y por equipo (entrenador)."""
    mis_equipos  = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id    = request.GET.get('equipo', '')

    jugadores_qs = Jugador.objects.filter(equipo__in=mis_equipos, activo=True).select_related('usuario', 'equipo')
    if equipo_id: jugadores_qs = jugadores_qs.filter(equipo_id=equipo_id)

    stats_deportistas = []
    for jug in jugadores_qs.order_by('usuario__last_name'):
        total       = Asistencia.objects.filter(jugador=jug).count()
        asistio     = Asistencia.objects.filter(jugador=jug, estado='asistio').count()
        no_asistio  = Asistencia.objects.filter(jugador=jug, estado='no_asistio').count()
        justificado = Asistencia.objects.filter(jugador=jug, estado='justificado').count()
        pct = round((asistio / total) * 100, 1) if total > 0 else 0
        stats_deportistas.append({
            'jugador': jug, 'total': total,
            'asistio': asistio, 'no_asistio': no_asistio,
            'justificado': justificado, 'pct': pct,
        })

    stats_equipos = []
    equipos_filtro = mis_equipos if not equipo_id else mis_equipos.filter(id=equipo_id)
    for eq in equipos_filtro:
        total       = Asistencia.objects.filter(entrenamiento__equipo=eq).count()
        asistio     = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='asistio').count()
        no_asistio  = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='no_asistio').count()
        justificado = Asistencia.objects.filter(entrenamiento__equipo=eq, estado='justificado').count()
        pct = round((asistio / total) * 100, 1) if total > 0 else 0
        stats_equipos.append({
            'equipo': eq, 'total': total,
            'asistio': asistio, 'no_asistio': no_asistio,
            'justificado': justificado, 'pct': pct,
        })

    dep_nombres = json.dumps([s['jugador'].usuario.get_full_name() or s['jugador'].usuario.username for s in stats_deportistas[:15]])
    dep_pct     = json.dumps([s['pct'] for s in stats_deportistas[:15]])
    eq_nombres  = json.dumps([s['equipo'].nombre for s in stats_equipos])
    eq_pct      = json.dumps([s['pct'] for s in stats_equipos])

    return render(request, 'entrenador/reportes/estadisticas_asistencia.html', {
        'equipos':           mis_equipos,
        'equipo_id':         equipo_id,
        'stats_deportistas': stats_deportistas,
        'stats_equipos':     stats_equipos,
        'dep_nombres':       dep_nombres,
        'dep_pct':           dep_pct,
        'eq_nombres':        eq_nombres,
        'eq_pct':            eq_pct,
    })


@solo_entrenador
def ent_estadisticas_partidos(request):
    """Estadisticas de partidos: ganados, perdidos, empatados y goles (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id   = request.GET.get('equipo', '')

    qs = PartidoCalendario.objects.filter(equipo_propio__in=mis_equipos, goles_favor__isnull=False)
    if equipo_id: qs = qs.filter(equipo_propio_id=equipo_id)

    total        = qs.count()
    ganados      = sum(1 for p in qs if p.goles_favor > p.goles_contra)
    perdidos     = sum(1 for p in qs if p.goles_favor < p.goles_contra)
    empatados    = sum(1 for p in qs if p.goles_favor == p.goles_contra)
    goles_favor  = qs.aggregate(t=Sum('goles_favor'))['t'] or 0
    goles_contra = qs.aggregate(t=Sum('goles_contra'))['t'] or 0

    stats_equipos = []
    equipos_filtro = mis_equipos if not equipo_id else mis_equipos.filter(id=equipo_id)
    for eq in equipos_filtro:
        pqs = PartidoCalendario.objects.filter(equipo_propio=eq, goles_favor__isnull=False)
        if not pqs.exists(): continue
        g  = sum(1 for p in pqs if p.goles_favor > p.goles_contra)
        pe = sum(1 for p in pqs if p.goles_favor < p.goles_contra)
        e  = sum(1 for p in pqs if p.goles_favor == p.goles_contra)
        gf = pqs.aggregate(t=Sum('goles_favor'))['t'] or 0
        gc = pqs.aggregate(t=Sum('goles_contra'))['t'] or 0
        stats_equipos.append({
            'equipo': eq, 'total': pqs.count(),
            'ganados': g, 'perdidos': pe, 'empatados': e,
            'goles_favor': gf, 'goles_contra': gc,
            'diferencia': gf - gc,
        })

    eq_nombres   = json.dumps([s['equipo'].nombre for s in stats_equipos])
    eq_ganados   = json.dumps([s['ganados']   for s in stats_equipos])
    eq_perdidos  = json.dumps([s['perdidos']  for s in stats_equipos])
    eq_empatados = json.dumps([s['empatados'] for s in stats_equipos])

    return render(request, 'entrenador/reportes/estadisticas_partidos.html', {
        'equipos':       mis_equipos,
        'equipo_id':     equipo_id,
        'total':         total,
        'ganados':       ganados,
        'perdidos':      perdidos,
        'empatados':     empatados,
        'goles_favor':   goles_favor,
        'goles_contra':  goles_contra,
        'stats_equipos': stats_equipos,
        'eq_nombres':    eq_nombres,
        'eq_ganados':    eq_ganados,
        'eq_perdidos':   eq_perdidos,
        'eq_empatados':  eq_empatados,
    })


# ─────────────────────────────────────────────
# EXPORTAR CSV
# ─────────────────────────────────────────────


@solo_admin
def exportar_csv(request, modulo):
    hoy = timezone.now()
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reporte_{modulo}_{hoy.strftime("%Y%m%d_%H%M")}.csv"'
    response.write('\ufeff')  # BOM para compatibilidad con Excel
    writer = csv.writer(response)

    if modulo == 'usuarios':
        buscar = request.GET.get('buscar', '').strip()
        rol    = request.GET.get('rol', '')
        estado = request.GET.get('estado', '')
        qs = UserModel.objects.all()
        if buscar:
            qs = qs.filter(
                Q(username__icontains=buscar) | Q(first_name__icontains=buscar) |
                Q(last_name__icontains=buscar) | Q(email__icontains=buscar) |
                Q(documento__icontains=buscar)
            )
        if rol:    qs = qs.filter(role=rol)
        if estado == 'activo':   qs = qs.filter(is_active=True)
        if estado == 'inactivo': qs = qs.filter(is_active=False)
        writer.writerow(['Username', 'Nombre', 'Email', 'Documento', 'Rol', 'Estado', 'Registro'])
        for u in qs.order_by('-date_joined'):
            writer.writerow([u.username, u.get_full_name(), u.email, u.documento,
                             u.get_role_display(), 'Activo' if u.is_active else 'Inactivo',
                             u.date_joined.strftime('%d/%m/%Y')])

    elif modulo == 'deportistas':
        buscar       = request.GET.get('buscar', '').strip()
        equipo_id    = request.GET.get('equipo', '')
        categoria_id = request.GET.get('categoria', '')
        estado       = request.GET.get('estado', '')
        qs = Jugador.objects.select_related('usuario', 'equipo__categoria').order_by('usuario__last_name')
        if buscar:
            qs = qs.filter(
                Q(usuario__first_name__icontains=buscar) |
                Q(usuario__last_name__icontains=buscar) |
                Q(usuario__documento__icontains=buscar)
            )
        if equipo_id:    qs = qs.filter(equipo_id=equipo_id)
        if categoria_id: qs = qs.filter(equipo__categoria_id=categoria_id)
        if estado == 'activo':   qs = qs.filter(activo=True)
        if estado == 'inactivo': qs = qs.filter(activo=False)
        writer.writerow(['Nombre', 'Documento', 'Email', 'Equipo', 'Categoria', 'Posicion', 'Camiseta', 'Estado'])
        for j in qs:
            writer.writerow([
                j.usuario.get_full_name() or j.usuario.username,
                j.usuario.documento or '',
                j.usuario.email,
                j.equipo.nombre if j.equipo else '',
                j.equipo.categoria.nombre if j.equipo else '',
                j.posicion or '', j.numero_camiseta or '',
                'Activo' if j.activo else 'Inactivo',
            ])

    elif modulo == 'asistencia':
        equipo_id = request.GET.get('equipo', '')
        jug_id    = request.GET.get('jugador', '')
        tipo      = request.GET.get('tipo', 'equipo')
        qs = Asistencia.objects.select_related('jugador__usuario', 'entrenamiento__equipo')
        if tipo == 'deportista' and jug_id:
            qs = qs.filter(jugador_id=jug_id)
        elif equipo_id:
            qs = qs.filter(entrenamiento__equipo_id=equipo_id)
        writer.writerow(['Fecha', 'Entrenamiento', 'Equipo', 'Deportista', 'Documento', 'Estado', 'Observacion'])
        for a in qs.order_by('-entrenamiento__fecha'):
            writer.writerow([
                a.entrenamiento.fecha.strftime('%d/%m/%Y'),
                a.entrenamiento.titulo,
                a.entrenamiento.equipo.nombre,
                a.jugador.usuario.get_full_name() or a.jugador.usuario.username,
                a.jugador.usuario.documento or '',
                a.get_estado_display(),
                a.observacion or '',
            ])

    elif modulo == 'reporte_entrenamientos':
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        buscar      = request.GET.get('buscar', '').strip()
        qs = Entrenamiento.objects.select_related('equipo__categoria', 'entrenador')
        if equipo_id:   qs = qs.filter(equipo_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        if buscar:      qs = qs.filter(Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar))
        writer.writerow(['Fecha', 'Horario', 'Titulo', 'Equipo', 'Categoria', 'Entrenador', 'Cancha'])
        for e in qs.order_by('-fecha'):
            writer.writerow([
                e.fecha.strftime('%d/%m/%Y'),
                f'{e.hora_inicio.strftime("%H:%M")} - {e.hora_fin.strftime("%H:%M")}',
                e.titulo, e.equipo.nombre, e.equipo.categoria.nombre,
                e.entrenador.get_full_name() if e.entrenador else '',
                e.lugar or '',
            ])

    elif modulo == 'reporte_partidos':
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        qs = PartidoCalendario.objects.select_related('equipo_propio')
        if equipo_id:   qs = qs.filter(equipo_propio_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        writer.writerow(['Fecha', 'Hora', 'Equipo Propio', 'Rival', 'Goles Favor', 'Goles Contra', 'Cancha'])
        for p in qs.order_by('-fecha'):
            writer.writerow([
                p.fecha.strftime('%d/%m/%Y'), p.hora.strftime('%H:%M'),
                p.equipo_propio.nombre, p.equipo_rival,
                p.goles_favor if p.goles_favor is not None else '',
                p.goles_contra if p.goles_contra is not None else '',
                p.cancha or '',
            ])

    elif modulo in ('pagos', 'reporte_financiero'):
        _marcar_vencidos()
        qs = Pago.objects.select_related('jugador__usuario', 'jugador__equipo')
        buscar      = request.GET.get('buscar', '').strip()
        estado      = request.GET.get('estado', '')
        concepto    = request.GET.get('concepto', '')
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if buscar:
            qs = qs.filter(
                Q(descripcion__icontains=buscar) |
                Q(jugador__usuario__first_name__icontains=buscar) |
                Q(jugador__usuario__last_name__icontains=buscar) |
                Q(jugador__usuario__documento__icontains=buscar)
            )
        if estado:      qs = qs.filter(estado=estado)
        if concepto:    qs = qs.filter(concepto=concepto)
        if equipo_id:   qs = qs.filter(jugador__equipo_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)
        writer.writerow(['Deportista', 'Documento', 'Equipo', 'Concepto', 'Descripcion',
                         'Monto', 'Vencimiento', 'Fecha Pago', 'Metodo', 'Estado'])
        for p in qs.order_by('-fecha_vencimiento'):
            writer.writerow([
                p.jugador.usuario.get_full_name() or p.jugador.usuario.username,
                p.jugador.usuario.documento or '',
                p.jugador.equipo.nombre if p.jugador.equipo else '',
                p.get_concepto_display(), p.descripcion, p.monto,
                p.fecha_vencimiento.strftime('%d/%m/%Y'),
                p.fecha_pago.strftime('%d/%m/%Y') if p.fecha_pago else '',
                p.get_metodo_pago_display() if p.metodo_pago else '',
                p.get_estado_display(),
            ])

    elif modulo == 'finanzas':
        qs = Finanza.objects.select_related('registrado_por')
        buscar      = request.GET.get('buscar', '').strip()
        tipo        = request.GET.get('tipo', '')
        categoria   = request.GET.get('categoria', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if buscar:      qs = qs.filter(descripcion__icontains=buscar)
        if tipo:        qs = qs.filter(tipo=tipo)
        if categoria:   qs = qs.filter(categoria=categoria)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        writer.writerow(['Descripcion', 'Tipo', 'Categoria', 'Monto', 'Fecha', 'Registrado Por'])
        for f in qs.order_by('-fecha'):
            writer.writerow([
                f.descripcion, f.tipo, f.get_categoria_display(),
                f.monto, f.fecha.strftime('%d/%m/%Y'),
                f.registrado_por.username if f.registrado_por else '',
            ])

    else:
        raise Http404

    return response


# ─────────────────────────────────────────────
# ESTADISTICAS DE RENDIMIENTO
# ─────────────────────────────────────────────

@solo_admin
def admin_estadisticas_rendimiento(request):
    """Estadisticas de rendimiento: promedios, evolucion y comparacion entre deportistas (admin)."""
    equipos      = Equipo.objects.filter(activo=True).order_by('nombre')
    categorias   = Categoria.objects.filter(activo=True).order_by('nombre')
    equipo_id    = request.GET.get('equipo', '')
    categoria_id = request.GET.get('categoria', '')
    tipo_eval    = request.GET.get('tipo', '')
    fecha_desde  = request.GET.get('fecha_desde', '')
    fecha_hasta  = request.GET.get('fecha_hasta', '')

    jugadores_qs = Jugador.objects.filter(activo=True).select_related('usuario', 'equipo__categoria')
    if equipo_id:    jugadores_qs = jugadores_qs.filter(equipo_id=equipo_id)
    if categoria_id: jugadores_qs = jugadores_qs.filter(equipo__categoria_id=categoria_id)

    ev_qs = Evaluacion.objects.filter(puntaje__isnull=False)
    if equipo_id:    ev_qs = ev_qs.filter(jugador__equipo_id=equipo_id)
    if categoria_id: ev_qs = ev_qs.filter(jugador__equipo__categoria_id=categoria_id)
    if tipo_eval:    ev_qs = ev_qs.filter(tipo=tipo_eval)
    if fecha_desde:  ev_qs = ev_qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:  ev_qs = ev_qs.filter(fecha__lte=fecha_hasta)

    # Promedio general y por tipo
    promedio_general = ev_qs.aggregate(avg=Avg('puntaje'))['avg']
    promedios_tipo   = (
        ev_qs.values('tipo')
        .annotate(avg=Avg('puntaje'), cantidad=Count('id'))
        .order_by('tipo')
    )

    # Comparacion: promedio por deportista (top 10)
    comparacion = (
        ev_qs.values('jugador__id', 'jugador__usuario__first_name', 'jugador__usuario__last_name',
                     'jugador__equipo__nombre')
        .annotate(avg=Avg('puntaje'), total_ev=Count('id'))
        .order_by('-avg')[:15]
    )

    # Evolucion: evaluaciones agrupadas por mes
    from django.db.models.functions import TruncMonth
    evolucion = (
        ev_qs
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(avg=Avg('puntaje'), cantidad=Count('id'))
        .order_by('mes')
    )

    evo_labels  = json.dumps([row['mes'].strftime('%b %Y') for row in evolucion])
    evo_valores = json.dumps([float(row['avg']) for row in evolucion])

    comp_nombres = json.dumps([
        f"{r['jugador__usuario__first_name']} {r['jugador__usuario__last_name']}".strip() or str(r['jugador__id'])
        for r in comparacion
    ])
    comp_valores = json.dumps([float(r['avg']) for r in comparacion])

    return render(request, 'panel/reportes/estadisticas_rendimiento.html', {
        'equipos':          equipos,
        'categorias':       categorias,
        'equipo_id':        equipo_id,
        'categoria_id':     categoria_id,
        'tipo_eval':        tipo_eval,
        'fecha_desde':      fecha_desde,
        'fecha_hasta':      fecha_hasta,
        'tipos':            Evaluacion.TIPO_CHOICES,
        'promedio_general': promedio_general,
        'promedios_tipo':   promedios_tipo,
        'comparacion':      comparacion,
        'evolucion':        evolucion,
        'evo_labels':       evo_labels,
        'evo_valores':      evo_valores,
        'comp_nombres':     comp_nombres,
        'comp_valores':     comp_valores,
        'total_evaluaciones': ev_qs.count(),
    })


@solo_entrenador
def ent_estadisticas_rendimiento(request):
    """Estadisticas de rendimiento: promedios, evolucion y comparacion entre deportistas (entrenador)."""
    mis_equipos  = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id    = request.GET.get('equipo', '')
    tipo_eval    = request.GET.get('tipo', '')
    fecha_desde  = request.GET.get('fecha_desde', '')
    fecha_hasta  = request.GET.get('fecha_hasta', '')

    ev_qs = Evaluacion.objects.filter(jugador__equipo__in=mis_equipos, puntaje__isnull=False)
    if equipo_id:   ev_qs = ev_qs.filter(jugador__equipo_id=equipo_id)
    if tipo_eval:   ev_qs = ev_qs.filter(tipo=tipo_eval)
    if fecha_desde: ev_qs = ev_qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta: ev_qs = ev_qs.filter(fecha__lte=fecha_hasta)

    promedio_general = ev_qs.aggregate(avg=Avg('puntaje'))['avg']
    promedios_tipo   = (
        ev_qs.values('tipo')
        .annotate(avg=Avg('puntaje'), cantidad=Count('id'))
        .order_by('tipo')
    )

    comparacion = (
        ev_qs.values('jugador__id', 'jugador__usuario__first_name', 'jugador__usuario__last_name',
                     'jugador__equipo__nombre')
        .annotate(avg=Avg('puntaje'), total_ev=Count('id'))
        .order_by('-avg')[:15]
    )

    from django.db.models.functions import TruncMonth
    evolucion = (
        ev_qs
        .annotate(mes=TruncMonth('fecha'))
        .values('mes')
        .annotate(avg=Avg('puntaje'), cantidad=Count('id'))
        .order_by('mes')
    )

    evo_labels  = json.dumps([row['mes'].strftime('%b %Y') for row in evolucion])
    evo_valores = json.dumps([float(row['avg']) for row in evolucion])
    comp_nombres = json.dumps([
        f"{r['jugador__usuario__first_name']} {r['jugador__usuario__last_name']}".strip() or str(r['jugador__id'])
        for r in comparacion
    ])
    comp_valores = json.dumps([float(r['avg']) for r in comparacion])

    return render(request, 'entrenador/reportes/estadisticas_rendimiento.html', {
        'equipos':          mis_equipos,
        'equipo_id':        equipo_id,
        'tipo_eval':        tipo_eval,
        'fecha_desde':      fecha_desde,
        'fecha_hasta':      fecha_hasta,
        'tipos':            Evaluacion.TIPO_CHOICES,
        'promedio_general': promedio_general,
        'promedios_tipo':   promedios_tipo,
        'comparacion':      comparacion,
        'evolucion':        evolucion,
        'evo_labels':       evo_labels,
        'evo_valores':      evo_valores,
        'comp_nombres':     comp_nombres,
        'comp_valores':     comp_valores,
        'total_evaluaciones': ev_qs.count(),
    })


# ─────────────────────────────────────────────
# EXPORTAR PDF (ADMIN)
# ─────────────────────────────────────────────

@solo_admin
def exportar_pdf(request, modulo):
    hoy = timezone.now()

    if modulo == 'usuarios':
        qs = UserModel.objects.all()
        buscar = request.GET.get('buscar', '').strip()
        rol    = request.GET.get('rol', '')
        estado = request.GET.get('estado', '')
        if buscar:
            qs = qs.filter(
                Q(username__icontains=buscar) | Q(first_name__icontains=buscar) |
                Q(last_name__icontains=buscar) | Q(email__icontains=buscar) |
                Q(documento__icontains=buscar)
            )
        if rol: qs = qs.filter(role=rol)
        if estado == 'activo':   qs = qs.filter(is_active=True)
        if estado == 'inactivo': qs = qs.filter(is_active=False)
        datos = {
            'titulo': 'Reporte de Usuarios', 'modulo': 'usuarios',
            'registros': qs.order_by('-date_joined'),
            'filtros': {'Buscar': buscar, 'Rol': rol, 'Estado': estado},
            'columnas': ['Username', 'Nombre', 'Email', 'Documento', 'Rol', 'Estado', 'Registro'],
        }

    elif modulo == 'pagos':
        _marcar_vencidos()
        qs = Pago.objects.select_related('jugador__usuario', 'jugador__equipo__categoria')
        buscar      = request.GET.get('buscar', '').strip()
        estado      = request.GET.get('estado', '')
        concepto    = request.GET.get('concepto', '')
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if buscar:
            qs = qs.filter(
                Q(descripcion__icontains=buscar) |
                Q(jugador__usuario__first_name__icontains=buscar) |
                Q(jugador__usuario__last_name__icontains=buscar) |
                Q(jugador__usuario__documento__icontains=buscar)
            )
        if estado:      qs = qs.filter(estado=estado)
        if concepto:    qs = qs.filter(concepto=concepto)
        if equipo_id:   qs = qs.filter(jugador__equipo_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)
        total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        datos = {
            'titulo': 'Reporte de Pagos', 'modulo': 'pagos',
            'registros': qs.order_by('-fecha_vencimiento'),
            'filtros': {'Buscar': buscar, 'Estado': estado, 'Concepto': concepto,
                        'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta},
            'columnas': ['Deportista', 'Equipo', 'Descripción', 'Concepto', 'Monto', 'Vencimiento', 'Método', 'Estado'],
            'total_pagado': total_pagado, 'total_pendiente': total_pendiente,
            'total_vencido': total_vencido, 'total_general': total_pagado + total_pendiente + total_vencido,
        }

    elif modulo == 'entrenamientos':
        qs = Entrenamiento.objects.select_related('equipo', 'entrenador')
        buscar      = request.GET.get('buscar', '').strip()
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if buscar:
            qs = qs.filter(
                Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar) |
                Q(descripcion__icontains=buscar)
            )
        if equipo_id:   qs = qs.filter(equipo_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        datos = {
            'titulo': 'Reporte de Entrenamientos', 'modulo': 'entrenamientos',
            'registros': qs.order_by('-fecha'),
            'filtros': {'Buscar': buscar, 'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta},
            'columnas': ['Título', 'Equipo', 'Entrenador', 'Fecha', 'Horario', 'Lugar'],
        }

    elif modulo == 'equipos':
        qs = Equipo.objects.select_related('categoria', 'entrenador').annotate(
            num_jugadores=Count('jugadores')).order_by('nombre')
        datos = {
            'titulo': 'Reporte de Equipos', 'modulo': 'equipos',
            'registros': qs, 'filtros': {},
            'columnas': ['Nombre', 'Categoría', 'Entrenador', 'Jugadores', 'Estado'],
        }

    elif modulo == 'torneos':
        qs = Torneo.objects.select_related('categoria').annotate(num_partidos=Count('partidos'))
        buscar       = request.GET.get('buscar', '').strip()
        categoria_id = request.GET.get('categoria', '')
        estado       = request.GET.get('estado', '')
        if buscar:
            qs = qs.filter(Q(nombre__icontains=buscar) | Q(lugar__icontains=buscar))
        if categoria_id: qs = qs.filter(categoria_id=categoria_id)
        if estado:       qs = qs.filter(estado=estado)
        datos = {
            'titulo': 'Reporte de Torneos', 'modulo': 'torneos',
            'registros': qs.order_by('-fecha_inicio'),
            'filtros': {'Buscar': buscar, 'Estado': estado},
            'columnas': ['Nombre', 'Categoría', 'Inicio', 'Fin', 'Estado', 'Partidos'],
        }

    elif modulo == 'finanzas':
        qs = Finanza.objects.select_related('registrado_por')
        buscar      = request.GET.get('buscar', '').strip()
        tipo        = request.GET.get('tipo', '')
        categoria   = request.GET.get('categoria', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if buscar:      qs = qs.filter(descripcion__icontains=buscar)
        if tipo:        qs = qs.filter(tipo=tipo)
        if categoria:   qs = qs.filter(categoria=categoria)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        total_ingresos = qs.filter(tipo='ingreso').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        total_egresos  = qs.filter(tipo='egreso').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        datos = {
            'titulo': 'Reporte Financiero', 'modulo': 'finanzas',
            'registros': qs.order_by('-fecha'),
            'filtros': {'Buscar': buscar, 'Tipo': tipo, 'Categoría': categoria,
                        'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta},
            'columnas': ['Descripción', 'Tipo', 'Categoría', 'Monto', 'Fecha', 'Registrado por'],
            'total_ingresos': total_ingresos, 'total_egresos': total_egresos,
            'balance': total_ingresos - total_egresos,
        }

    elif modulo == 'deportistas':
        qs = Jugador.objects.select_related(
            'usuario', 'equipo__categoria'
        ).order_by('usuario__last_name', 'usuario__first_name')
        buscar       = request.GET.get('buscar', '').strip()
        equipo_id    = request.GET.get('equipo', '')
        categoria_id = request.GET.get('categoria', '')
        estado       = request.GET.get('estado', '')
        if buscar:
            qs = qs.filter(
                Q(usuario__first_name__icontains=buscar) |
                Q(usuario__last_name__icontains=buscar)  |
                Q(usuario__documento__icontains=buscar)
            )
        if equipo_id:    qs = qs.filter(equipo_id=equipo_id)
        if categoria_id: qs = qs.filter(equipo__categoria_id=categoria_id)
        if estado == 'activo':   qs = qs.filter(activo=True)
        if estado == 'inactivo': qs = qs.filter(activo=False)
        datos = {
            'titulo': 'Reporte de Deportistas', 'modulo': 'deportistas',
            'registros': qs,
            'filtros': {'Buscar': buscar, 'Estado': estado},
            'columnas': ['Nombre', 'Documento', 'Equipo', 'Categoría', 'Posición', 'Camiseta', 'Estado'],
        }

    elif modulo == 'asistencia':
        equipo_id = request.GET.get('equipo', '')
        jug_id    = request.GET.get('jugador', '')
        tipo      = request.GET.get('tipo', 'equipo')
        qs = Asistencia.objects.select_related(
            'jugador__usuario', 'entrenamiento__equipo'
        )
        titulo_extra = 'General'
        if tipo == 'deportista' and jug_id:
            jugador = get_object_or_404(Jugador, id=jug_id)
            qs = qs.filter(jugador=jugador)
            titulo_extra = jugador.usuario.get_full_name() or jugador.usuario.username
        elif equipo_id:
            equipo = get_object_or_404(Equipo, id=equipo_id)
            qs = qs.filter(entrenamiento__equipo=equipo)
            titulo_extra = equipo.nombre
        qs = qs.order_by('-entrenamiento__fecha', 'jugador__usuario__last_name')
        datos = {
            'titulo': f'Reporte de Asistencia — {titulo_extra}', 'modulo': 'asistencia',
            'registros': qs,
            'filtros': {'Tipo': tipo},
            'columnas': ['Fecha', 'Entrenamiento', 'Equipo', 'Deportista', 'Documento', 'Estado', 'Observación'],
        }

    elif modulo == 'reporte_entrenamientos':
        qs = Entrenamiento.objects.select_related('equipo__categoria', 'entrenador')
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        buscar      = request.GET.get('buscar', '').strip()
        if equipo_id:   qs = qs.filter(equipo_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        if buscar:      qs = qs.filter(Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar))
        qs = qs.order_by('-fecha')
        datos = {
            'titulo': 'Reporte de Entrenamientos', 'modulo': 'reporte_entrenamientos',
            'registros': qs,
            'filtros': {'Buscar': buscar, 'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta},
            'columnas': ['Fecha', 'Horario', 'Título', 'Equipo', 'Categoría', 'Entrenador', 'Cancha'],
        }

    elif modulo == 'reporte_partidos':
        qs = PartidoCalendario.objects.select_related(
            'equipo_propio__categoria'
        ).prefetch_related('estadisticas__jugador__usuario')
        equipo_id   = request.GET.get('equipo', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        if equipo_id:   qs = qs.filter(equipo_propio_id=equipo_id)
        if fecha_desde: qs = qs.filter(fecha__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha__lte=fecha_hasta)
        qs = qs.order_by('-fecha')
        datos = {
            'titulo': 'Reporte de Partidos', 'modulo': 'reporte_partidos',
            'registros': qs,
            'filtros': {'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta},
            'columnas': ['Fecha', 'Hora', 'Equipo Propio', 'Rival', 'Resultado', 'Cancha', 'Estadísticas'],
        }

    elif modulo == 'reporte_financiero':
        _marcar_vencidos()
        qs = Pago.objects.select_related('jugador__usuario', 'jugador__equipo__categoria', 'registrado_por')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')
        estado      = request.GET.get('estado', '')
        equipo_id   = request.GET.get('equipo', '')
        concepto    = request.GET.get('concepto', '')
        if fecha_desde: qs = qs.filter(fecha_vencimiento__gte=fecha_desde)
        if fecha_hasta: qs = qs.filter(fecha_vencimiento__lte=fecha_hasta)
        if estado:      qs = qs.filter(estado=estado)
        if equipo_id:   qs = qs.filter(jugador__equipo_id=equipo_id)
        if concepto:    qs = qs.filter(concepto=concepto)
        qs = qs.order_by('-fecha_vencimiento')
        total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
        datos = {
            'titulo': 'Reporte Financiero', 'modulo': 'reporte_financiero',
            'registros': qs,
            'filtros': {'Fecha desde': fecha_desde, 'Fecha hasta': fecha_hasta,
                        'Estado': estado, 'Concepto': concepto},
            'columnas': ['Deportista', 'Equipo', 'Concepto', 'Descripción', 'Monto',
                         'Vencimiento', 'Fecha Pago', 'Método', 'Estado'],
            'total_pagado': total_pagado, 'total_pendiente': total_pendiente,
            'total_vencido': total_vencido,
            'total_general': total_pagado + total_pendiente + total_vencido,
        }

    else:
        raise Http404

    datos['fecha_generacion'] = hoy
    datos['generado_por']     = request.user.get_full_name() or request.user.username
    datos['total_registros']  = datos['registros'].count()

    html = render_to_string('panel/pdf/base_pdf.html', datos)
    buffer = BytesIO()
    pisa.CreatePDF(html.encode('utf-8'), dest=buffer)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{modulo}_{hoy.strftime("%Y%m%d_%H%M")}.pdf"'
    response['Content-Length'] = len(pdf)
    return response


# ─────────────────────────────────────────────
# ENVÍO MASIVO DE CORREOS (ADMIN)
# ─────────────────────────────────────────────

@solo_admin
def admin_enviar_correo_usuario(request):
    """Enviar mensaje a un usuario especifico (admin)."""
    usuarios = UserModel.objects.filter(is_active=True).exclude(role='pendiente').order_by('last_name', 'first_name')

    if request.method == 'POST':
        usuario_id = request.POST.get('usuario')
        asunto     = request.POST.get('asunto', '').strip()
        mensaje    = request.POST.get('mensaje', '').strip()

        errores = []
        if not usuario_id: errores.append('Selecciona un destinatario.')
        if not asunto:     errores.append('El asunto es obligatorio.')
        if not mensaje:    errores.append('El mensaje es obligatorio.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            usuario = get_object_or_404(UserModel, id=usuario_id)
            Notificacion.objects.create(usuario=usuario, asunto=asunto, mensaje=mensaje, emisor=request.user)
            # Intentar enviar correo
            if usuario.email and '@' in usuario.email:
                try:
                    from django.core.mail import send_mail
                    send_mail(
                        subject=asunto,
                        message=mensaje,
                        from_email=django_settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[usuario.email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
            messages.success(request, f'Mensaje enviado a {usuario.get_full_name() or usuario.username} correctamente.')

    return render(request, 'panel/correos/form_usuario.html', {
        'usuarios': usuarios,
        'post_data': request.POST if request.method == 'POST' else {},
    })


@solo_entrenador
def ent_enviar_mensaje_grupo(request):
    """Enviar mensaje a todos los deportistas de un equipo propio (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True).select_related('categoria')

    if request.method == 'POST':
        equipo_id = request.POST.get('equipo_id')
        asunto    = request.POST.get('asunto', '').strip()
        mensaje   = request.POST.get('mensaje', '').strip()

        errores = []
        if not equipo_id: errores.append('Selecciona un equipo.')
        if not asunto:    errores.append('El asunto es obligatorio.')
        if not mensaje:   errores.append('El mensaje es obligatorio.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            equipo = get_object_or_404(Equipo, id=equipo_id, entrenador=request.user)
            usuarios = UserModel.objects.filter(
                jugador__equipo=equipo, is_active=True
            ).distinct()
            if not usuarios.exists():
                messages.error(request, 'El equipo no tiene deportistas activos.')
            else:
                Notificacion.objects.bulk_create([
                    Notificacion(usuario=u, asunto=asunto, mensaje=mensaje, emisor=request.user)
                    for u in usuarios
                ])
                for u in usuarios:
                    if u.email and '@' in u.email:
                        try:
                            send_mail(
                                subject=asunto,
                                message=mensaje,
                                from_email=django_settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[u.email],
                                fail_silently=True,
                            )
                        except Exception:
                            pass
                messages.success(request, f'Mensaje enviado a {usuarios.count()} deportista(s) del equipo {equipo.nombre}.')

    return render(request, 'entrenador/mensajes_grupo.html', {
        'mis_equipos': mis_equipos,
        'post_data': request.POST if request.method == 'POST' else {},
    })


@solo_entrenador
def ent_enviar_mensaje(request):
    """Enviar mensaje a un usuario especifico (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    # El entrenador puede escribir a sus deportistas y al admin
    destinatarios = UserModel.objects.filter(
        Q(role='administrador') |
        Q(jugador__equipo__in=mis_equipos)
    ).filter(is_active=True).distinct().order_by('last_name', 'first_name')

    if request.method == 'POST':
        usuario_id = request.POST.get('usuario')
        asunto     = request.POST.get('asunto', '').strip()
        mensaje    = request.POST.get('mensaje', '').strip()

        errores = []
        if not usuario_id: errores.append('Selecciona un destinatario.')
        if not asunto:     errores.append('El asunto es obligatorio.')
        if not mensaje:    errores.append('El mensaje es obligatorio.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            usuario = get_object_or_404(UserModel, id=usuario_id)
            Notificacion.objects.create(usuario=usuario, asunto=asunto, mensaje=mensaje, emisor=request.user)
            if usuario.email and '@' in usuario.email:
                try:
                    from django.core.mail import send_mail
                    send_mail(
                        subject=asunto,
                        message=mensaje,
                        from_email=django_settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[usuario.email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
            messages.success(request, f'Mensaje enviado a {usuario.get_full_name() or usuario.username} correctamente.')

    return render(request, 'entrenador/mensajes_enviar.html', {
        'destinatarios': destinatarios,
        'post_data': request.POST if request.method == 'POST' else {},
    })


@solo_admin
def admin_enviar_correo(request):
    roles_opciones = [
        ('todos', 'Todos los usuarios'),
        ('administrador', 'Administradores'),
        ('entrenador', 'Entrenadores'),
        ('deportista', 'Deportistas'),
    ]
    equipos    = Equipo.objects.filter(activo=True).select_related('categoria').order_by('nombre')
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')

    if request.method == 'POST':
        asunto            = request.POST.get('asunto', '').strip()
        mensaje           = request.POST.get('mensaje', '').strip()
        tipo_destinatario = request.POST.get('tipo_destinatario', 'rol')
        destinatario      = request.POST.get('destinatario', 'todos')
        equipo_id         = request.POST.get('equipo_id', '')
        categoria_id      = request.POST.get('categoria_id', '')

        ctx_error = {
            'roles_opciones': roles_opciones,
            'equipos': equipos,
            'categorias': categorias,
            'post_data': request.POST,
        }

        if not asunto or not mensaje:
            messages.error(request, 'El asunto y el mensaje son obligatorios.')
            return render(request, 'panel/correos/form.html', ctx_error)

        qs = UserModel.objects.filter(is_active=True).exclude(role='pendiente')

        if tipo_destinatario == 'equipo':
            if not equipo_id:
                messages.error(request, 'Selecciona un equipo.')
                return render(request, 'panel/correos/form.html', ctx_error)
            qs = qs.filter(jugador__equipo_id=equipo_id)
        elif tipo_destinatario == 'categoria':
            if not categoria_id:
                messages.error(request, 'Selecciona una categoría.')
                return render(request, 'panel/correos/form.html', ctx_error)
            qs = qs.filter(jugador__equipo__categoria_id=categoria_id)
        else:
            if destinatario != 'todos':
                qs = qs.filter(role=destinatario)

        usuarios = list(qs.distinct())
        if not usuarios:
            messages.error(request, 'No hay usuarios para ese grupo.')
            return render(request, 'panel/correos/form.html', ctx_error)

        Notificacion.objects.bulk_create([
            Notificacion(usuario=u, asunto=asunto, mensaje=mensaje, emisor=request.user)
            for u in usuarios
        ])

        emails = [u.email for u in usuarios if u.email and '@' in u.email]
        enviados = 0
        error_smtp = None
        if emails:
            try:
                connection = get_connection(
                    backend='django.core.mail.backends.smtp.EmailBackend',
                    fail_silently=False,
                )
                connection.open()
                for email in emails:
                    try:
                        send_mail(
                            subject=asunto,
                            message=mensaje,
                            from_email=django_settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[email],
                            connection=connection,
                            fail_silently=False,
                        )
                        enviados += 1
                    except Exception:
                        pass
                connection.close()
            except Exception as e:
                error_smtp = str(e)

        if error_smtp:
            messages.warning(
                request,
                f'Notificacion guardada para {len(usuarios)} usuario(s). '
                f'El correo no se pudo enviar ({error_smtp}). '
                f'Los usuarios veran el mensaje al ingresar al sistema.'
            )
        else:
            messages.success(
                request,
                f'Correos enviados: {enviados} de {len(emails)}. '
                f'Notificacion guardada para {len(usuarios)} usuario(s).'
            )

    return render(request, 'panel/correos/form.html', {
        'roles_opciones': roles_opciones,
        'equipos': equipos,
        'categorias': categorias,
    })


# ─────────────────────────────────────────────
# DASHBOARDS DE OTROS ROLES
# ─────────────────────────────────────────────

@login_required
def entrenador_dashboard(request):
    """Dashboard principal del entrenador con estadísticas propias"""
    if request.user.role != 'entrenador':
        return HttpResponseForbidden("Sin permisos.")

    mis_equipos       = Equipo.objects.filter(entrenador=request.user, activo=True)
    total_jugadores   = Jugador.objects.filter(equipo__in=mis_equipos, activo=True).count()
    total_entrenamientos = Entrenamiento.objects.filter(entrenador=request.user).count()
    proximos          = Entrenamiento.objects.filter(
        entrenador=request.user
    ).order_by('fecha')[:5]

    return render(request, 'entrenador/dashboard.html', {
        'mis_equipos':          mis_equipos,
        'total_equipos':        mis_equipos.count(),
        'total_jugadores':      total_jugadores,
        'total_entrenamientos': total_entrenamientos,
        'proximos':             proximos,
        'notificaciones':       Notificacion.objects.filter(usuario=request.user, leida=False).order_by('-creado'),
    })


# ── Decorador solo entrenador ────────────────────────────────────────────────


# ── Entrenamientos del entrenador ────────────────────────────────────────────

@solo_entrenador
def ent_entrenamientos_lista(request):
    """Lista de entrenamientos propios con filtros"""
    qs = Entrenamiento.objects.filter(entrenador=request.user).select_related('equipo')

    buscar      = request.GET.get('buscar', '').strip()
    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar))
    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    paginator = Paginator(qs.order_by('-fecha'), 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'entrenador/entrenamientos/lista.html', {
        'entrenamientos': page,
        'buscar': buscar,
        'equipo_id': equipo_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'equipos': Equipo.objects.filter(entrenador=request.user, activo=True),
        'total_filtrados': qs.count(),
    })


@solo_entrenador
def ent_entrenamiento_crear(request):
    """Crear entrenamiento asignado al entrenador actual"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)

    if request.method == 'POST':
        titulo      = request.POST.get('titulo', '').strip()
        equipo_id   = request.POST.get('equipo')
        fecha       = request.POST.get('fecha')
        hora_inicio = request.POST.get('hora_inicio')
        hora_fin    = request.POST.get('hora_fin')
        lugar       = request.POST.get('lugar', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()

        if not all([titulo, equipo_id, fecha, hora_inicio, hora_fin]):
            messages.error(request, 'Completa todos los campos obligatorios.')
        elif not mis_equipos.filter(id=equipo_id).exists():
            messages.error(request, 'No tienes permiso sobre ese equipo.')
        else:
            ent = Entrenamiento.objects.create(
                titulo=titulo, equipo_id=equipo_id,
                entrenador=request.user,
                fecha=fecha, hora_inicio=hora_inicio, hora_fin=hora_fin,
                lugar=lugar, descripcion=descripcion
            )
            _notificar_entrenamiento(ent, es_nuevo=True, emisor=request.user)
            messages.success(request, f'Entrenamiento "{titulo}" creado correctamente.')
            return ent_entrenamientos_lista(request)

    return render(request, 'entrenador/entrenamientos/form.html', {
        'titulo': 'Nuevo Entrenamiento',
        'equipos': mis_equipos,
    })


@solo_entrenador
def ent_entrenamiento_editar(request, ent_id):
    """Editar entrenamiento propio"""
    entrenamiento = get_object_or_404(Entrenamiento, id=ent_id, entrenador=request.user)
    mis_equipos   = Equipo.objects.filter(entrenador=request.user, activo=True)

    if request.method == 'POST':
        entrenamiento.titulo      = request.POST.get('titulo', '').strip()
        entrenamiento.equipo_id   = request.POST.get('equipo')
        entrenamiento.fecha       = request.POST.get('fecha')
        entrenamiento.hora_inicio = request.POST.get('hora_inicio')
        entrenamiento.hora_fin    = request.POST.get('hora_fin')
        entrenamiento.lugar       = request.POST.get('lugar', '').strip()
        entrenamiento.descripcion = request.POST.get('descripcion', '').strip()
        entrenamiento.save()
        _notificar_entrenamiento(entrenamiento, es_nuevo=False, emisor=request.user)
        messages.success(request, f'Entrenamiento "{entrenamiento.titulo}" actualizado.')

    return render(request, 'entrenador/entrenamientos/form.html', {
        'titulo': 'Editar Entrenamiento',
        'entrenamiento': entrenamiento,
        'equipos': mis_equipos,
    })


@solo_entrenador
def ent_entrenamiento_eliminar(request, ent_id):
    """Eliminar entrenamiento propio"""
    entrenamiento = get_object_or_404(Entrenamiento, id=ent_id, entrenador=request.user)
    nombre = entrenamiento.titulo
    entrenamiento.delete()
    messages.success(request, f'Entrenamiento "{nombre}" eliminado.')
    return ent_entrenamientos_lista(request)


# ── Equipos del entrenador ───────────────────────────────────────────────────

@solo_entrenador
def ent_equipos_lista(request):
    """Lista de equipos asignados al entrenador"""
    equipos = Equipo.objects.filter(
        entrenador=request.user, activo=True
    ).annotate(num_jugadores=Count('jugadores'))

    return render(request, 'entrenador/equipos/lista.html', {'equipos': equipos})


@solo_entrenador
def ent_equipo_detalle(request, equipo_id):
    """Detalle de un equipo: jugadores y entrenamientos.
    Logica cruzada:
    - Solo muestra deportistas activos
    - Al asignar, crea o actualiza el perfil Jugador
    - Al quitar, desvincula sin eliminar el perfil
    """
    equipo         = get_object_or_404(Equipo, id=equipo_id, entrenador=request.user)
    jugadores      = Jugador.objects.filter(equipo=equipo).select_related('usuario')
    entrenamientos = Entrenamiento.objects.filter(equipo=equipo).order_by('-fecha')[:10]

    # Deportistas activos que NO están ya en este equipo
    ids_en_equipo = jugadores.values_list('usuario_id', flat=True)
    disponibles   = UserModel.objects.filter(
        role='deportista', is_active=True
    ).exclude(id__in=ids_en_equipo)

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'asignar_jugador':
            user_id = request.POST.get('deportista_id')
            usuario = get_object_or_404(UserModel, id=user_id, role='deportista', is_active=True)

            # Verificar si ya tiene equipo asignado
            jugador_existente = Jugador.objects.filter(usuario=usuario).first()
            if jugador_existente and jugador_existente.equipo and jugador_existente.equipo != equipo:
                messages.warning(
                    request,
                    f'{usuario.get_full_name() or usuario.username} ya pertenece al equipo '
                    f'"{jugador_existente.equipo.nombre}". Se reasignó a "{equipo.nombre}".')

            jugador, _ = Jugador.objects.get_or_create(
                usuario=usuario,
                defaults={'equipo': equipo, 'activo': True}
            )
            jugador.equipo = equipo
            jugador.activo = True
            jugador.save()
            messages.success(
                request,
                f'{usuario.get_full_name() or usuario.username} asignado a "{equipo.nombre}" '
                f'(Categoría: {equipo.categoria.nombre}).')

        elif accion == 'quitar_jugador':
            jug_id  = request.POST.get('jugador_id')
            jugador = get_object_or_404(Jugador, id=jug_id, equipo=equipo)
            nombre  = jugador.usuario.get_full_name() or jugador.usuario.username
            jugador.equipo = None
            jugador.save()
            messages.success(request, f'{nombre} removido del equipo (perfil conservado).')

        return ent_equipo_detalle(request, equipo_id)

    return render(request, 'entrenador/equipos/detalle.html', {
        'equipo':         equipo,
        'jugadores':      jugadores,
        'entrenamientos': entrenamientos,
        'disponibles':    disponibles,
    })


# ── Jugadores del entrenador ─────────────────────────────────────────────────

@solo_entrenador
def ent_jugadores_lista(request):
    """Lista de jugadores en los equipos del entrenador con filtros"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = Jugador.objects.filter(equipo__in=mis_equipos).select_related('usuario', 'equipo')

    buscar    = request.GET.get('buscar', '').strip()
    equipo_id = request.GET.get('equipo', '')

    if buscar:
        qs = qs.filter(
            Q(usuario__first_name__icontains=buscar) |
            Q(usuario__last_name__icontains=buscar)  |
            Q(usuario__email__icontains=buscar)       |
            Q(posicion__icontains=buscar)
        )
    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)

    return render(request, 'entrenador/jugadores/lista.html', {
        'jugadores':       qs,
        'buscar':          buscar,
        'equipo_id':       equipo_id,
        'equipos':         mis_equipos,
        'total_filtrados': qs.count(),
    })


@solo_entrenador
def ent_jugador_crear(request):
    """Crear perfil de jugador para un deportista existente"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    # Deportistas sin perfil de jugador aún
    sin_perfil = UserModel.objects.filter(
        role='deportista', is_active=True
    ).exclude(jugador__isnull=False)

    if request.method == 'POST':
        usuario_id      = request.POST.get('usuario')
        equipo_id       = request.POST.get('equipo')
        posicion        = request.POST.get('posicion', '').strip()
        numero_camiseta = request.POST.get('numero_camiseta') or None

        if not usuario_id:
            messages.error(request, 'Selecciona un deportista.')
        elif Jugador.objects.filter(usuario_id=usuario_id).exists():
            messages.error(request, 'Este deportista ya tiene perfil de jugador.')
        elif equipo_id and not mis_equipos.filter(id=equipo_id).exists():
            messages.error(request, 'No tienes permiso sobre ese equipo.')
        else:
            Jugador.objects.create(
                usuario_id=usuario_id,
                equipo_id=equipo_id or None,
                posicion=posicion,
                numero_camiseta=numero_camiseta,
                activo=True
            )
            messages.success(request, 'Jugador creado correctamente.')
            return ent_jugadores_lista(request)

    return render(request, 'entrenador/jugadores/form.html', {
        'titulo':     'Nuevo Jugador',
        'equipos':    mis_equipos,
        'sin_perfil': sin_perfil,
    })


@solo_entrenador
def ent_jugador_editar(request, jug_id):
    """Editar jugador de los equipos del entrenador"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugador     = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)

    if request.method == 'POST':
        equipo_id       = request.POST.get('equipo')
        posicion        = request.POST.get('posicion', '').strip()
        numero_camiseta = request.POST.get('numero_camiseta') or None

        if equipo_id and not mis_equipos.filter(id=equipo_id).exists():
            messages.error(request, 'No tienes permiso sobre ese equipo.')
        else:
            jugador.equipo_id       = equipo_id or None
            jugador.posicion        = posicion
            jugador.numero_camiseta = numero_camiseta
            jugador.activo          = request.POST.get('activo') == 'on'
            jugador.save()
            messages.success(request, f'Jugador "{jugador.usuario.get_full_name() or jugador.usuario.username}" actualizado.')

    return render(request, 'entrenador/jugadores/form.html', {
        'titulo':  'Editar Jugador',
        'jugador': jugador,
        'equipos': mis_equipos,
    })


@solo_entrenador
def ent_jugador_eliminar(request, jug_id):
    """Eliminar jugador de los equipos del entrenador"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugador     = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)
    nombre      = jugador.usuario.get_full_name() or jugador.usuario.username
    jugador.delete()
    messages.success(request, f'Jugador "{nombre}" eliminado.')
    return ent_jugadores_lista(request)


# ── Evaluaciones del entrenador ──────────────────────────────────────────────

# ── Observaciones del entrenador ──────────────────────────────────────────────────

@solo_entrenador
def ent_observaciones_lista(request):
    """Lista de observaciones del entrenador con filtros."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = ObservacionDeportista.objects.filter(
        entrenador=request.user
    ).select_related('jugador__usuario', 'jugador__equipo')

    jugador_id  = request.GET.get('jugador', '')
    tipo        = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if jugador_id:
        qs = qs.filter(jugador_id=jugador_id)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    jugadores_filtro = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario').order_by('usuario__last_name')

    paginator = Paginator(qs.order_by('-fecha'), 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'entrenador/observaciones/lista.html', {
        'observaciones':   page,
        'jugador_id':      jugador_id,
        'tipo':            tipo,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'jugadores_filtro': jugadores_filtro,
        'tipos':           ObservacionDeportista.TIPO_CHOICES,
        'total_filtrados': qs.count(),
    })


@solo_entrenador
def ent_observacion_crear(request):
    """Crear observación para un jugador propio."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugadores   = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario', 'equipo').order_by('usuario__last_name')

    if request.method == 'POST':
        jugador_id  = request.POST.get('jugador')
        tipo        = request.POST.get('tipo', 'tecnica')
        fecha       = request.POST.get('fecha')
        descripcion = request.POST.get('descripcion', '').strip()

        errores = []
        if not jugador_id:  errores.append('El deportista es obligatorio.')
        elif not jugadores.filter(id=jugador_id).exists():
            errores.append('No tenés permiso sobre ese deportista.')
        if not fecha:       errores.append('La fecha es obligatoria.')
        if not descripcion: errores.append('La descripción es obligatoria.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            ObservacionDeportista.objects.create(
                jugador_id=jugador_id,
                entrenador=request.user,
                tipo=tipo,
                fecha=fecha,
                descripcion=descripcion,
            )
            messages.success(request, 'Observación registrada correctamente.')
            return ent_observaciones_lista(request)

    return render(request, 'entrenador/observaciones/form.html', {
        'titulo':    'Nueva Observación',
        'jugadores': jugadores,
        'tipos':     ObservacionDeportista.TIPO_CHOICES,
    })


@solo_entrenador
def ent_observacion_editar(request, obs_id):
    """Editar descripción y fecha de una observación propia."""
    mis_equipos  = Equipo.objects.filter(entrenador=request.user, activo=True)
    observacion  = get_object_or_404(ObservacionDeportista, id=obs_id, entrenador=request.user)
    jugadores    = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario', 'equipo').order_by('usuario__last_name')

    if request.method == 'POST':
        descripcion = request.POST.get('descripcion', '').strip()
        fecha       = request.POST.get('fecha')
        tipo        = request.POST.get('tipo', observacion.tipo)

        errores = []
        if not fecha:       errores.append('La fecha es obligatoria.')
        if not descripcion: errores.append('La descripción es obligatoria.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            observacion.fecha       = fecha
            observacion.tipo        = tipo
            observacion.descripcion = descripcion
            observacion.save()
            messages.success(request, 'Observación actualizada correctamente.')
            return ent_observaciones_lista(request)

    return render(request, 'entrenador/observaciones/form.html', {
        'titulo':      'Editar Observación',
        'observacion': observacion,
        'jugadores':   jugadores,
        'tipos':       ObservacionDeportista.TIPO_CHOICES,
    })


@solo_entrenador
def ent_observacion_eliminar(request, obs_id):
    """Eliminar observación propia."""
    observacion = get_object_or_404(ObservacionDeportista, id=obs_id, entrenador=request.user)
    observacion.delete()
    messages.success(request, 'Observación eliminada correctamente.')
    return ent_observaciones_lista(request)


# ── Reporte de rendimiento PDF (admin) ──────────────────────────────────

@solo_admin
def admin_reporte_deportistas(request):
    """Reporte de deportistas con filtros por nombre, documento, categoría, equipo y estado."""
    qs = Jugador.objects.select_related(
        'usuario', 'equipo__categoria'
    ).order_by('usuario__last_name', 'usuario__first_name')

    buscar       = request.GET.get('buscar', '').strip()
    equipo_id    = request.GET.get('equipo', '')
    categoria_id = request.GET.get('categoria', '')
    estado       = request.GET.get('estado', '')

    if buscar:
        qs = qs.filter(
            Q(usuario__first_name__icontains=buscar) |
            Q(usuario__last_name__icontains=buscar)  |
            Q(usuario__documento__icontains=buscar)
        )
    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)
    if categoria_id:
        qs = qs.filter(equipo__categoria_id=categoria_id)
    if estado == 'activo':
        qs = qs.filter(activo=True)
    elif estado == 'inactivo':
        qs = qs.filter(activo=False)

    # Construir query string para el PDF preservando filtros
    params = request.GET.urlencode()

    return render(request, 'panel/reportes/deportistas.html', {
        'deportistas':     qs,
        'buscar':          buscar,
        'equipo_id':       equipo_id,
        'categoria_id':    categoria_id,
        'estado':          estado,
        'equipos':         Equipo.objects.filter(activo=True).order_by('nombre'),
        'categorias':      Categoria.objects.filter(activo=True).order_by('nombre'),
        'total_filtrados': qs.count(),
        'query_params':    params,
    })


@solo_admin
def admin_reporte_asistencia(request):
    """Reporte de asistencia por equipo o deportista (admin)."""
    equipos   = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id = request.GET.get('equipo', '')
    jug_id    = request.GET.get('jugador', '')
    tipo      = request.GET.get('tipo', 'equipo')  # 'equipo' | 'deportista'

    jugadores   = Jugador.objects.none()
    asistencias = Asistencia.objects.none()
    jugador     = None
    equipo      = None
    resumen     = {}

    if equipo_id:
        equipo    = get_object_or_404(Equipo, id=equipo_id)
        jugadores = Jugador.objects.filter(equipo=equipo, activo=True).select_related('usuario').order_by('usuario__last_name')

    if tipo == 'deportista' and jug_id:
        jugador     = get_object_or_404(Jugador, id=jug_id)
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()
    elif tipo == 'equipo' and equipo_id:
        asistencias = Asistencia.objects.filter(
            entrenamiento__equipo=equipo
        ).select_related('jugador__usuario', 'entrenamiento').order_by('-entrenamiento__fecha', 'jugador__usuario__last_name')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    params = request.GET.urlencode()
    return render(request, 'panel/reportes/asistencia.html', {
        'equipos':    equipos,
        'equipo_id':  equipo_id,
        'equipo':     equipo,
        'jugadores':  jugadores,
        'jug_id':     jug_id,
        'jugador':    jugador,
        'tipo':       tipo,
        'asistencias': asistencias,
        'resumen':    resumen,
        'estados':    Asistencia.ESTADO_CHOICES,
        'query_params': params,
    })


@solo_admin
def admin_reporte_entrenamientos(request):
    """Reporte de entrenamientos con filtros (admin)."""
    qs = Entrenamiento.objects.select_related('equipo__categoria', 'entrenador')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    buscar      = request.GET.get('buscar', '').strip()

    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    if buscar:
        qs = qs.filter(Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar))

    qs = qs.order_by('-fecha')
    params = request.GET.urlencode()
    return render(request, 'panel/reportes/entrenamientos.html', {
        'entrenamientos':  qs,
        'equipos':         Equipo.objects.filter(activo=True).order_by('nombre'),
        'equipo_id':       equipo_id,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'buscar':          buscar,
        'total_filtrados': qs.count(),
        'query_params':    params,
    })


@solo_admin
def admin_reporte_partidos(request):
    """Reporte de partidos con estadisticas generales (admin)."""
    qs = PartidoCalendario.objects.select_related(
        'equipo_propio__categoria'
    ).prefetch_related('estadisticas__jugador__usuario')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if equipo_id:
        qs = qs.filter(equipo_propio_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by('-fecha')
    params = request.GET.urlencode()
    return render(request, 'panel/reportes/partidos.html', {
        'partidos':        qs,
        'equipos':         Equipo.objects.filter(activo=True).order_by('nombre'),
        'equipo_id':       equipo_id,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_filtrados': qs.count(),
        'query_params':    params,
    })


@solo_entrenador
def ent_reporte_asistencia(request):
    """Reporte de asistencia por equipo o deportista (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id   = request.GET.get('equipo', '')
    jug_id      = request.GET.get('jugador', '')
    tipo        = request.GET.get('tipo', 'equipo')

    jugadores   = Jugador.objects.none()
    asistencias = Asistencia.objects.none()
    jugador     = None
    equipo      = None
    resumen     = {}

    if equipo_id:
        equipo    = get_object_or_404(Equipo, id=equipo_id, entrenador=request.user)
        jugadores = Jugador.objects.filter(equipo=equipo, activo=True).select_related('usuario').order_by('usuario__last_name')

    if tipo == 'deportista' and jug_id:
        jugador     = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()
    elif tipo == 'equipo' and equipo_id:
        asistencias = Asistencia.objects.filter(
            entrenamiento__equipo=equipo
        ).select_related('jugador__usuario', 'entrenamiento').order_by('-entrenamiento__fecha', 'jugador__usuario__last_name')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    params = request.GET.urlencode()
    return render(request, 'entrenador/reportes/asistencia.html', {
        'equipos':     mis_equipos,
        'equipo_id':   equipo_id,
        'equipo':      equipo,
        'jugadores':   jugadores,
        'jug_id':      jug_id,
        'jugador':     jugador,
        'tipo':        tipo,
        'asistencias': asistencias,
        'resumen':     resumen,
        'estados':     Asistencia.ESTADO_CHOICES,
        'query_params': params,
    })


@solo_entrenador
def ent_reporte_entrenamientos(request):
    """Reporte de entrenamientos propios con filtros (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = Entrenamiento.objects.filter(
        equipo__in=mis_equipos
    ).select_related('equipo__categoria', 'entrenador')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    buscar      = request.GET.get('buscar', '').strip()

    if equipo_id:
        qs = qs.filter(equipo_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)
    if buscar:
        qs = qs.filter(Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar))

    qs = qs.order_by('-fecha')
    params = request.GET.urlencode()
    return render(request, 'entrenador/reportes/entrenamientos.html', {
        'entrenamientos':  qs,
        'equipos':         mis_equipos,
        'equipo_id':       equipo_id,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'buscar':          buscar,
        'total_filtrados': qs.count(),
        'query_params':    params,
    })


@solo_entrenador
def ent_reporte_partidos(request):
    """Reporte de partidos con estadisticas generales (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = PartidoCalendario.objects.filter(
        equipo_propio__in=mis_equipos
    ).select_related(
        'equipo_propio__categoria'
    ).prefetch_related('estadisticas__jugador__usuario')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if equipo_id:
        qs = qs.filter(equipo_propio_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by('-fecha')
    params = request.GET.urlencode()
    return render(request, 'entrenador/reportes/partidos.html', {
        'partidos':        qs,
        'equipos':         mis_equipos,
        'equipo_id':       equipo_id,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_filtrados': qs.count(),
        'query_params':    params,
    })


@solo_admin
def admin_reporte_rendimiento(request):
    """Página de selección para generar reporte de rendimiento por jugador o equipo."""
    jugadores = Jugador.objects.filter(
        activo=True
    ).select_related('usuario', 'equipo').order_by('usuario__last_name')
    equipos = Equipo.objects.filter(activo=True).select_related('categoria').order_by('nombre')
    return render(request, 'panel/reportes/rendimiento.html', {
        'jugadores': jugadores,
        'equipos':   equipos,
    })


@solo_admin
def admin_reporte_rendimiento_pdf(request, tipo, objeto_id):
    """Genera PDF de rendimiento para un jugador o equipo (admin, sin restriccion de equipo)."""
    hoy = timezone.now()

    if tipo == 'jugador':
        jugador       = get_object_or_404(Jugador, id=objeto_id)
        evaluaciones  = Evaluacion.objects.filter(jugador=jugador).select_related('entrenador').order_by('-fecha')
        observaciones = ObservacionDeportista.objects.filter(jugador=jugador).select_related('entrenador').order_by('-fecha')
        stats         = EstadisticaPartido.objects.filter(jugador=jugador).select_related('partido__equipo_propio').order_by('-partido__fecha')
        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )
        datos = {
            'tipo': 'jugador',
            'titulo': f'Reporte de Rendimiento — {jugador.usuario.get_full_name() or jugador.usuario.username}',
            'jugador': jugador,
            'evaluaciones': evaluaciones,
            'observaciones': observaciones,
            'stats': stats,
            'totales': totales,
            'fecha_generacion': hoy,
            'generado_por': request.user.get_full_name() or request.user.username,
        }
        nombre_archivo = f'reporte_jugador_{objeto_id}_{hoy.strftime("%Y%m%d")}.pdf'

    elif tipo == 'equipo':
        equipo       = get_object_or_404(Equipo, id=objeto_id)
        jugadores_qs = Jugador.objects.filter(equipo=equipo, activo=True).select_related('usuario').order_by('usuario__last_name')
        datos_jugadores = []
        for jug in jugadores_qs:
            evs  = Evaluacion.objects.filter(jugador=jug).order_by('-fecha')
            obs  = ObservacionDeportista.objects.filter(jugador=jug).order_by('-fecha')
            sts  = EstadisticaPartido.objects.filter(jugador=jug).select_related('partido__equipo_propio').order_by('-partido__fecha')
            tots = sts.aggregate(
                total_goles=Sum('goles'),
                total_asistencias=Sum('asistencias'),
                total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
                total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
                promedio_calif=Avg('calificacion'),
            )
            datos_jugadores.append({'jugador': jug, 'evaluaciones': evs, 'observaciones': obs, 'stats': sts, 'totales': tots})
        datos = {
            'tipo': 'equipo',
            'titulo': f'Reporte de Rendimiento — Equipo {equipo.nombre}',
            'equipo': equipo,
            'datos_jugadores': datos_jugadores,
            'fecha_generacion': hoy,
            'generado_por': request.user.get_full_name() or request.user.username,
        }
        nombre_archivo = f'reporte_equipo_{objeto_id}_{hoy.strftime("%Y%m%d")}.pdf'

    else:
        raise Http404

    html = render_to_string('entrenador/reporte_rendimiento.html', datos)
    buffer = BytesIO()
    pisa.CreatePDF(html.encode('utf-8'), dest=buffer)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Content-Length'] = len(pdf)
    return response


# ── Reporte de rendimiento PDF (entrenador) ────────────────────────────────────

@solo_entrenador
def ent_reporte_rendimiento(request):
    """Página de selección para generar reporte de rendimiento por jugador o equipo (entrenador)."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugadores = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario', 'equipo').order_by('usuario__last_name')
    return render(request, 'entrenador/reportes/rendimiento.html', {
        'jugadores': jugadores,
        'equipos':   mis_equipos.select_related('categoria').order_by('nombre'),
    })


@solo_entrenador
def ent_reporte_rendimiento_pdf(request, tipo, objeto_id):
    """Genera PDF de rendimiento para un jugador o equipo completo."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    hoy = timezone.now()

    if tipo == 'jugador':
        jugador = get_object_or_404(Jugador, id=objeto_id, equipo__in=mis_equipos)
        evaluaciones  = Evaluacion.objects.filter(jugador=jugador).select_related('entrenador').order_by('-fecha')
        observaciones = ObservacionDeportista.objects.filter(jugador=jugador).select_related('entrenador').order_by('-fecha')
        stats = EstadisticaPartido.objects.filter(jugador=jugador).select_related('partido__equipo_propio').order_by('-partido__fecha')
        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )
        datos = {
            'tipo': 'jugador',
            'titulo': f'Reporte de Rendimiento — {jugador.usuario.get_full_name() or jugador.usuario.username}',
            'jugador': jugador,
            'evaluaciones': evaluaciones,
            'observaciones': observaciones,
            'stats': stats,
            'totales': totales,
            'fecha_generacion': hoy,
            'generado_por': request.user.get_full_name() or request.user.username,
        }
        nombre_archivo = f'reporte_jugador_{objeto_id}_{hoy.strftime("%Y%m%d")}.pdf'

    elif tipo == 'equipo':
        equipo = get_object_or_404(Equipo, id=objeto_id, entrenador=request.user)
        jugadores_qs = Jugador.objects.filter(equipo=equipo, activo=True).select_related('usuario').order_by('usuario__last_name')
        datos_jugadores = []
        for jug in jugadores_qs:
            evs   = Evaluacion.objects.filter(jugador=jug).order_by('-fecha')
            obs   = ObservacionDeportista.objects.filter(jugador=jug).order_by('-fecha')
            sts   = EstadisticaPartido.objects.filter(jugador=jug).select_related('partido__equipo_propio').order_by('-partido__fecha')
            tots  = sts.aggregate(
                total_goles=Sum('goles'),
                total_asistencias=Sum('asistencias'),
                total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
                total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
                promedio_calif=Avg('calificacion'),
            )
            datos_jugadores.append({'jugador': jug, 'evaluaciones': evs, 'observaciones': obs, 'stats': sts, 'totales': tots})
        datos = {
            'tipo': 'equipo',
            'titulo': f'Reporte de Rendimiento — Equipo {equipo.nombre}',
            'equipo': equipo,
            'datos_jugadores': datos_jugadores,
            'fecha_generacion': hoy,
            'generado_por': request.user.get_full_name() or request.user.username,
        }
        nombre_archivo = f'reporte_equipo_{objeto_id}_{hoy.strftime("%Y%m%d")}.pdf'
    else:
        raise Http404

    html = render_to_string('entrenador/reporte_rendimiento.html', datos)
    buffer = BytesIO()
    pisa.CreatePDF(html.encode('utf-8'), dest=buffer)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Content-Length'] = len(pdf)
    return response


# ── Historial de deportista (entrenador) ────────────────────────────────────────

@solo_entrenador
def ent_historial_jugador(request, jug_id):
    """Historial completo de un jugador: evaluaciones, observaciones y estadísticas de partido."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugador     = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)

    evaluaciones = Evaluacion.objects.filter(
        jugador=jugador
    ).select_related('entrenador').order_by('-fecha')

    observaciones = ObservacionDeportista.objects.filter(
        jugador=jugador
    ).select_related('entrenador').order_by('-fecha')

    stats = EstadisticaPartido.objects.filter(
        jugador=jugador
    ).select_related('partido__equipo_propio').order_by('-partido__fecha')

    totales = stats.aggregate(
        total_goles=Sum('goles'),
        total_asistencias=Sum('asistencias'),
        total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
        total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
        promedio_calif=Avg('calificacion'),
    )

    # Datos para gráfico de evolución: evaluaciones con puntaje ordenadas cronológicamente
    evs_con_puntaje = Evaluacion.objects.filter(
        jugador=jugador, puntaje__isnull=False
    ).order_by('fecha', 'id')

    graf_labels   = json.dumps([str(e.fecha) for e in evs_con_puntaje])
    graf_tecnica  = json.dumps([float(e.puntaje) if e.tipo == 'tecnica'  else None for e in evs_con_puntaje])
    graf_tactica  = json.dumps([float(e.puntaje) if e.tipo == 'tactica'  else None for e in evs_con_puntaje])
    graf_fisica   = json.dumps([float(e.puntaje) if e.tipo == 'fisica'   else None for e in evs_con_puntaje])
    graf_todas    = json.dumps([float(e.puntaje) for e in evs_con_puntaje])

    return render(request, 'entrenador/jugadores/historial.html', {
        'jugador':       jugador,
        'evaluaciones':  evaluaciones,
        'observaciones': observaciones,
        'stats':         stats,
        'totales':       totales,
        'graf_labels':   graf_labels,
        'graf_tecnica':  graf_tecnica,
        'graf_tactica':  graf_tactica,
        'graf_fisica':   graf_fisica,
        'graf_todas':    graf_todas,
        'tiene_grafico': evs_con_puntaje.exists(),
    })


# ── Evaluaciones del entrenador ──────────────────────────────────────────────────

@solo_entrenador
def ent_evaluaciones_lista(request):
    """Lista de evaluaciones creadas por el entrenador con filtros"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = Evaluacion.objects.filter(
        entrenador=request.user
    ).select_related('jugador__usuario', 'jugador__equipo')

    buscar      = request.GET.get('buscar', '').strip()
    equipo_id   = request.GET.get('equipo', '')
    jugador_id  = request.GET.get('jugador', '')
    tipo        = request.GET.get('tipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(titulo__icontains=buscar) |
            Q(jugador__usuario__first_name__icontains=buscar) |
            Q(jugador__usuario__last_name__icontains=buscar)
        )
    if equipo_id:
        qs = qs.filter(jugador__equipo_id=equipo_id)
    if jugador_id:
        qs = qs.filter(jugador_id=jugador_id)
    if tipo:
        qs = qs.filter(tipo=tipo)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    # Jugadores disponibles para el filtro (del equipo seleccionado o de todos)
    if equipo_id:
        jugadores_filtro = Jugador.objects.filter(
            equipo_id=equipo_id, equipo__in=mis_equipos, activo=True
        ).select_related('usuario').order_by('usuario__last_name')
    else:
        jugadores_filtro = Jugador.objects.filter(
            equipo__in=mis_equipos, activo=True
        ).select_related('usuario').order_by('usuario__last_name')

    paginator = Paginator(qs.order_by('-fecha'), 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'entrenador/evaluaciones/lista.html', {
        'evaluaciones':    page,
        'buscar':          buscar,
        'equipo_id':       equipo_id,
        'jugador_id':      jugador_id,
        'tipo':            tipo,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'equipos':         mis_equipos,
        'jugadores_filtro': jugadores_filtro,
        'tipos':           Evaluacion.TIPO_CHOICES,
        'total_filtrados': qs.count(),
    })


@solo_entrenador
def ent_evaluacion_crear(request):
    """Crear evaluación para un jugador del entrenador"""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugadores   = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario', 'equipo')

    if request.method == 'POST':
        jugador_id   = request.POST.get('jugador')
        titulo       = request.POST.get('titulo', '').strip()
        tipo         = request.POST.get('tipo', 'tecnica')
        descripcion  = request.POST.get('descripcion', '').strip()
        puntaje      = request.POST.get('puntaje') or None
        observaciones = request.POST.get('observaciones', '').strip()
        fecha        = request.POST.get('fecha')

        if not all([jugador_id, titulo, fecha]):
            messages.error(request, 'Jugador, título y fecha son obligatorios.')
        elif not jugadores.filter(id=jugador_id).exists():
            messages.error(request, 'No tienes permiso sobre ese jugador.')
        else:
            Evaluacion.objects.create(
                jugador_id=jugador_id,
                entrenador=request.user,
                titulo=titulo,
                tipo=tipo,
                descripcion=descripcion,
                puntaje=puntaje,
                observaciones=observaciones,
                fecha=fecha,
            )
            messages.success(request, f'Evaluación "{titulo}" creada correctamente.')
            return ent_evaluaciones_lista(request)

    return render(request, 'entrenador/evaluaciones/form.html', {
        'titulo':    'Nueva Evaluación',
        'jugadores': jugadores,
    })


@solo_entrenador
def ent_evaluacion_editar(request, eval_id):
    """Editar evaluación propia"""
    evaluacion  = get_object_or_404(Evaluacion, id=eval_id, entrenador=request.user)
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    jugadores   = Jugador.objects.filter(
        equipo__in=mis_equipos, activo=True
    ).select_related('usuario', 'equipo')

    if request.method == 'POST':
        evaluacion.jugador_id    = request.POST.get('jugador')
        evaluacion.titulo        = request.POST.get('titulo', '').strip()
        evaluacion.tipo          = request.POST.get('tipo', 'tecnica')
        evaluacion.descripcion   = request.POST.get('descripcion', '').strip()
        evaluacion.puntaje       = request.POST.get('puntaje') or None
        evaluacion.observaciones = request.POST.get('observaciones', '').strip()
        evaluacion.fecha         = request.POST.get('fecha')
        evaluacion.save()
        messages.success(request, f'Evaluación "{evaluacion.titulo}" actualizada.')

    return render(request, 'entrenador/evaluaciones/form.html', {
        'titulo':     'Editar Evaluación',
        'evaluacion': evaluacion,
        'jugadores':  jugadores,
    })


@solo_entrenador
def ent_evaluacion_eliminar(request, eval_id):
    """Eliminar evaluación propia"""
    evaluacion = get_object_or_404(Evaluacion, id=eval_id, entrenador=request.user)
    nombre     = evaluacion.titulo
    evaluacion.delete()
    messages.success(request, f'Evaluación "{nombre}" eliminada.')
    return ent_evaluaciones_lista(request)


# ── Pagos del entrenador ─────────────────────────────────────────────────────

@solo_entrenador
def ent_pagos_lista(request):
    """Lista de pagos de los jugadores del entrenador con filtros"""
    _marcar_vencidos()
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = Pago.objects.filter(
        jugador__equipo__in=mis_equipos
    ).select_related('jugador__usuario', 'jugador__equipo')

    buscar   = request.GET.get('buscar', '').strip()
    estado   = request.GET.get('estado', '')
    equipo_id = request.GET.get('equipo', '')

    if buscar:
        qs = qs.filter(
            Q(descripcion__icontains=buscar) |
            Q(jugador__usuario__first_name__icontains=buscar) |
            Q(jugador__usuario__last_name__icontains=buscar)
        )
    if estado:
        qs = qs.filter(estado=estado)
    if equipo_id:
        qs = qs.filter(jugador__equipo_id=equipo_id)

    paginator = Paginator(qs.order_by('-fecha_vencimiento'), 10)
    page = paginator.get_page(request.GET.get('page'))

    # Totales sobre el queryset filtrado
    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    alertas_vencidos = qs.filter(estado='vencido').count()

    return render(request, 'entrenador/pagos/lista.html', {
        'pagos':            page,
        'buscar':           buscar,
        'estado':           estado,
        'equipo_id':        equipo_id,
        'equipos':          mis_equipos,
        'estados':          Pago.ESTADO_CHOICES,
        'total_pagado':     total_pagado,
        'total_pendiente':  total_pendiente,
        'total_vencido':    total_vencido,
        'alertas_vencidos': alertas_vencidos,
        'total_filtrados':  qs.count(),
    })


# ─────────────────────────────────────────────
# PARTIDOS CALENDARIO (ADMIN)
# ─────────────────────────────────────────────

@solo_admin
def admin_partidos(request):
    """Lista de partidos del calendario con filtros."""
    qs = PartidoCalendario.objects.select_related('equipo_propio', 'registrado_por')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if equipo_id:
        qs = qs.filter(equipo_propio_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    return render(request, 'panel/partidos/lista.html', {
        'partidos':    qs,
        'equipos':     Equipo.objects.filter(activo=True).order_by('nombre'),
        'equipo_id':   equipo_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total':       qs.count(),
    })


@solo_admin
def admin_partido_crear(request):
    """Crear partido de calendario."""
    equipos = Equipo.objects.filter(activo=True).order_by('nombre')
    if request.method == 'POST':
        equipo_id    = request.POST.get('equipo_propio')
        equipo_rival = request.POST.get('equipo_rival', '').strip()
        fecha        = request.POST.get('fecha')
        hora         = request.POST.get('hora')
        cancha       = request.POST.get('cancha', '').strip()
        descripcion  = request.POST.get('descripcion', '').strip()

        errores = []
        if not equipo_id:    errores.append('El equipo propio es obligatorio.')
        if not equipo_rival: errores.append('El equipo rival es obligatorio.')
        if not fecha:        errores.append('La fecha es obligatoria.')
        if not hora:         errores.append('La hora es obligatoria.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            partido = PartidoCalendario.objects.create(
                equipo_propio_id=equipo_id,
                equipo_rival=equipo_rival,
                fecha=fecha, hora=hora,
                cancha=cancha, descripcion=descripcion,
                registrado_por=request.user,
            )
            _notificar_partido(partido, es_nuevo=True, emisor=request.user)
            messages.success(request, 'Partido creado correctamente.')
            return admin_partidos(request)

    return render(request, 'panel/partidos/form.html', {
        'titulo': 'Nuevo Partido', 'equipos': equipos,
    })


@solo_admin
def admin_partido_editar(request, partido_id):
    """Editar partido de calendario."""
    partido = get_object_or_404(PartidoCalendario, id=partido_id)
    equipos = Equipo.objects.filter(activo=True).order_by('nombre')
    if request.method == 'POST':
        partido.equipo_propio_id = request.POST.get('equipo_propio')
        partido.equipo_rival     = request.POST.get('equipo_rival', '').strip()
        partido.fecha            = request.POST.get('fecha')
        partido.hora             = request.POST.get('hora')
        partido.cancha           = request.POST.get('cancha', '').strip()
        partido.descripcion      = request.POST.get('descripcion', '').strip()
        partido.save()
        _notificar_partido(partido, es_nuevo=False, emisor=request.user)
        messages.success(request, 'Partido actualizado correctamente.')
        return admin_partidos(request)

    return render(request, 'panel/partidos/form.html', {
        'titulo': 'Editar Partido', 'partido': partido, 'equipos': equipos,
    })


@solo_admin
def admin_partido_eliminar(request, partido_id):
    """Eliminar partido de calendario."""
    partido = get_object_or_404(PartidoCalendario, id=partido_id)
    partido.delete()
    messages.success(request, 'Partido eliminado correctamente.')
    return admin_partidos(request)


@solo_admin
def admin_estadisticas_equipo(request):
    """Admin: resultados y estadísticas de partidos filtrados por equipo."""
    equipos   = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id = request.GET.get('equipo', '')
    partidos  = None

    if equipo_id:
        partidos = PartidoCalendario.objects.filter(
            equipo_propio_id=equipo_id
        ).prefetch_related('estadisticas__jugador__usuario').order_by('-fecha')

    return render(request, 'panel/estadisticas/por_equipo.html', {
        'equipos':   equipos,
        'equipo_id': equipo_id,
        'partidos':  partidos,
    })


@solo_admin
def admin_estadisticas_jugador(request):
    """Admin: estadísticas acumuladas de un jugador en todos sus partidos."""
    equipos   = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id = request.GET.get('equipo', '')
    jug_id    = request.GET.get('jugador', '')
    jugadores = Jugador.objects.none()
    jugador   = None
    stats     = None
    totales   = {}

    if equipo_id:
        jugadores = Jugador.objects.filter(
            equipo_id=equipo_id, activo=True
        ).select_related('usuario').order_by('usuario__last_name')

    if jug_id:
        jugador = get_object_or_404(Jugador, id=jug_id)
        stats   = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')
        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )

    return render(request, 'panel/estadisticas/por_jugador.html', {
        'equipos':   equipos,
        'equipo_id': equipo_id,
        'jugadores': jugadores,
        'jug_id':    jug_id,
        'jugador':   jugador,
        'stats':     stats,
        'totales':   totales,
    })


@solo_entrenador
def ent_estadisticas_equipo(request):
    """Entrenador: resultados y estadísticas de sus partidos filtrados por equipo."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id   = request.GET.get('equipo', '')
    partidos    = None

    if equipo_id:
        partidos = PartidoCalendario.objects.filter(
            equipo_propio_id=equipo_id, equipo_propio__in=mis_equipos
        ).prefetch_related('estadisticas__jugador__usuario').order_by('-fecha')

    return render(request, 'entrenador/estadisticas/por_equipo.html', {
        'equipos':   mis_equipos,
        'equipo_id': equipo_id,
        'partidos':  partidos,
    })


@solo_entrenador
def ent_estadisticas_jugador(request):
    """Entrenador: estadísticas acumuladas de un jugador de sus equipos."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id   = request.GET.get('equipo', '')
    jug_id      = request.GET.get('jugador', '')
    jugadores   = Jugador.objects.none()
    jugador     = None
    stats       = None
    totales     = {}

    if equipo_id:
        jugadores = Jugador.objects.filter(
            equipo_id=equipo_id, activo=True
        ).select_related('usuario').order_by('usuario__last_name')

    if jug_id:
        jugador = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)
        stats   = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')
        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )

    return render(request, 'entrenador/estadisticas/por_jugador.html', {
        'equipos':   mis_equipos,
        'equipo_id': equipo_id,
        'jugadores': jugadores,
        'jug_id':    jug_id,
        'jugador':   jugador,
        'stats':     stats,
        'totales':   totales,
    })


@solo_admin
def admin_asistencia(request):
    """
    Admin: consulta asistencia por entrenamiento.
    Filtra por equipo, entrenamiento y fecha.
    """
    equipos       = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id     = request.GET.get('equipo', '')
    ent_id        = request.GET.get('entrenamiento', '')
    asistencias   = None
    entrenamiento = None
    resumen       = {}

    entrenamientos = Entrenamiento.objects.none()
    if equipo_id:
        entrenamientos = Entrenamiento.objects.filter(
            equipo_id=equipo_id
        ).order_by('-fecha')

    if ent_id:
        entrenamiento = get_object_or_404(Entrenamiento, id=ent_id)
        asistencias   = Asistencia.objects.filter(
            entrenamiento=entrenamiento
        ).select_related('jugador__usuario').order_by('jugador__usuario__last_name')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    return render(request, 'panel/asistencia/por_entrenamiento.html', {
        'equipos':        equipos,
        'equipo_id':      equipo_id,
        'entrenamientos': entrenamientos,
        'ent_id':         ent_id,
        'entrenamiento':  entrenamiento,
        'asistencias':    asistencias,
        'resumen':        resumen,
    })


@solo_admin
def admin_asistencia_deportista(request):
    """
    Admin: consulta asistencia por deportista.
    Filtra por equipo y jugador.
    """
    equipos   = Equipo.objects.filter(activo=True).order_by('nombre')
    equipo_id = request.GET.get('equipo', '')
    jug_id    = request.GET.get('jugador', '')
    jugadores = Jugador.objects.none()
    jugador   = None
    asistencias = None
    resumen     = {}

    if equipo_id:
        jugadores = Jugador.objects.filter(
            equipo_id=equipo_id, activo=True
        ).select_related('usuario').order_by('usuario__last_name')

    if jug_id:
        jugador     = get_object_or_404(Jugador, id=jug_id)
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    return render(request, 'panel/asistencia/por_deportista.html', {
        'equipos':    equipos,
        'equipo_id':  equipo_id,
        'jugadores':  jugadores,
        'jug_id':     jug_id,
        'jugador':    jugador,
        'asistencias': asistencias,
        'resumen':    resumen,
    })


# ── Partidos del entrenador ─────────────────────────────────────────────────────────

@solo_entrenador
def ent_partidos(request):
    """Lista de partidos de los equipos del entrenador."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    qs = PartidoCalendario.objects.filter(
        equipo_propio__in=mis_equipos
    ).select_related('equipo_propio')

    equipo_id   = request.GET.get('equipo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if equipo_id:
        qs = qs.filter(equipo_propio_id=equipo_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    return render(request, 'entrenador/partidos/lista.html', {
        'partidos':    qs,
        'equipos':     mis_equipos,
        'equipo_id':   equipo_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'total':       qs.count(),
    })


@solo_entrenador
def ent_estadisticas(request, partido_id):
    """
    Registrar o actualizar estadisticas por jugador en un partido.
    Usa upsert (get_or_create) para cada jugador del equipo.
    """
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    partido     = get_object_or_404(PartidoCalendario, id=partido_id, equipo_propio__in=mis_equipos)
    jugadores   = Jugador.objects.filter(
        equipo=partido.equipo_propio, activo=True
    ).select_related('usuario').order_by('usuario__last_name')

    existentes = {
        e.jugador_id: e
        for e in EstadisticaPartido.objects.filter(partido=partido)
    }

    if request.method == 'POST':
        for jugador in jugadores:
            goles       = int(request.POST.get(f'goles_{jugador.id}', 0) or 0)
            asistencias = int(request.POST.get(f'asistencias_{jugador.id}', 0) or 0)
            t_amarilla  = bool(request.POST.get(f'amarilla_{jugador.id}'))
            t_roja      = bool(request.POST.get(f'roja_{jugador.id}'))
            calificacion = request.POST.get(f'calificacion_{jugador.id}', '').strip() or None

            if jugador.id in existentes:
                est = existentes[jugador.id]
                est.goles            = goles
                est.asistencias      = asistencias
                est.tarjeta_amarilla = t_amarilla
                est.tarjeta_roja     = t_roja
                est.calificacion     = calificacion
                est.save()
            else:
                EstadisticaPartido.objects.create(
                    partido=partido,
                    jugador=jugador,
                    goles=goles,
                    asistencias=asistencias,
                    tarjeta_amarilla=t_amarilla,
                    tarjeta_roja=t_roja,
                    calificacion=calificacion,
                    registrado_por=request.user,
                )
        messages.success(request, 'Estadisticas guardadas correctamente.')
        return ent_partidos(request)

    filas = []
    for jugador in jugadores:
        est = existentes.get(jugador.id)
        filas.append({
            'jugador':     jugador,
            'goles':       est.goles            if est else 0,
            'asistencias': est.asistencias       if est else 0,
            'amarilla':    est.tarjeta_amarilla  if est else False,
            'roja':        est.tarjeta_roja       if est else False,
            'calificacion': est.calificacion     if est else '',
        })

    return render(request, 'entrenador/partidos/estadisticas.html', {
        'partido': partido,
        'filas':   filas,
    })


@solo_entrenador
def ent_resultado(request, partido_id):
    """
    Registrar o actualizar el resultado de un partido propio.
    """
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    partido     = get_object_or_404(PartidoCalendario, id=partido_id, equipo_propio__in=mis_equipos)

    if request.method == 'POST':
        goles_favor  = request.POST.get('goles_favor', '').strip()
        goles_contra = request.POST.get('goles_contra', '').strip()
        observaciones = request.POST.get('observaciones', '').strip()

        errores = []
        if goles_favor == '':  errores.append('Los goles a favor son obligatorios.')
        if goles_contra == '': errores.append('Los goles en contra son obligatorios.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            partido.goles_favor   = int(goles_favor)
            partido.goles_contra  = int(goles_contra)
            partido.observaciones = observaciones
            partido.save()
            messages.success(request, 'Resultado registrado correctamente.')
            return ent_partidos(request)

    return render(request, 'entrenador/partidos/resultado.html', {'partido': partido})


@solo_entrenador
def ent_convocatoria(request, partido_id):
    """
    Crear o actualizar la convocatoria de jugadores para un partido.
    Muestra todos los jugadores activos del equipo propio del partido.
    Permite seleccionar cuales van convocados y agregar una nota.
    """
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    partido     = get_object_or_404(PartidoCalendario, id=partido_id, equipo_propio__in=mis_equipos)

    jugadores = Jugador.objects.filter(
        equipo=partido.equipo_propio, activo=True
    ).select_related('usuario').order_by('usuario__last_name')

    convocatoria, _ = Convocatoria.objects.get_or_create(partido=partido)

    if request.method == 'POST':
        ids_seleccionados = request.POST.getlist('jugadores')
        nota              = request.POST.get('nota', '').strip()
        convocatoria.jugadores.set(ids_seleccionados)
        convocatoria.nota = nota
        convocatoria.save()
        # Notificar a cada jugador convocado
        jugadores_convocados = Jugador.objects.filter(
            id__in=ids_seleccionados
        ).select_related('usuario')
        asunto_conv = f'Convocatoria: {partido.equipo_propio.nombre} vs {partido.equipo_rival}'
        mensaje_conv = (
            f'Has sido convocado para el siguiente partido:\n\n'
            f'Equipo:  {partido.equipo_propio.nombre}\n'
            f'Rival:   {partido.equipo_rival}\n'
            f'Fecha:   {partido.fecha.strftime("%d/%m/%Y")}\n'
            f'Hora:    {partido.hora.strftime("%H:%M")}\n'
            f'Lugar:   {partido.cancha or "Por confirmar"}'
            + (f'\nNota:    {convocatoria.nota}' if convocatoria.nota else '')
        )
        Notificacion.objects.bulk_create([
            Notificacion(usuario=j.usuario, asunto=asunto_conv, mensaje=mensaje_conv, emisor=request.user)
            for j in jugadores_convocados
        ])
        for j in jugadores_convocados:
            if j.usuario.email and '@' in j.usuario.email:
                try:
                    send_mail(
                        subject=asunto_conv,
                        message=mensaje_conv,
                        from_email=django_settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[j.usuario.email],
                        fail_silently=True,
                    )
                except Exception:
                    pass
        messages.success(request, f'Convocatoria guardada: {len(ids_seleccionados)} jugador(es) convocado(s).')
        return ent_partidos(request)

    convocados_ids = set(convocatoria.jugadores.values_list('id', flat=True))

    return render(request, 'entrenador/partidos/convocatoria.html', {
        'partido':      partido,
        'jugadores':    jugadores,
        'convocados':   convocados_ids,
        'convocatoria': convocatoria,
    })


@solo_entrenador
def ent_partido_crear(request):
    """Crear partido para un equipo propio."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    if request.method == 'POST':
        equipo_id    = request.POST.get('equipo_propio')
        equipo_rival = request.POST.get('equipo_rival', '').strip()
        fecha        = request.POST.get('fecha')
        hora         = request.POST.get('hora')
        cancha       = request.POST.get('cancha', '').strip()
        descripcion  = request.POST.get('descripcion', '').strip()

        errores = []
        if not equipo_id:    errores.append('El equipo propio es obligatorio.')
        elif not mis_equipos.filter(id=equipo_id).exists():
            errores.append('No tenés permiso sobre ese equipo.')
        if not equipo_rival: errores.append('El equipo rival es obligatorio.')
        if not fecha:        errores.append('La fecha es obligatoria.')
        if not hora:         errores.append('La hora es obligatoria.')

        if errores:
            for e in errores: messages.error(request, e)
        else:
            partido = PartidoCalendario.objects.create(
                equipo_propio_id=equipo_id,
                equipo_rival=equipo_rival,
                fecha=fecha, hora=hora,
                cancha=cancha, descripcion=descripcion,
                registrado_por=request.user,
            )
            _notificar_partido(partido, es_nuevo=True, emisor=request.user)
            messages.success(request, 'Partido creado correctamente.')
            return ent_partidos(request)

    return render(request, 'entrenador/partidos/form.html', {
        'titulo': 'Nuevo Partido', 'equipos': mis_equipos,
    })


@solo_entrenador
def ent_partido_editar(request, partido_id):
    """Editar partido propio."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    partido     = get_object_or_404(PartidoCalendario, id=partido_id, equipo_propio__in=mis_equipos)
    if request.method == 'POST':
        equipo_id = request.POST.get('equipo_propio')
        if not mis_equipos.filter(id=equipo_id).exists():
            messages.error(request, 'No tenés permiso sobre ese equipo.')
        else:
            partido.equipo_propio_id = equipo_id
            partido.equipo_rival     = request.POST.get('equipo_rival', '').strip()
            partido.fecha            = request.POST.get('fecha')
            partido.hora             = request.POST.get('hora')
            partido.cancha           = request.POST.get('cancha', '').strip()
            partido.descripcion      = request.POST.get('descripcion', '').strip()
            partido.save()
            _notificar_partido(partido, es_nuevo=False, emisor=request.user)
            messages.success(request, 'Partido actualizado correctamente.')
            return ent_partidos(request)

    return render(request, 'entrenador/partidos/form.html', {
        'titulo': 'Editar Partido', 'partido': partido, 'equipos': mis_equipos,
    })


@solo_entrenador
def ent_partido_eliminar(request, partido_id):
    """Eliminar partido propio."""
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    partido     = get_object_or_404(PartidoCalendario, id=partido_id, equipo_propio__in=mis_equipos)
    partido.delete()
    messages.success(request, 'Partido eliminado correctamente.')
    return ent_partidos(request)


@solo_entrenador
def ent_asistencia_consulta(request):
    """
    Entrenador: consulta asistencia por entrenamiento propio.
    Muestra todos los entrenamientos del equipo, no solo los creados por el entrenador.
    """
    mis_equipos    = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id      = request.GET.get('equipo', '')
    ent_id         = request.GET.get('entrenamiento', '')
    asistencias    = None
    entrenamiento  = None
    resumen        = {}

    entrenamientos = Entrenamiento.objects.none()
    if equipo_id:
        entrenamientos = Entrenamiento.objects.filter(
            equipo_id=equipo_id, equipo__in=mis_equipos
        ).order_by('-fecha')

    if ent_id:
        entrenamiento = get_object_or_404(Entrenamiento, id=ent_id, equipo__in=mis_equipos)
        asistencias   = Asistencia.objects.filter(
            entrenamiento=entrenamiento
        ).select_related('jugador__usuario').order_by('jugador__usuario__last_name')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    return render(request, 'entrenador/asistencia/por_entrenamiento.html', {
        'equipos':        mis_equipos,
        'equipo_id':      equipo_id,
        'entrenamientos': entrenamientos,
        'ent_id':         ent_id,
        'entrenamiento':  entrenamiento,
        'asistencias':    asistencias,
        'resumen':        resumen,
    })


@solo_entrenador
def ent_asistencia_deportista(request):
    """
    Entrenador: consulta asistencia por deportista de sus equipos.
    """
    mis_equipos = Equipo.objects.filter(entrenador=request.user, activo=True)
    equipo_id   = request.GET.get('equipo', '')
    jug_id      = request.GET.get('jugador', '')
    jugadores   = Jugador.objects.none()
    jugador     = None
    asistencias = None
    resumen     = {}

    if equipo_id:
        jugadores = Jugador.objects.filter(
            equipo_id=equipo_id, activo=True
        ).select_related('usuario').order_by('usuario__last_name')

    if jug_id:
        jugador     = get_object_or_404(Jugador, id=jug_id, equipo__in=mis_equipos)
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()

    return render(request, 'entrenador/asistencia/por_deportista.html', {
        'equipos':     mis_equipos,
        'equipo_id':   equipo_id,
        'jugadores':   jugadores,
        'jug_id':      jug_id,
        'jugador':     jugador,
        'asistencias': asistencias,
        'resumen':     resumen,
    })


@solo_entrenador
def ent_asistencia(request, ent_id):
    """
    Registrar o actualizar asistencia de los jugadores de un entrenamiento.
    Carga todos los jugadores activos del equipo. Si ya existe un registro
    de asistencia para ese jugador lo muestra, si no lo crea al guardar.
    """
    entrenamiento = get_object_or_404(Entrenamiento, id=ent_id, entrenador=request.user)
    jugadores     = Jugador.objects.filter(
        equipo=entrenamiento.equipo, activo=True
    ).select_related('usuario').order_by('usuario__last_name')

    # Asistencias ya registradas para este entrenamiento
    asistencias_existentes = {
        a.jugador_id: a
        for a in Asistencia.objects.filter(entrenamiento=entrenamiento)
    }

    if request.method == 'POST':
        for jugador in jugadores:
            estado      = request.POST.get(f'estado_{jugador.id}', 'asistio')
            observacion = request.POST.get(f'obs_{jugador.id}', '').strip()
            if jugador.id in asistencias_existentes:
                a = asistencias_existentes[jugador.id]
                a.estado      = estado
                a.observacion = observacion
                a.save()
            else:
                Asistencia.objects.create(
                    entrenamiento=entrenamiento,
                    jugador=jugador,
                    estado=estado,
                    observacion=observacion,
                    registrado_por=request.user,
                )
        messages.success(request, 'Asistencia registrada correctamente.')
        return ent_entrenamientos_lista(request)

    # Armar lista con datos existentes para prellenar el form
    filas = []
    for jugador in jugadores:
        asistencia = asistencias_existentes.get(jugador.id)
        filas.append({
            'jugador':    jugador,
            'estado':     asistencia.estado      if asistencia else 'asistio',
            'observacion': asistencia.observacion if asistencia else '',
        })

    return render(request, 'entrenador/entrenamientos/asistencia.html', {
        'entrenamiento': entrenamiento,
        'filas':         filas,
        'estados':       Asistencia.ESTADO_CHOICES,
    })


@solo_entrenador
def ent_pago_crear(request):
    """Registrar pago para un jugador del entrenador"""
    return HttpResponseForbidden("La gestión de pagos es exclusiva del administrador.")



@solo_entrenador
def ent_pago_editar(request, pago_id):
    """Editar pago de un jugador del entrenador"""
    return HttpResponseForbidden("La gestión de pagos es exclusiva del administrador.")



@solo_entrenador
def ent_pago_eliminar(request, pago_id):
    """Eliminar pago de un jugador del entrenador y su movimiento financiero asociado."""
    return HttpResponseForbidden("La gestión de pagos es exclusiva del administrador.")


@login_required
def deportista_dashboard(request):
    """Dashboard del deportista con resumen completo."""
    if request.user.role != 'deportista':
        return HttpResponseForbidden("Sin permisos.")

    jugador = entrenador = equipo = categoria = None
    entrenamientos = evaluaciones = pagos = torneos = []
    total_pendiente = 0

    try:
        jugador = request.user.jugador
        if jugador.equipo:
            equipo     = jugador.equipo
            categoria  = equipo.categoria
            entrenador = equipo.entrenador
            entrenamientos = Entrenamiento.objects.filter(
                equipo=equipo).order_by('fecha')[:5]
            torneos = Torneo.objects.filter(
                categoria=categoria,
                estado__in=['planificado', 'en_curso']
            ).order_by('fecha_inicio')[:3]

        evaluaciones = Evaluacion.objects.filter(
            jugador=jugador).order_by('-fecha')[:5]
        pagos = Pago.objects.filter(
            jugador=jugador).order_by('-fecha_vencimiento')[:5]
        total_pendiente = Pago.objects.filter(
            jugador=jugador, estado='pendiente'
        ).count()
    except Exception:
        pass

    return render(request, 'deportista/dashboard.html', {
        'jugador':         jugador,
        'equipo':          equipo,
        'categoria':       categoria,
        'entrenador':      entrenador,
        'entrenamientos':  entrenamientos,
        'evaluaciones':    evaluaciones,
        'pagos':           pagos,
        'torneos':         torneos,
        'total_pendiente': total_pendiente,
        'notificaciones':  Notificacion.objects.filter(usuario=request.user, leida=False).order_by('-creado'),
    })


# ── Decorador solo deportista ────────────────────────────────────────────────
def solo_deportista(view_func):
    """Decorador que verifica que el usuario sea deportista"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, 'auth/login.html')
        if request.user.role != 'deportista':
            return HttpResponseForbidden("Sin permisos de deportista.")
        return view_func(request, *args, **kwargs)
    return wrapper


@solo_deportista
def dep_perfil(request):
    """Ver y actualizar datos personales del deportista."""
    usuario = request.user

    if request.method == 'POST':
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()

        errores = []
        if phone and not re.match(r'^\d{7,15}$', phone):
            errores.append('El teléfono debe contener solo números (7-15 dígitos).')
        if email and UserModel.objects.filter(
                email__iexact=email).exclude(id=usuario.id).exists():
            errores.append(f'El correo "{email}" ya está en uso por otra cuenta.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            usuario.phone = phone
            if email:
                usuario.email = email
            usuario.save()
            messages.success(request, 'Datos de contacto actualizados correctamente.')

    # Obtener perfil de jugador si existe
    jugador = None
    try:
        jugador = usuario.jugador
    except Exception:
        pass

    return render(request, 'deportista/perfil.html', {
        'usuario': usuario,
        'jugador': jugador,
    })


@solo_deportista
def dep_entrenamientos(request):
    """Lista de entrenamientos del equipo del deportista con filtros."""
    jugador = None
    qs      = Entrenamiento.objects.none()

    try:
        jugador = request.user.jugador
        if jugador.equipo:
            qs = Entrenamiento.objects.filter(
                equipo=jugador.equipo
            ).select_related('entrenador', 'equipo')
    except Exception:
        pass

    # Filtros
    buscar      = request.GET.get('buscar', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(titulo__icontains=buscar) | Q(lugar__icontains=buscar)
        )
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    paginator = Paginator(qs.order_by('-fecha'), 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'deportista/entrenamientos.html', {
        'entrenamientos':  page,
        'jugador':         jugador,
        'buscar':          buscar,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_filtrados': qs.count(),
    })


@solo_deportista
def dep_evaluaciones(request):
    """Lista de evaluaciones del deportista con filtros."""
    jugador = None
    qs      = Evaluacion.objects.none()

    try:
        jugador = request.user.jugador
        qs = Evaluacion.objects.filter(
            jugador=jugador
        ).select_related('entrenador')
    except Exception:
        pass

    # Filtros
    buscar      = request.GET.get('buscar', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(titulo__icontains=buscar) | Q(observaciones__icontains=buscar)
        )
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    # Promedio de puntajes
    promedio = qs.filter(
        puntaje__isnull=False
    ).aggregate(avg=Avg('puntaje'))['avg']

    paginator = Paginator(qs.order_by('-fecha'), 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'deportista/evaluaciones.html', {
        'evaluaciones':    page,
        'jugador':         jugador,
        'buscar':          buscar,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_filtrados': qs.count(),
        'promedio':        promedio,
    })


@solo_deportista
def dep_pagos(request):
    """Historial de pagos del deportista con filtros."""
    _marcar_vencidos()
    jugador = None
    qs      = Pago.objects.none()

    try:
        jugador = request.user.jugador
        qs = Pago.objects.filter(jugador=jugador)
    except Exception:
        pass

    # Filtros
    estado   = request.GET.get('estado', '')
    concepto = request.GET.get('concepto', '')

    if estado:
        qs = qs.filter(estado=estado)
    if concepto:
        qs = qs.filter(concepto=concepto)

    # Totales
    total_pagado   = qs.filter(estado='pagado').aggregate(
        t=Sum('monto'))['t'] or 0
    total_pendiente = qs.filter(estado='pendiente').aggregate(
        t=Sum('monto'))['t'] or 0
    total_vencido  = qs.filter(estado='vencido').aggregate(
        t=Sum('monto'))['t'] or 0

    paginator = Paginator(qs.order_by('-fecha_vencimiento'), 10)
    page = paginator.get_page(request.GET.get('page'))

    es_mayor = _deportista_es_mayor_de_edad(request.user)
    return render(request, 'deportista/pagos.html', {
        'pagos':            page,
        'jugador':          jugador,
        'estado':           estado,
        'concepto':         concepto,
        'estados':          Pago.ESTADO_CHOICES,
        'conceptos':        Pago.CONCEPTO_CHOICES,
        'total_pagado':     total_pagado,
        'total_pendiente':  total_pendiente,
        'total_vencido':    total_vencido,
        'total_filtrados':  qs.count(),
        'es_mayor':         es_mayor,
    })


@solo_deportista
def dep_pago_reportar(request, pago_id):
    """
    El deportista reporta que ya realizó el pago en la vida real.
    - Solo puede reportar pagos propios en estado pendiente o vencido.
    - Cambia el estado a 'en_revision' para que admin/entrenador confirme.
    - El deportista indica el método y puede agregar una nota.
    """
    try:
        jugador = request.user.jugador
    except Exception:
        messages.error(request, 'No tienes perfil de jugador.')
        return dep_pagos(request)

    pago = get_object_or_404(
        Pago,
        id=pago_id,
        jugador=jugador,
        estado__in=['pendiente', 'vencido']
    )

    if request.method == 'POST':
        metodo_pago = request.POST.get('metodo_pago', '').strip()
        fecha_pago  = request.POST.get('fecha_pago', '').strip() or timezone.now().date()

        if not metodo_pago:
            messages.error(request, 'Debes indicar el método de pago.')
            return render(request, 'deportista/pago_reportar.html', {
                'pago':    pago,
                'metodos': Pago.METODO_CHOICES,
            })

        pago.metodo_pago = metodo_pago
        pago.fecha_pago  = fecha_pago
        pago.estado      = 'en_revision'
        pago.save()
        messages.success(
            request,
            f'Pago reportado correctamente. '
            f'Un administrador o entrenador confirmará tu pago pronto.'
        )
        return dep_pagos(request)

    return render(request, 'deportista/pago_reportar.html', {
        'pago':    pago,
        'metodos': Pago.METODO_CHOICES,
    })


@solo_deportista
def dep_historial(request):
    """Historial propio del deportista: evaluaciones, observaciones y estadísticas de partido."""
    jugador = None
    evaluaciones  = Evaluacion.objects.none()
    observaciones = ObservacionDeportista.objects.none()
    stats         = EstadisticaPartido.objects.none()
    totales       = {}
    graf_labels = graf_tecnica = graf_tactica = graf_fisica = graf_todas = '[]'
    tiene_grafico = False

    try:
        jugador = request.user.jugador
        evaluaciones = Evaluacion.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')

        observaciones = ObservacionDeportista.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')

        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')

        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )

        # Datos para gráfico de evolución
        evs_con_puntaje = Evaluacion.objects.filter(
            jugador=jugador, puntaje__isnull=False
        ).order_by('fecha', 'id')

        graf_labels   = json.dumps([str(e.fecha) for e in evs_con_puntaje])
        graf_tecnica  = json.dumps([float(e.puntaje) if e.tipo == 'tecnica'  else None for e in evs_con_puntaje])
        graf_tactica  = json.dumps([float(e.puntaje) if e.tipo == 'tactica'  else None for e in evs_con_puntaje])
        graf_fisica   = json.dumps([float(e.puntaje) if e.tipo == 'fisica'   else None for e in evs_con_puntaje])
        graf_todas    = json.dumps([float(e.puntaje) for e in evs_con_puntaje])
        tiene_grafico = evs_con_puntaje.exists()
    except Exception:
        pass

    return render(request, 'deportista/historial.html', {
        'jugador':       jugador,
        'evaluaciones':  evaluaciones,
        'observaciones': observaciones,
        'stats':         stats,
        'totales':       totales,
        'graf_labels':   graf_labels,
        'graf_tecnica':  graf_tecnica,
        'graf_tactica':  graf_tactica,
        'graf_fisica':   graf_fisica,
        'graf_todas':    graf_todas,
        'tiene_grafico': tiene_grafico,
    })


@solo_deportista
def dep_factura_pago(request, pago_id):
    """El deportista descarga la factura PDF de uno de sus propios pagos."""
    try:
        jugador = request.user.jugador
    except Exception:
        from django.http import Http404
        raise Http404
    pago = get_object_or_404(
        Pago.objects.select_related('jugador__usuario', 'jugador__equipo__categoria', 'registrado_por'),
        id=pago_id,
        jugador=jugador
    )
    hoy  = timezone.now()
    html = render_to_string('panel/pagos/factura.html', {
        'pago':          pago,
        'fecha_emision': hoy,
        'emitido_por':   'Sistema Deportivo',
    })
    buffer = BytesIO()
    pisa.CreatePDF(html.encode('utf-8'), dest=buffer)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="factura_pago_{pago_id}.pdf"'
    response['Content-Length'] = len(pdf)
    return response


@solo_deportista
def dep_torneos(request):
    """Torneos de la categoría del deportista con filtros."""
    jugador   = None
    categoria = None
    qs        = Torneo.objects.none()

    try:
        jugador = request.user.jugador
        if jugador.equipo:
            categoria = jugador.equipo.categoria
            qs = Torneo.objects.filter(
                categoria=categoria
            ).annotate(num_partidos=Count('partidos'))
    except Exception:
        pass

    # Filtros
    estado = request.GET.get('estado', '')
    if estado:
        qs = qs.filter(estado=estado)

    return render(request, 'deportista/torneos.html', {
        'torneos':         qs.order_by('-fecha_inicio'),
        'jugador':         jugador,
        'categoria':       categoria,
        'estado':          estado,
        'estados':         Torneo.ESTADO_CHOICES,
        'total_filtrados': qs.count(),
    })


@solo_deportista
def dep_calendario(request):
    """Calendario de entrenamientos y partidos del equipo del deportista."""
    jugador = None
    entrenamientos = Entrenamiento.objects.none()
    partidos = PartidoCalendario.objects.none()

    try:
        jugador = request.user.jugador
        if jugador.equipo:
            entrenamientos = Entrenamiento.objects.filter(
                equipo=jugador.equipo
            ).select_related('entrenador').order_by('fecha', 'hora_inicio')
            partidos = PartidoCalendario.objects.filter(
                equipo_propio=jugador.equipo
            ).order_by('fecha', 'hora')
    except Exception:
        pass

    return render(request, 'deportista/calendario.html', {
        'jugador': jugador,
        'entrenamientos': entrenamientos,
        'partidos': partidos,
    })


@solo_deportista
def dep_convocatorias(request):
    """Convocatorias a partidos en las que el deportista ha sido incluido."""
    jugador = None
    convocatorias_data = []

    try:
        jugador = request.user.jugador
        convocatorias = Convocatoria.objects.filter(
            jugadores=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')
        respuestas_map = {
            r.convocatoria_id: r.respuesta
            for r in RespuestaConvocatoria.objects.filter(jugador=jugador)
        }
        for conv in convocatorias:
            convocatorias_data.append({
                'conv': conv,
                'respuesta': respuestas_map.get(conv.id),
            })
    except Exception:
        pass

    return render(request, 'deportista/convocatorias.html', {
        'jugador': jugador,
        'convocatorias_data': convocatorias_data,
    })


@solo_deportista
def dep_responder_convocatoria(request, conv_id):
    """Confirmar o rechazar participación en una convocatoria."""
    jugador = get_object_or_404(Jugador, usuario=request.user)
    convocatoria = get_object_or_404(Convocatoria, id=conv_id, jugadores=jugador)

    if request.method == 'POST':
        respuesta = request.POST.get('respuesta')
        if respuesta in ('confirmado', 'rechazado'):
            RespuestaConvocatoria.objects.update_or_create(
                convocatoria=convocatoria,
                jugador=jugador,
                defaults={'respuesta': respuesta},
            )
            labels = {'confirmado': 'confirmada', 'rechazado': 'rechazada'}
            messages.success(request, f'Participación {labels[respuesta]} correctamente.')
        else:
            messages.error(request, 'Respuesta no válida.')

    from django.shortcuts import redirect
    return redirect('/deportista/convocatorias/')


@solo_deportista
def dep_asistencia_historial(request):
    """Historial de asistencia a entrenamientos del deportista."""
    jugador = None
    asistencias = Asistencia.objects.none()
    resumen = {}

    try:
        jugador = request.user.jugador
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()
    except Exception:
        pass

    return render(request, 'deportista/asistencia_historial.html', {
        'jugador':    jugador,
        'asistencias': asistencias,
        'resumen':    resumen,
        'estados':    Asistencia.ESTADO_CHOICES,
    })


@solo_deportista
def dep_rendimiento(request):
    """Evaluaciones, observaciones y estadísticas de rendimiento del deportista."""
    jugador = None
    evaluaciones  = Evaluacion.objects.none()
    observaciones = ObservacionDeportista.objects.none()
    stats         = EstadisticaPartido.objects.none()
    totales       = {}
    graf_labels = graf_tecnica = graf_tactica = graf_fisica = graf_todas = '[]'
    tiene_grafico = False

    try:
        jugador = request.user.jugador
        evaluaciones = Evaluacion.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')

        observaciones = ObservacionDeportista.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')

        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')

        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )

        evs_con_puntaje = Evaluacion.objects.filter(
            jugador=jugador, puntaje__isnull=False
        ).order_by('fecha', 'id')
        graf_labels  = json.dumps([str(e.fecha) for e in evs_con_puntaje])
        graf_tecnica = json.dumps([float(e.puntaje) if e.tipo == 'tecnica' else None for e in evs_con_puntaje])
        graf_tactica = json.dumps([float(e.puntaje) if e.tipo == 'tactica' else None for e in evs_con_puntaje])
        graf_fisica  = json.dumps([float(e.puntaje) if e.tipo == 'fisica'  else None for e in evs_con_puntaje])
        graf_todas   = json.dumps([float(e.puntaje) for e in evs_con_puntaje])
        tiene_grafico = evs_con_puntaje.exists()
    except Exception:
        pass

    return render(request, 'deportista/rendimiento.html', {
        'jugador':       jugador,
        'evaluaciones':  evaluaciones,
        'observaciones': observaciones,
        'stats':         stats,
        'totales':       totales,
        'graf_labels':   graf_labels,
        'graf_tecnica':  graf_tecnica,
        'graf_tactica':  graf_tactica,
        'graf_fisica':   graf_fisica,
        'graf_todas':    graf_todas,
        'tiene_grafico': tiene_grafico,
    })


@solo_deportista
def dep_partidos_resultados(request):
    """Partidos en los que el deportista ha participado (tiene estadística registrada)."""
    jugador = None
    partidos_data = []

    try:
        jugador = request.user.jugador
        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha', '-partido__hora')
        for s in stats:
            partidos_data.append({
                'partido': s.partido,
                'stat':    s,
            })
    except Exception:
        pass

    return render(request, 'deportista/partidos_resultados.html', {
        'jugador':       jugador,
        'partidos_data': partidos_data,
    })


@solo_deportista
def dep_estadisticas(request):
    """Estadísticas deportivas del deportista: goles, asistencias, rendimiento y participación."""
    jugador = None
    stats   = EstadisticaPartido.objects.none()
    totales = {}
    total_partidos = total_convocado = 0
    pct_participacion = 0

    try:
        jugador = request.user.jugador
        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')

        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )
        total_partidos    = stats.count()
        total_convocado   = jugador.convocatorias.count()
        pct_participacion = round((total_partidos / total_convocado) * 100) if total_convocado else 0
    except Exception:
        pass

    return render(request, 'deportista/estadisticas.html', {
        'jugador':           jugador,
        'stats':             stats,
        'totales':           totales,
        'total_partidos':    total_partidos,
        'total_convocado':   total_convocado,
        'pct_participacion': pct_participacion,
    })


@login_required
def marcar_notificacion_leida(request, noti_id):
    noti = get_object_or_404(Notificacion, id=noti_id, usuario=request.user)
    noti.leida = True
    noti.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))


@solo_entrenador
def ent_mensajes(request):

    qs = Notificacion.objects.filter(usuario=request.user).order_by('-creado')

    tipo        = request.GET.get('tipo', '')
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    TIPOS_NOTIF = [
        ('entrenamiento', 'Entrenamiento'),
        ('partido',       'Partido'),
        ('convocatoria',  'Convocatoria'),
        ('pago',          'Pago'),
        ('mensaje',       'Mensaje'),
    ]
    if tipo:        qs = qs.filter(asunto__icontains=tipo)
    if leida == '1': qs = qs.filter(leida=True)
    elif leida == '0': qs = qs.filter(leida=False)
    if fecha_desde: qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta: qs = qs.filter(creado__date__lte=fecha_hasta)

    no_leidos = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'entrenador/mensajes.html', {
        'mensajes': page, 'no_leidos': no_leidos,
        'tipo': tipo, 'leida': leida,
        'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta,
        'tipos_notif': TIPOS_NOTIF,
    })


@solo_entrenador
def ent_mensajes_leer_todos(request):
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    return HttpResponseRedirect('/entrenador/mensajes/')


@solo_admin
def admin_mensajes_enviados(request):
    """Lista de mensajes enviados por el administrador con filtros y totales."""
    qs = Notificacion.objects.filter(emisor=request.user).select_related('usuario').order_by('-creado')

    buscar      = request.GET.get('buscar', '').strip()
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(asunto__icontains=buscar) |
            Q(usuario__first_name__icontains=buscar) |
            Q(usuario__last_name__icontains=buscar)
        )
    if leida == '1':
        qs = qs.filter(leida=True)
    elif leida == '0':
        qs = qs.filter(leida=False)
    if fecha_desde:
        qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(creado__date__lte=fecha_hasta)

    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'panel/correos/enviados.html', {
        'mensajes':        page,
        'buscar':          buscar,
        'leida':           leida,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_enviados':  qs.count(),
        'total_leidos':    qs.filter(leida=True).count(),
        'total_no_leidos': qs.filter(leida=False).count(),
    })


@solo_admin
def admin_mensaje_eliminar(request, noti_id):
    """Eliminar un mensaje enviado por el administrador."""
    noti = get_object_or_404(Notificacion, id=noti_id, emisor=request.user)
    noti.delete()
    messages.success(request, 'Mensaje eliminado correctamente.')
    return HttpResponseRedirect('/panel/correos/enviados/')


@solo_entrenador
def ent_mensajes_enviados(request):
    """Lista de mensajes enviados por el entrenador con filtros y totales."""
    qs = Notificacion.objects.filter(emisor=request.user).select_related('usuario').order_by('-creado')

    buscar      = request.GET.get('buscar', '').strip()
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if buscar:
        qs = qs.filter(
            Q(asunto__icontains=buscar) |
            Q(usuario__first_name__icontains=buscar) |
            Q(usuario__last_name__icontains=buscar)
        )
    if leida == '1':
        qs = qs.filter(leida=True)
    elif leida == '0':
        qs = qs.filter(leida=False)
    if fecha_desde:
        qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(creado__date__lte=fecha_hasta)

    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'entrenador/mensajes_enviados.html', {
        'mensajes':        page,
        'buscar':          buscar,
        'leida':           leida,
        'fecha_desde':     fecha_desde,
        'fecha_hasta':     fecha_hasta,
        'total_enviados':  qs.count(),
        'total_leidos':    qs.filter(leida=True).count(),
        'total_no_leidos': qs.filter(leida=False).count(),
    })


@solo_entrenador
def ent_mensaje_eliminar(request, noti_id):
    """Eliminar un mensaje enviado por el entrenador."""
    noti = get_object_or_404(Notificacion, id=noti_id, emisor=request.user)
    noti.delete()
    messages.success(request, 'Mensaje eliminado correctamente.')
    return HttpResponseRedirect('/entrenador/mensajes/enviados/')


@solo_deportista
def dep_mensajes(request):

    qs = Notificacion.objects.filter(usuario=request.user).order_by('-creado')

    tipo        = request.GET.get('tipo', '')
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    TIPOS_NOTIF = [
        ('entrenamiento', 'Entrenamiento'),
        ('partido',       'Partido'),
        ('convocatoria',  'Convocatoria'),
        ('pago',          'Pago'),
        ('mensaje',       'Mensaje'),
    ]
    if tipo:        qs = qs.filter(asunto__icontains=tipo)
    if leida == '1': qs = qs.filter(leida=True)
    elif leida == '0': qs = qs.filter(leida=False)
    if fecha_desde: qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta: qs = qs.filter(creado__date__lte=fecha_hasta)

    no_leidos = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'deportista/mensajes.html', {

        'mensajes': page, 'no_leidos': no_leidos,
        'tipo': tipo, 'leida': leida,
        'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta,
        'tipos_notif': TIPOS_NOTIF,
    })


@solo_admin
def admin_notificaciones(request):
    """Lista de notificaciones recibidas por el administrador con filtros."""
    qs = Notificacion.objects.filter(usuario=request.user).order_by('-creado')

    tipo        = request.GET.get('tipo', '')
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    TIPOS_NOTIF = [
        ('entrenamiento', 'Entrenamiento'),
        ('partido',       'Partido'),
        ('convocatoria',  'Convocatoria'),
        ('pago',          'Pago'),
        ('mensaje',       'Mensaje'),
    ]
    if tipo:         qs = qs.filter(asunto__icontains=tipo)
    if leida == '1': qs = qs.filter(leida=True)
    elif leida == '0': qs = qs.filter(leida=False)
    if fecha_desde:  qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta:  qs = qs.filter(creado__date__lte=fecha_hasta)

    no_leidos = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'panel/notificaciones.html', {
        'mensajes': page, 'no_leidos': no_leidos,
        'tipo': tipo, 'leida': leida,
        'fecha_desde': fecha_desde, 'fecha_hasta': fecha_hasta,
        'tipos_notif': TIPOS_NOTIF,
    })


@solo_admin
def admin_notificaciones_leer_todos(request):
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    return HttpResponseRedirect('/panel/notificaciones/')


@solo_admin
def admin_notificacion_eliminar(request, noti_id):
    """Eliminar una notificacion propia del administrador."""
    noti = get_object_or_404(Notificacion, id=noti_id, usuario=request.user)
    noti.delete()
    messages.success(request, 'Notificacion eliminada.')
    return HttpResponseRedirect('/panel/notificaciones/')


@solo_entrenador
def ent_notificacion_eliminar(request, noti_id):
    """Eliminar una notificacion propia del entrenador."""
    noti = get_object_or_404(Notificacion, id=noti_id, usuario=request.user)
    noti.delete()
    messages.success(request, 'Notificacion eliminada.')
    return HttpResponseRedirect('/entrenador/mensajes/')


@solo_deportista
def dep_notificacion_eliminar(request, noti_id):
    """Eliminar una notificacion propia del deportista."""
    noti = get_object_or_404(Notificacion, id=noti_id, usuario=request.user)
    noti.delete()
    messages.success(request, 'Notificacion eliminada.')
    return HttpResponseRedirect('/deportista/mensajes/')


@solo_deportista
def dep_mensajes_leer_todos(request):
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    return HttpResponseRedirect('/deportista/mensajes/')


# ─────────────────────────────────────────────
# CARGA MASIVA (ADMIN)
# ─────────────────────────────────────────────

@solo_admin
def carga_masiva(request):
    secciones = [
        ('usuarios',   'Usuarios',   'primary', 'fas fa-users',        '/panel/carga-masiva/usuarios/'),
        ('categorias', 'Categorías', 'success', 'fas fa-tags',         '/panel/carga-masiva/categorias/'),
        ('equipos',    'Equipos',    'warning',  'fas fa-shield-alt',   '/panel/carga-masiva/equipos/'),
        ('pagos',      'Pagos',      'danger',  'fas fa-money-bill-wave', '/panel/carga-masiva/pagos/'),
    ]
    return render(request, 'panel/carga_masiva/index.html', {'secciones': secciones})


@solo_admin
def carga_plantilla(request, modulo):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    plantillas = {
        'usuarios':    ['email', 'first_name', 'last_name', 'documento', 'phone', 'role', 'password', 'equipo_nombre', 'posicion'],
        'categorias':  ['nombre', 'descripcion'],
        'equipos':     ['nombre', 'categoria_nombre', 'entrenador_email'],
        'pagos':       ['deportista_documento', 'concepto', 'descripcion', 'monto', 'fecha_vencimiento'],
    }
    if modulo not in plantillas:
        raise Http404

    ws.append(plantillas[modulo])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="plantilla_{modulo}.xlsx"'
    wb.save(response)
    return response


@solo_admin
def carga_usuarios(request):
    if request.method != 'POST' or 'archivo' not in request.FILES:
        return HttpResponseRedirect('/panel/carga-masiva/')

    import openpyxl
    archivo = request.FILES['archivo']
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    creados = errores = 0
    filas_error = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            email, first_name, last_name, documento, phone, role, password, equipo_nombre, posicion = [
                str(c).strip() if c is not None else '' for c in (row + (None,) * 9)[:9]
            ]
            if not email or not documento or not password:
                raise ValueError('email, documento y password son obligatorios')
            if UserModel.objects.filter(email__iexact=email).exists():
                raise ValueError(f'email {email} ya existe')
            if UserModel.objects.filter(documento=documento).exists():
                raise ValueError(f'documento {documento} ya existe')

            role = role if role in ['administrador', 'entrenador', 'deportista'] else 'deportista'
            prefijos = {'administrador': 'admin', 'entrenador': 'entrenador', 'deportista': 'deportista'}
            prefijo = prefijos[role]
            count = UserModel.objects.filter(role=role).count() + 1
            username = f'{prefijo}{count:03d}'
            while UserModel.objects.filter(username=username).exists():
                count += 1
                username = f'{prefijo}{count:03d}'

            usuario = UserModel.objects.create_user(
                username=username, email=email, password=password,
                first_name=first_name, last_name=last_name,
                role=role, phone=phone, documento=documento, is_active=True,
            )
            if role == 'deportista' and equipo_nombre:
                equipo = Equipo.objects.filter(nombre__iexact=equipo_nombre, activo=True).first()
                Jugador.objects.update_or_create(
                    usuario=usuario,
                    defaults={'equipo': equipo, 'posicion': posicion, 'activo': True}
                )
            creados += 1
        except Exception as e:
            errores += 1
            filas_error.append(f'Fila {i}: {e}')

    if creados:
        messages.success(request, f'{creados} usuario(s) creado(s) correctamente.')
    if errores:
        messages.warning(request, f'{errores} fila(s) con error: ' + ' | '.join(filas_error[:5]))
    return HttpResponseRedirect('/panel/carga-masiva/')


@solo_admin
def carga_categorias(request):
    if request.method != 'POST' or 'archivo' not in request.FILES:
        return HttpResponseRedirect('/panel/carga-masiva/')

    import openpyxl
    wb = openpyxl.load_workbook(request.FILES['archivo'])
    ws = wb.active
    creados = errores = 0
    filas_error = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            nombre = str(row[0]).strip() if row[0] else ''
            descripcion = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not nombre:
                raise ValueError('nombre es obligatorio')
            if Categoria.objects.filter(nombre__iexact=nombre).exists():
                raise ValueError(f'"{nombre}" ya existe')
            Categoria.objects.create(nombre=nombre, descripcion=descripcion)
            creados += 1
        except Exception as e:
            errores += 1
            filas_error.append(f'Fila {i}: {e}')

    if creados:
        messages.success(request, f'{creados} categoría(s) creada(s) correctamente.')
    if errores:
        messages.warning(request, f'{errores} fila(s) con error: ' + ' | '.join(filas_error[:5]))
    return HttpResponseRedirect('/panel/carga-masiva/')


@solo_admin
def carga_equipos(request):
    if request.method != 'POST' or 'archivo' not in request.FILES:
        return HttpResponseRedirect('/panel/carga-masiva/')

    import openpyxl
    wb = openpyxl.load_workbook(request.FILES['archivo'])
    ws = wb.active
    creados = errores = 0
    filas_error = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            nombre = str(row[0]).strip() if row[0] else ''
            cat_nombre = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            ent_email = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not nombre or not cat_nombre:
                raise ValueError('nombre y categoria_nombre son obligatorios')
            categoria = Categoria.objects.filter(nombre__iexact=cat_nombre, activo=True).first()
            if not categoria:
                raise ValueError(f'Categoría "{cat_nombre}" no existe')
            entrenador = None
            if ent_email:
                entrenador = UserModel.objects.filter(email__iexact=ent_email, role='entrenador').first()
            Equipo.objects.create(nombre=nombre, categoria=categoria, entrenador=entrenador)
            creados += 1
        except Exception as e:
            errores += 1
            filas_error.append(f'Fila {i}: {e}')

    if creados:
        messages.success(request, f'{creados} equipo(s) creado(s) correctamente.')
    if errores:
        messages.warning(request, f'{errores} fila(s) con error: ' + ' | '.join(filas_error[:5]))
    return HttpResponseRedirect('/panel/carga-masiva/')


@solo_admin
def carga_pagos(request):
    if request.method != 'POST' or 'archivo' not in request.FILES:
        return HttpResponseRedirect('/panel/carga-masiva/')

    import openpyxl
    wb = openpyxl.load_workbook(request.FILES['archivo'])
    ws = wb.active
    creados = errores = 0
    filas_error = []
    conceptos_validos = [v for v, _ in Pago.CONCEPTO_CHOICES]

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            documento  = str(row[0]).strip() if row[0] else ''
            concepto   = str(row[1]).strip() if len(row) > 1 and row[1] else 'cuota_mensual'
            descripcion = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            monto      = row[3] if len(row) > 3 else None
            fecha_venc = row[4] if len(row) > 4 else None

            if not documento or not descripcion or not monto or not fecha_venc:
                raise ValueError('documento, descripcion, monto y fecha_vencimiento son obligatorios')

            usuario = UserModel.objects.filter(documento=documento, role='deportista').first()
            if not usuario:
                raise ValueError(f'Deportista con documento {documento} no encontrado')
            jugador = Jugador.objects.filter(usuario=usuario).first()
            if not jugador:
                raise ValueError(f'El deportista {documento} no tiene perfil de jugador')

            concepto = concepto if concepto in conceptos_validos else 'otro'
            monto_dec = Decimal(str(monto))
            if monto_dec <= 0:
                raise ValueError('monto debe ser mayor a 0')

            if hasattr(fecha_venc, 'date'):
                fecha_venc = fecha_venc.date()

            Pago.objects.create(
                jugador=jugador,
                concepto=concepto,
                descripcion=descripcion,
                monto=monto_dec,
                fecha_vencimiento=fecha_venc,
                estado='pendiente',
                registrado_por=request.user,
            )
            creados += 1
        except Exception as e:
            errores += 1
            filas_error.append(f'Fila {i}: {e}')

    if creados:
        messages.success(request, f'{creados} pago(s) creado(s) correctamente.')
    if errores:
        messages.warning(request, f'{errores} fila(s) con error: ' + ' | '.join(filas_error[:5]))
    return HttpResponseRedirect('/panel/carga-masiva/')


# ═══════════════════════════════════════════════════════════════════════════════
# ACUDIENTE
# ═══════════════════════════════════════════════════════════════════════════════

def solo_acudiente(view_func):
    """Decorador que verifica que el usuario sea acudiente"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return render(request, 'auth/login.html')
        if request.user.role != 'acudiente':
            return HttpResponseForbidden('Sin permisos de acudiente.')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def acudiente_dashboard(request):
    """Dashboard del acudiente con resumen del deportista a cargo."""
    if request.user.role != 'acudiente':
        return HttpResponseForbidden('Sin permisos.')

    acudiente = jugador = equipo = categoria = None
    entrenamientos = []
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        if jugador.equipo:
            equipo = jugador.equipo
            categoria = equipo.categoria
            entrenamientos = Entrenamiento.objects.filter(equipo=equipo).order_by('fecha')[:5]
    except Exception:
        pass

    return render(request, 'acudiente/dashboard.html', {
        'acudiente': acudiente,
        'jugador': jugador,
        'equipo': equipo,
        'categoria': categoria,
        'entrenamientos': entrenamientos,
        'notificaciones': Notificacion.objects.filter(usuario=request.user, leida=False).order_by('-creado'),
    })


@solo_acudiente
def acu_deportista(request):
    """Información del deportista a cargo del acudiente."""
    acudiente = jugador = None
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
    except Exception:
        pass
    return render(request, 'acudiente/deportista.html', {
        'acudiente': acudiente,
        'jugador': jugador,
    })


@solo_acudiente
def acu_calendario(request):
    """Calendario de entrenamientos y partidos del deportista a cargo."""
    jugador = None
    entrenamientos = Entrenamiento.objects.none()
    partidos = PartidoCalendario.objects.none()
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        if jugador.equipo:
            entrenamientos = Entrenamiento.objects.filter(
                equipo=jugador.equipo
            ).select_related('entrenador').order_by('fecha', 'hora_inicio')
            partidos = PartidoCalendario.objects.filter(
                equipo_propio=jugador.equipo
            ).order_by('fecha', 'hora')
    except Exception:
        pass
    return render(request, 'acudiente/calendario.html', {
        'jugador': jugador,
        'entrenamientos': entrenamientos,
        'partidos': partidos,
    })


@solo_acudiente
def acu_convocatorias(request):
    """Convocatorias a partidos del deportista a cargo."""
    jugador = None
    convocatorias_data = []
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        convocatorias = Convocatoria.objects.filter(
            jugadores=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')
        respuestas_map = {
            r.convocatoria_id: r.respuesta
            for r in RespuestaConvocatoria.objects.filter(jugador=jugador)
        }
        for conv in convocatorias:
            convocatorias_data.append({'conv': conv, 'respuesta': respuestas_map.get(conv.id)})
    except Exception:
        pass
    return render(request, 'acudiente/convocatorias.html', {
        'jugador': jugador,
        'convocatorias_data': convocatorias_data,
    })


@solo_acudiente
def acu_responder_convocatoria(request, conv_id):
    """El acudiente confirma o rechaza la participación del deportista en una convocatoria."""
    try:
        jugador = request.user.acudiente.jugador
    except Exception:
        return HttpResponseForbidden('Sin deportista vinculado.')
    convocatoria = get_object_or_404(Convocatoria, id=conv_id, jugadores=jugador)
    if request.method == 'POST':
        respuesta = request.POST.get('respuesta')
        if respuesta in ('confirmado', 'rechazado'):
            RespuestaConvocatoria.objects.update_or_create(
                convocatoria=convocatoria,
                jugador=jugador,
                defaults={'respuesta': respuesta},
            )
            labels = {'confirmado': 'confirmada', 'rechazado': 'rechazada'}
            messages.success(request, f'Participación {labels[respuesta]} correctamente.')
        else:
            messages.error(request, 'Respuesta no válida.')
    from django.shortcuts import redirect
    return redirect('/acudiente/convocatorias/')


@solo_acudiente
def acu_asistencia(request):
    """Historial de asistencia a entrenamientos del deportista a cargo."""
    jugador = None
    asistencias = Asistencia.objects.none()
    resumen = {}
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        asistencias = Asistencia.objects.filter(
            jugador=jugador
        ).select_related('entrenamiento__equipo').order_by('-entrenamiento__fecha')
        for estado, _ in Asistencia.ESTADO_CHOICES:
            resumen[estado] = asistencias.filter(estado=estado).count()
    except Exception:
        pass
    return render(request, 'acudiente/asistencia.html', {
        'jugador': jugador,
        'asistencias': asistencias,
        'resumen': resumen,
    })


@solo_acudiente
def acu_rendimiento(request):
    """Evaluaciones, observaciones y estadísticas de rendimiento del deportista a cargo."""
    jugador = None
    evaluaciones  = Evaluacion.objects.none()
    observaciones = ObservacionDeportista.objects.none()
    stats         = EstadisticaPartido.objects.none()
    totales       = {}
    graf_labels = graf_tecnica = graf_tactica = graf_fisica = graf_todas = '[]'
    tiene_grafico = False
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        evaluaciones = Evaluacion.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')
        observaciones = ObservacionDeportista.objects.filter(
            jugador=jugador
        ).select_related('entrenador').order_by('-fecha')
        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha')
        totales = stats.aggregate(
            total_goles=Sum('goles'),
            total_asistencias=Sum('asistencias'),
            total_amarillas=Count('id', filter=Q(tarjeta_amarilla=True)),
            total_rojas=Count('id', filter=Q(tarjeta_roja=True)),
            promedio_calif=Avg('calificacion'),
        )
        evs_con_puntaje = Evaluacion.objects.filter(
            jugador=jugador, puntaje__isnull=False
        ).order_by('fecha', 'id')
        graf_labels  = json.dumps([str(e.fecha) for e in evs_con_puntaje])
        graf_tecnica = json.dumps([float(e.puntaje) if e.tipo == 'tecnica' else None for e in evs_con_puntaje])
        graf_tactica = json.dumps([float(e.puntaje) if e.tipo == 'tactica' else None for e in evs_con_puntaje])
        graf_fisica  = json.dumps([float(e.puntaje) if e.tipo == 'fisica'  else None for e in evs_con_puntaje])
        graf_todas   = json.dumps([float(e.puntaje) for e in evs_con_puntaje])
        tiene_grafico = evs_con_puntaje.exists()
    except Exception:
        pass
    return render(request, 'acudiente/rendimiento.html', {
        'jugador':       jugador,
        'evaluaciones':  evaluaciones,
        'observaciones': observaciones,
        'stats':         stats,
        'totales':       totales,
        'graf_labels':   graf_labels,
        'graf_tecnica':  graf_tecnica,
        'graf_tactica':  graf_tactica,
        'graf_fisica':   graf_fisica,
        'graf_todas':    graf_todas,
        'tiene_grafico': tiene_grafico,
    })


@solo_acudiente
def acu_partidos_resultados(request):
    """Partidos en los que ha participado el deportista a cargo."""
    jugador = None
    partidos_data = []
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        stats = EstadisticaPartido.objects.filter(
            jugador=jugador
        ).select_related('partido__equipo_propio').order_by('-partido__fecha', '-partido__hora')
        for s in stats:
            partidos_data.append({'partido': s.partido, 'stat': s})
    except Exception:
        pass
    return render(request, 'acudiente/partidos_resultados.html', {
        'jugador':       jugador,
        'partidos_data': partidos_data,
    })


@solo_acudiente
def acu_pagos(request):
    """Historial de pagos, facturas y saldos pendientes del deportista a cargo."""
    _marcar_vencidos()
    jugador = None
    qs = Pago.objects.none()
    try:
        acudiente = request.user.acudiente
        jugador = acudiente.jugador
        qs = Pago.objects.filter(jugador=jugador)
    except Exception:
        pass

    estado   = request.GET.get('estado', '')
    concepto = request.GET.get('concepto', '')
    if estado:   qs = qs.filter(estado=estado)
    if concepto: qs = qs.filter(concepto=concepto)

    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or 0
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or 0
    total_vencido   = qs.filter(estado='vencido').aggregate(t=Sum('monto'))['t'] or 0

    paginator = Paginator(qs.order_by('-fecha_vencimiento'), 10)
    page = paginator.get_page(request.GET.get('page'))
    deportista_es_mayor = jugador and _deportista_es_mayor_de_edad(jugador.usuario)
    return render(request, 'acudiente/pagos.html', {
        'pagos':               page,
        'jugador':             jugador,
        'estado':              estado,
        'concepto':            concepto,
        'estados':             Pago.ESTADO_CHOICES,
        'conceptos':           Pago.CONCEPTO_CHOICES,
        'total_pagado':        total_pagado,
        'total_pendiente':     total_pendiente,
        'total_vencido':       total_vencido,
        'total_filtrados':     qs.count(),
        'deportista_es_mayor': deportista_es_mayor,
    })


@solo_acudiente
def acu_mensajes(request):
    """Notificaciones y mensajes recibidos por el acudiente."""
    qs = Notificacion.objects.filter(usuario=request.user).order_by('-creado')

    tipo        = request.GET.get('tipo', '')
    leida       = request.GET.get('leida', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    TIPOS_NOTIF = [
        ('entrenamiento', 'Entrenamiento'),
        ('partido',       'Partido'),
        ('convocatoria',  'Convocatoria'),
        ('pago',          'Pago'),
        ('mensaje',       'Mensaje'),
    ]
    if tipo:             qs = qs.filter(asunto__icontains=tipo)
    if leida == '1':     qs = qs.filter(leida=True)
    elif leida == '0':   qs = qs.filter(leida=False)
    if fecha_desde:      qs = qs.filter(creado__date__gte=fecha_desde)
    if fecha_hasta:      qs = qs.filter(creado__date__lte=fecha_hasta)

    no_leidos = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'acudiente/mensajes.html', {
        'mensajes':    page,
        'no_leidos':   no_leidos,
        'tipo':        tipo,
        'leida':       leida,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipos_notif': TIPOS_NOTIF,
    })


@solo_acudiente
def acu_mensajes_leer_todos(request):
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    return HttpResponseRedirect('/acudiente/mensajes/')


@solo_acudiente
def acu_notificacion_eliminar(request, noti_id):
    noti = get_object_or_404(Notificacion, id=noti_id, usuario=request.user)
    noti.delete()
    messages.success(request, 'Notificación eliminada.')
    return HttpResponseRedirect('/acudiente/mensajes/')


# ─────────────────────────────────────────────
# NÓMINA DE ENTRENADORES (ADMIN)
# ─────────────────────────────────────────────

def _crear_finanza_por_nomina(nomina, registrado_por):
    """Crea un egreso en Finanza por el pago de nómina a un entrenador si no existe ya."""
    descripcion_clave = f'Nómina: {nomina.entrenador.get_full_name() or nomina.entrenador.username} — {nomina.mes:02d}/{nomina.anio}'
    if not Finanza.objects.filter(descripcion=descripcion_clave).exists():
        Finanza.objects.create(
            tipo='egreso',
            categoria='otro',
            descripcion=descripcion_clave,
            monto=nomina.monto,
            fecha=nomina.fecha_pago or timezone.now().date(),
            registrado_por=registrado_por,
        )


@solo_admin
def admin_nomina_lista(request):
    """Lista de pagos de nómina a entrenadores con filtros."""
    qs = PagoEntrenador.objects.select_related('entrenador', 'registrado_por')

    entrenador_id = request.GET.get('entrenador', '')
    estado        = request.GET.get('estado', '')
    anio          = request.GET.get('anio', '')

    if entrenador_id:
        qs = qs.filter(entrenador_id=entrenador_id)
    if estado:
        qs = qs.filter(estado=estado)
    if anio:
        qs = qs.filter(anio=anio)

    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')

    paginator = Paginator(qs.order_by('-anio', '-mes'), 15)
    page = paginator.get_page(request.GET.get('page'))

    entrenadores = UserModel.objects.filter(role='entrenador', is_active=True).order_by('last_name', 'first_name')

    return render(request, 'panel/pagos/nomina_lista.html', {
        'nominas':         page,
        'entrenador_id':   entrenador_id,
        'estado':          estado,
        'anio':            anio,
        'entrenadores':    entrenadores,
        'estados':         PagoEntrenador.ESTADO_CHOICES,
        'total_pagado':    total_pagado,
        'total_pendiente': total_pendiente,
        'total_filtrados': qs.count(),
    })


@solo_admin
def admin_nomina_crear(request):
    """Registrar pago de nómina mensual a un entrenador."""
    entrenadores = UserModel.objects.filter(role='entrenador', is_active=True).order_by('last_name', 'first_name')

    if request.method == 'POST':
        entrenador_id = request.POST.get('entrenador')
        mes           = request.POST.get('mes', '').strip()
        anio          = request.POST.get('anio', '').strip()
        monto         = request.POST.get('monto', '').strip()
        estado        = request.POST.get('estado', 'pendiente')
        metodo_pago   = request.POST.get('metodo_pago', '')
        descripcion   = request.POST.get('descripcion', '').strip()
        fecha_pago    = request.POST.get('fecha_pago') or None

        errores = []
        if not entrenador_id:
            errores.append('Debes seleccionar un entrenador.')
        if not mes or not mes.isdigit() or not (1 <= int(mes) <= 12):
            errores.append('El mes debe ser un número entre 1 y 12.')
        if not anio or not anio.isdigit():
            errores.append('El año es obligatorio.')
        try:
            monto_dec = Decimal(monto)
            if monto_dec <= 0:
                errores.append('El monto debe ser mayor a 0.')
        except Exception:
            errores.append('El monto ingresado no es válido.')

        if not errores and PagoEntrenador.objects.filter(
            entrenador_id=entrenador_id, mes=int(mes), anio=int(anio)
        ).exists():
            errores.append('Ya existe un pago de nómina para ese entrenador en ese mes/año.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            nomina = PagoEntrenador.objects.create(
                entrenador_id=entrenador_id,
                mes=int(mes),
                anio=int(anio),
                monto=monto_dec,
                estado=estado,
                metodo_pago=metodo_pago,
                descripcion=descripcion,
                fecha_pago=fecha_pago,
                registrado_por=request.user,
            )
            if estado == 'pagado':
                _crear_finanza_por_nomina(nomina, request.user)
            # Notificar al entrenador
            Notificacion.objects.create(
                usuario=nomina.entrenador,
                asunto=f'Pago de nómina {nomina.mes:02d}/{nomina.anio}',
                mensaje=(
                    f'Se ha registrado tu pago de nómina.\n\n'
                    f'Mes/Año:  {nomina.mes:02d}/{nomina.anio}\n'
                    f'Monto:    {nomina.monto}\n'
                    f'Estado:   {nomina.get_estado_display()}'
                    + (f'\nFecha:    {nomina.fecha_pago.strftime("%d/%m/%Y")}' if nomina.fecha_pago else '')
                ),
                emisor=request.user,
            )
            messages.success(request, 'Pago de nómina registrado correctamente.')
            return admin_nomina_lista(request)

    return render(request, 'panel/pagos/nomina_form.html', {
        'titulo':      'Registrar Nómina',
        'entrenadores': entrenadores,
        'estados':     PagoEntrenador.ESTADO_CHOICES,
        'metodos':     PagoEntrenador.METODO_CHOICES,
    })


@solo_admin
def admin_nomina_editar(request, nomina_id):
    """Editar pago de nómina existente."""
    nomina       = get_object_or_404(PagoEntrenador, id=nomina_id)
    entrenadores = UserModel.objects.filter(role='entrenador', is_active=True).order_by('last_name', 'first_name')

    if request.method == 'POST':
        monto       = request.POST.get('monto', '').strip()
        estado      = request.POST.get('estado', nomina.estado)
        metodo_pago = request.POST.get('metodo_pago', '')
        descripcion = request.POST.get('descripcion', '').strip()
        fecha_pago  = request.POST.get('fecha_pago') or None

        errores = []
        try:
            monto_dec = Decimal(monto)
            if monto_dec <= 0:
                errores.append('El monto debe ser mayor a 0.')
        except Exception:
            errores.append('El monto ingresado no es válido.')

        if errores:
            for e in errores:
                messages.error(request, e)
        else:
            estado_anterior = nomina.estado
            nomina.monto       = monto_dec
            nomina.estado      = estado
            nomina.metodo_pago = metodo_pago
            nomina.descripcion = descripcion
            nomina.fecha_pago  = fecha_pago
            nomina.save()
            if estado == 'pagado' and estado_anterior != 'pagado':
                _crear_finanza_por_nomina(nomina, request.user)
            messages.success(request, 'Nómina actualizada correctamente.')
            return admin_nomina_lista(request)

    return render(request, 'panel/pagos/nomina_form.html', {
        'titulo':      'Editar Nómina',
        'nomina':      nomina,
        'entrenadores': entrenadores,
        'estados':     PagoEntrenador.ESTADO_CHOICES,
        'metodos':     PagoEntrenador.METODO_CHOICES,
    })


@solo_admin
def admin_nomina_eliminar(request, nomina_id):
    """Eliminar pago de nómina."""
    nomina = get_object_or_404(PagoEntrenador, id=nomina_id)
    nomina.delete()
    messages.success(request, 'Pago de nómina eliminado correctamente.')
    return admin_nomina_lista(request)


@solo_admin
def admin_nomina_marcar_pagado(request, nomina_id):
    """Acción rápida: marcar nómina como PAGADA y registrar egreso en finanzas."""
    nomina = get_object_or_404(PagoEntrenador, id=nomina_id)
    if nomina.estado != 'pagado':
        nomina.estado = 'pagado'
        if not nomina.fecha_pago:
            nomina.fecha_pago = timezone.now().date()
        nomina.save()
        _crear_finanza_por_nomina(nomina, request.user)
        Notificacion.objects.create(
            usuario=nomina.entrenador,
            asunto=f'Nómina pagada {nomina.mes:02d}/{nomina.anio}',
            mensaje=(
                f'Tu nómina ha sido pagada.\n\n'
                f'Mes/Año:  {nomina.mes:02d}/{nomina.anio}\n'
                f'Monto:    {nomina.monto}\n'
                f'Fecha:    {nomina.fecha_pago.strftime("%d/%m/%Y")}'
            ),
            emisor=request.user,
        )
        messages.success(request, f'Nómina de {nomina.entrenador.get_full_name() or nomina.entrenador.username} marcada como pagada.')
    return admin_nomina_lista(request)


# ─────────────────────────────────────────────
# NÓMINA — HISTORIAL DEL ENTRENADOR
# ─────────────────────────────────────────────

@solo_entrenador
def ent_nomina_lista(request):
    """Historial de pagos de nómina recibidos por el entrenador (solo lectura)."""
    qs = PagoEntrenador.objects.filter(
        entrenador=request.user
    ).select_related('registrado_por').order_by('-anio', '-mes')

    estado = request.GET.get('estado', '')
    anio   = request.GET.get('anio', '')

    if estado:
        qs = qs.filter(estado=estado)
    if anio:
        qs = qs.filter(anio=anio)

    total_pagado    = qs.filter(estado='pagado').aggregate(t=Sum('monto'))['t'] or Decimal('0')
    total_pendiente = qs.filter(estado='pendiente').aggregate(t=Sum('monto'))['t'] or Decimal('0')

    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'entrenador/pagos/nomina_lista.html', {
        'nominas':         page,
        'estado':          estado,
        'anio':            anio,
        'estados':         PagoEntrenador.ESTADO_CHOICES,
        'total_pagado':    total_pagado,
        'total_pendiente': total_pendiente,
        'total_filtrados': qs.count(),
    })


# ─────────────────────────────────────────────
# VALIDACIÓN DE MAYORÍA DE EDAD EN PAGOS
# ─────────────────────────────────────────────

def _deportista_es_mayor_de_edad(usuario):
    """Retorna True si el deportista tiene 18 años o más."""
    if not usuario.birth_date:
        return False
    hoy = timezone.now().date()
    edad = hoy.year - usuario.birth_date.year - (
        (hoy.month, hoy.day) < (usuario.birth_date.month, usuario.birth_date.day)
    )
    return edad >= 18
